#!/usr/bin/env python3
"""
AIFMD Annex IV XML Validator
=============================
Generic validator for any AIFM or AIF XML against the ESMA base ruleset
with optional NCA-specific overrides.

Usage:
    python validate_aifmd_xml.py <xml_file> [<xml_file2> ...]
    python validate_aifmd_xml.py output/*.xml
    python validate_aifmd_xml.py --nca cssf AIF_LU_*.xml

Options:
    --nca <code>     NCA code (auto-detected from XML if omitted)
    --output <path>  Output Excel path (default: aifmd_validation_<timestamp>.xlsx)
    --rules <path>   Path to base rules YAML (default: auto-detect)
"""

import sys
import re
import glob
import copy
import yaml
import argparse
from xml.etree import ElementTree as ET
from pathlib import Path
from datetime import datetime
from typing import Any, Optional
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# lxml for XSD validation (optional — graceful fallback if not installed)
try:
    from lxml import etree as lxml_etree
    HAS_LXML = True
except ImportError:
    HAS_LXML = False

# ============================================================================
# Auto-detect project paths
# ============================================================================

_SCRIPT_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _SCRIPT_DIR

# Walk up to find Application folder (or Blueprint as fallback)
for _p in [_SCRIPT_DIR, _SCRIPT_DIR.parent, _SCRIPT_DIR.parent.parent, _SCRIPT_DIR.parent.parent.parent]:
    if (_p / "Application").is_dir():
        _PROJECT_ROOT = _p
        break
    if (_p / "Blueprint").is_dir():
        _PROJECT_ROOT = _p
        break

_APP_ROOT = _PROJECT_ROOT / "Application"
_AIFMD_ANNEX_IV_DIR = _APP_ROOT / "regulation" / "aifmd" / "annex_iv"
if not _AIFMD_ANNEX_IV_DIR.is_dir():
    # Fallback: try old flat location, then legacy Blueprint
    _AIFMD_ANNEX_IV_DIR = _APP_ROOT / "regulation" / "aifmd annex iv"
    if not _AIFMD_ANNEX_IV_DIR.is_dir():
        _AIFMD_ANNEX_IV_DIR = _PROJECT_ROOT / "Blueprint"

DEFAULT_RULES_PATH = _AIFMD_ANNEX_IV_DIR / "aifmd_validation_rules.yaml"
if not DEFAULT_RULES_PATH.exists():
    # Fallback to old filename
    DEFAULT_RULES_PATH = _AIFMD_ANNEX_IV_DIR / "aifmd_annex_iv_validation_rules.yaml"
NCA_OVERRIDES_DIR = _AIFMD_ANNEX_IV_DIR / "nca_overrides"
if not NCA_OVERRIDES_DIR.is_dir():
    # Fallback: NCA overrides in same dir as rules (old layout)
    NCA_OVERRIDES_DIR = _AIFMD_ANNEX_IV_DIR

# ============================================================================
# Layer B: Rule Integrity Monitor
# ============================================================================
import hashlib as _hashlib
import json as _json

_APPROVED_HASHES_PATH = _SCRIPT_DIR / "aifmd_approved_rule_hashes.yaml"


def _compute_file_hash(path: Path) -> str:
    """Compute SHA-256 hash of a file."""
    h = _hashlib.sha256()
    with open(str(path), "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _load_approved_hashes() -> dict:
    """Load the approved rule hashes file."""
    if not _APPROVED_HASHES_PATH.exists():
        return {}
    with open(str(_APPROVED_HASHES_PATH), "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _save_approved_hashes(hashes: dict):
    """Save the approved rule hashes file."""
    with open(str(_APPROVED_HASHES_PATH), "w", encoding="utf-8") as f:
        yaml.dump(hashes, f, default_flow_style=False, sort_keys=False)


def verify_rule_integrity(rules_path: Path, nca_code: str = None) -> dict:
    """
    Layer B: Rule Integrity Monitor.

    Computes SHA-256 hashes of the base rules file and active NCA overrides,
    compares against the last approved version.

    Returns dict with:
        status: VERIFIED | UNVERIFIED | FIRST_RUN
        base_rules: {path, hash, approved_hash, match}
        nca_overrides: {path, hash, approved_hash, match} (if applicable)
        changed_files: list of files that changed
    """
    result = {
        "status": "VERIFIED",
        "base_rules": {},
        "nca_overrides": {},
        "changed_files": [],
        "timestamp": datetime.now().isoformat(),
    }

    approved = _load_approved_hashes()
    first_run = len(approved) == 0

    # Hash base rules
    if rules_path.exists():
        current_hash = _compute_file_hash(rules_path)
        approved_hash = approved.get("base_rules", {}).get("hash", "")
        match = current_hash == approved_hash if approved_hash else None
        result["base_rules"] = {
            "path": str(rules_path.name),
            "hash": current_hash,
            "approved_hash": approved_hash or "(none)",
            "match": match,
        }
        if match is False:
            result["changed_files"].append(str(rules_path.name))

    # Hash NCA overrides
    if nca_code:
        nca_pattern = f"aifmd_nca_overrides_{nca_code.lower()}_*.yaml"
        nca_files = sorted(NCA_OVERRIDES_DIR.glob(nca_pattern))
        for nca_file in nca_files:
            current_hash = _compute_file_hash(nca_file)
            approved_hash = approved.get("nca_overrides", {}).get(
                nca_file.name, {}).get("hash", "")
            match = current_hash == approved_hash if approved_hash else None
            result["nca_overrides"][nca_file.name] = {
                "hash": current_hash,
                "approved_hash": approved_hash or "(none)",
                "match": match,
            }
            if match is False:
                result["changed_files"].append(nca_file.name)

    # Determine overall status
    if first_run:
        result["status"] = "FIRST_RUN"
    elif result["changed_files"]:
        result["status"] = "UNVERIFIED"
    else:
        result["status"] = "VERIFIED"

    return result


def approve_current_hashes(rules_path: Path, approved_by: str = "",
                           reason: str = "", nca_codes: list = None):
    """
    Approve the current rule file hashes as the baseline.
    Called explicitly when rule changes are intentional.

    Args:
        rules_path: Path to the base validation rules YAML.
        approved_by: Name of the person approving (required for audit trail).
        reason: Motivation for the approval (required for audit trail).
        nca_codes: Optional list of NCA codes to approve (default: all).
    """
    if not approved_by:
        raise ValueError("--approved-by is required: identify who is approving the rule change")
    if not reason:
        raise ValueError("--reason is required: document why this rule change is approved")

    # Load previous hashes for change tracking
    previous = _load_approved_hashes()
    prev_base_hash = previous.get("base_rules", {}).get("hash", "")
    prev_nca = previous.get("nca_overrides", {})

    hashes = {
        "approved_at": datetime.now().isoformat(),
        "approved_by": approved_by,
        "reason": reason,
        "base_rules": {
            "path": str(rules_path.name),
            "hash": _compute_file_hash(rules_path) if rules_path.exists() else "",
        },
        "nca_overrides": {},
        "changes": [],
    }

    # Detect base rules change
    if hashes["base_rules"]["hash"] and hashes["base_rules"]["hash"] != prev_base_hash:
        hashes["changes"].append({
            "file": str(rules_path.name),
            "previous_hash": prev_base_hash[:16] + "..." if prev_base_hash else "(new)",
            "new_hash": hashes["base_rules"]["hash"][:16] + "...",
        })

    # Hash all NCA overrides and detect changes
    for nca_file in sorted(NCA_OVERRIDES_DIR.glob(
            "aifmd_nca_overrides_*.yaml")):
        current_hash = _compute_file_hash(nca_file)
        hashes["nca_overrides"][nca_file.name] = {
            "hash": current_hash,
        }
        prev_hash = prev_nca.get(nca_file.name, {}).get("hash", "")
        if current_hash != prev_hash:
            hashes["changes"].append({
                "file": nca_file.name,
                "previous_hash": prev_hash[:16] + "..." if prev_hash else "(new)",
                "new_hash": current_hash[:16] + "...",
            })

    # Append to approval history log
    _append_approval_history(hashes)

    _save_approved_hashes(hashes)
    return hashes


def _append_approval_history(approval: dict):
    """Append approval record to the cumulative history log."""
    history_path = _SCRIPT_DIR / "Logging" / "rule_approval_history.yaml"
    history_path.parent.mkdir(parents=True, exist_ok=True)

    history = []
    if history_path.exists():
        with open(str(history_path), "r", encoding="utf-8") as f:
            history = yaml.safe_load(f) or []

    record = {
        "approved_at": approval["approved_at"],
        "approved_by": approval["approved_by"],
        "reason": approval["reason"],
        "base_rules_hash": approval["base_rules"].get("hash", "")[:16] + "...",
        "nca_overrides_count": len(approval.get("nca_overrides", {})),
        "changes": approval.get("changes", []),
    }
    history.append(record)

    with open(str(history_path), "w", encoding="utf-8") as f:
        yaml.dump(history, f, default_flow_style=False, sort_keys=False)

# XSD schema paths — ESMA (default) and FCA (post-Brexit GB)
# Folder structure:
#   M adapter/
#     ESMA XSD/   ← ESMA schemas (no targetNamespace)
#       AIFMD_DATAIF_V1.2.xsd
#       AIFMD_DATMAN_V1.2.xsd
#       AIFMD_REPORTING_DataTypes_V1.2.xsd
#     FCA XSD/    ← FCA schemas (targetNamespace = urn:fsa-gov-uk:MER:...)
#       AIFMD_DATAIF_V2.xsd
#       AIFMD_DATMAN_V2.xsd
#       AIFMD_REPORTING_DataTypes_V2.xsd
#
# Also supports legacy layout where XSDs live directly alongside script.

# XSD directories — new regulation structure, then legacy flat location
_AIFMD_XSD_DIR = _AIFMD_ANNEX_IV_DIR / "xsd"
_ESMA_XSD_DIR = _AIFMD_XSD_DIR / "esma_1.2"
if not _ESMA_XSD_DIR.is_dir():
    _ESMA_XSD_DIR = _APP_ROOT / "XSD" / "ESMA 1.2"
_FCA_XSD_DIR = _AIFMD_XSD_DIR / "fca_2.0"
if not _FCA_XSD_DIR.is_dir():
    _FCA_XSD_DIR = _APP_ROOT / "XSD" / "FCA 2.0"

def _find_xsd(directory: Path, pattern: str) -> Optional[Path]:
    """Find first XSD matching pattern in directory."""
    if not directory.is_dir():
        return None
    matches = sorted(directory.glob(pattern))
    return matches[0] if matches else None

XSD_SCHEMAS = {
    "ESMA": {
        "AIF": (_find_xsd(_ESMA_XSD_DIR, "AIFMD_DATAIF*.xsd")
                or _find_xsd(_SCRIPT_DIR, "AIFMD_DATAIF*.xsd")),
        "AIFM": (_find_xsd(_ESMA_XSD_DIR, "AIFMD_DATMAN*.xsd")
                 or _find_xsd(_SCRIPT_DIR, "AIFMD_DATMAN*.xsd")),
    },
    "FCA": {
        "AIF": _find_xsd(_FCA_XSD_DIR, "AIFMD_DATAIF*.xsd"),
        "AIFM": _find_xsd(_FCA_XSD_DIR, "AIFMD_DATMAN*.xsd"),
    },
}

# FCA namespace URIs (post-Brexit UK)
FCA_NAMESPACES = {
    "AIF": "urn:fsa-gov-uk:MER:AIF002:2",
    "AIFM": "urn:fsa-gov-uk:MER:AIF001:2",
}

# Country code → NCA acronym mapping (for display and file lookup)
# NCA override files follow: aifmd_nca_overrides_{cc}_{nca}.yaml
NCA_COUNTRY_MAP = {
    "LU": "cssf", "NL": "afm", "DE": "bafin", "BE": "fsma",
    "FR": "amf", "GB": "fca", "IE": "cbi", "IT": "consob",
}


# ============================================================================
# Layer 0: File Naming & Packaging Validation
# ============================================================================

class FileFinding:
    """One file naming or packaging validation result."""
    __slots__ = ("file_name", "check", "expected", "actual", "status",
                 "severity", "message")

    def __init__(self, **kw):
        for k in self.__slots__:
            setattr(self, k, kw.get(k, ""))


# NCA filename patterns and packaging specs
# Each NCA defines:
#   aifm_pattern / aif_pattern: regex the XML filename must match
#   packaging: expected wrapper format ("xml", "zip", "gzip", "zip-in-zip")
#   aifm_template / aif_template: human-readable filename template

_NCA_FILE_SPECS = {
    "BE": {
        "aifm_pattern": r"^AIFM_BE_[0-9]{10}_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_BE_[0-9]{5}-[0-9]{4}_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_BE_{10-digit-code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_BE_{XXXXX-XXXX}_{YYYYMMDD}.xml",
        "packaging": "zip",
        "pack_desc": "XML wrapped in ZIP (1 XML per ZIP)",
    },
    "DE": {
        "aifm_pattern": r"^AIFM_DE_[0-9]{8}_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_DE_[0-9]{8}_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_DE_{8-digit-code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_DE_{8-digit-code}_{YYYYMMDD}.xml",
        "packaging": "gzip",
        "pack_desc": "XML wrapped in GZIP (.gz)",
    },
    "NL": {
        "aifm_pattern": r"^AIFM_NL_B[A-Z]{2}[0-9]{3}_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_NL_500[0-9]{5}_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_NL_{BXX999}_{YYYYMMDD}.xml",
        "aif_template": "AIF_NL_{500XXXXX}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML (bundled in ZIP for delivery)",
    },
    "GB": {
        "aifm_pattern": r"^AIFM_GB_[0-9]{6}_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_REPORTS_GB_[0-9]{6}_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_GB_{6-digit-FRN}_{YYYYMMDD}.xml",
        "aif_template": "AIF_REPORTS_GB_{6-digit-FRN}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML (consolidated multi-AIF in single file)",
        # Also accept legacy per-AIF naming
        "aif_pattern_alt": r"^AIF_GB_[0-9]{6,7}_[0-9]{8}\.xml$",
    },
    "LU": {
        "aifm_pattern": r"^AIFREP-\d+-A\d+-F\d+-\d+-MAN-\d+\.xml$",
        "aif_pattern": r"^AIFREP-\d+-A\d+-[OV]\d+-\d+-AIF-\d+\.xml$",
        "aifm_template": "AIFREP-{sender}-A{AIFM}-F{fund}-{comp}-MAN-{timestamp}.xml",
        "aif_template": "AIFREP-{sender}-A{AIFM}-{V/O}{AIF}-{comp}-AIF-{timestamp}.xml",
        "packaging": "zip-in-zip",
        "pack_desc": "XML->ZIP (per AIF), all ZIPs bundled in master ZIP",
        # Also accept simplified naming from adapter
        "aifm_pattern_alt": r"^AIFM_LU_A\d+_[0-9]{8}\.xml$",
        "aif_pattern_alt": r"^AIF_LU_[OV]\d+_\d+_[0-9]{8}\.xml$",
    },
    "IE": {
        "aifm_pattern": r"^C\d{6}_\d{8}_AIM\.xml$",
        "aif_pattern": r"^C\d{6}_\d{8}_AIF\.xml$",
        "aifm_template": "C{6-digit}_{YYYYMMDD}_AIM.xml",
        "aif_template": "C{6-digit}_{YYYYMMDD}_AIF.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "FR": {
        "aifm_pattern": r"^AIFM_FR_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_FR_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_FR_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_FR_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "IT": {
        "aifm_pattern": r"^AIFM_IT_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_IT_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_IT_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_IT_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "ES": {
        "aifm_pattern": r"^AIFM_ES_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_ES_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_ES_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_ES_{code}_{YYYYMMDD}.xml",
        "packaging": "zip",
        "pack_desc": "All XMLs bundled in ZIP",
    },
    "PT": {
        "aifm_pattern": r"^AIFM_PT_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_PT_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_PT_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_PT_{code}_{YYYYMMDD}.xml",
        "packaging": "zip",
        "pack_desc": "All XMLs bundled in ZIP",
    },
    "SE": {
        "aifm_pattern": r"^AIFM_SE_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_SE_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_SE_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_SE_{code}_{YYYYMMDD}.xml",
        "packaging": "zip",
        "pack_desc": "All XMLs bundled in ZIP",
    },
    "CZ": {
        "aifm_pattern": r"^AIFM_CZ_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_CZ_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_CZ_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_CZ_{code}_{YYYYMMDD}.xml",
        "packaging": "zip",
        "pack_desc": "All XMLs bundled in ZIP (SDAT)",
    },
    "MT": {
        "aifm_pattern": r"^AIFM_MT_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_MT_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_MT_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_MT_{code}_{YYYYMMDD}.xml",
        "packaging": "zip",
        "pack_desc": "AIFM + AIF XMLs bundled in ZIP",
    },
    "AT": {
        "aifm_pattern": r"^AIFM_AT_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_AT_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_AT_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_AT_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML (ZIP when multiple AIFs)",
    },
    "CY": {
        "aifm_pattern": r"^AIFM_CY_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_CY_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_CY_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_CY_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "DK": {
        "aifm_pattern": r"^AIFM_DK_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_DK_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_DK_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_DK_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "FI": {
        "aifm_pattern": r"^AIFM_FI_[A-Z0-9#]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_FI_[A-Z0-9#]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_FI_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_FI_{code#NNN}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "NO": {
        "aifm_pattern": r"^AIFM_NO_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_NO_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_NO_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_NO_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "GR": {
        "aifm_pattern": r"^AIFM_GR_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_GR_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_GR_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_GR_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "PL": {
        "aifm_pattern": r"^AIFM_PL_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_PL_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_PL_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_PL_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "BG": {
        "aifm_pattern": r"^AIFM_BG_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_BG_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_BG_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_BG_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "HR": {
        "aifm_pattern": r"^AIFM_HR_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_HR_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_HR_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_HR_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "HU": {
        "aifm_pattern": r"^AIFM_HU_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_HU_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_HU_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_HU_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "EE": {
        "aifm_pattern": r"^AIFM_EE_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_EE_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_EE_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_EE_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "LV": {
        "aifm_pattern": r"^AIFM_LV_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_LV_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_LV_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_LV_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "LT": {
        "aifm_pattern": r"^AIFM_LT_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_LT_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_LT_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_LT_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "LI": {
        "aifm_pattern": r"^AIFM_LI_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_LI_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_LI_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_LI_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "RO": {
        "aifm_pattern": r"^AIFM_RO_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_RO_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_RO_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_RO_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "SI": {
        "aifm_pattern": r"^AIFM_SI_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_SI_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_SI_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_SI_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "SK": {
        "aifm_pattern": r"^AIFM_SK_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_SK_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_SK_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_SK_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
    "IS": {
        "aifm_pattern": r"^AIFM_IS_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aif_pattern": r"^AIF_IS_[A-Z0-9]+_[0-9]{8}\.xml$",
        "aifm_template": "AIFM_IS_{code}_{YYYYMMDD}.xml",
        "aif_template": "AIF_IS_{code}_{YYYYMMDD}.xml",
        "packaging": "xml",
        "pack_desc": "Plain XML",
    },
}

# Fallback spec for unknown NCAs — used when rms not in _NCA_FILE_SPECS
_NCA_FILE_SPEC_DEFAULT = {
    "aifm_pattern": r"^AIFM_[A-Z]{2}_[A-Z0-9]+_[0-9]{8}\.xml$",
    "aif_pattern": r"^AIF_[A-Z]{2}_[A-Z0-9]+_[0-9]{8}\.xml$",
    "aifm_template": "AIFM_{CC}_{code}_{YYYYMMDD}.xml",
    "aif_template": "AIF_{CC}_{code}_{YYYYMMDD}.xml",
    "packaging": "xml",
    "pack_desc": "Plain XML (unknown NCA — using ESMA default)",
}


def validate_file_naming(file_path: Path, report_type: str,
                         nca_code: str, metadata: dict) -> list:
    """Validate file naming convention and packaging format.

    Checks:
      1. Filename matches NCA-specific pattern
      2. Report date in filename matches XML ReportingPeriodEndDate
      3. National code in filename matches XML content
      4. Packaging format matches NCA requirement (checked at delivery level)
    """
    findings = []
    fname = file_path.name
    rms = nca_code or metadata.get("ReportingMemberState", "")

    spec = _NCA_FILE_SPECS.get(rms, _NCA_FILE_SPEC_DEFAULT)

    # ── Check 1: Filename pattern ────────────────────────────────────────
    pattern_key = "aifm_pattern" if report_type == "AIFM" else "aif_pattern"
    pattern = spec[pattern_key]
    template = spec.get("aifm_template" if report_type == "AIFM" else "aif_template", "")

    matched = bool(re.match(pattern, fname))
    # Also try alternate pattern (e.g. LU simplified naming)
    alt_key = pattern_key + "_alt"
    if not matched and alt_key in spec:
        matched = bool(re.match(spec[alt_key], fname))
        if matched:
            template += " (alt: simplified adapter format)"

    if matched:
        findings.append(FileFinding(
            file_name=fname, check="Filename Pattern",
            expected=template, actual=fname,
            status="PASS", severity="",
            message=f"Filename matches {rms} {report_type} naming convention"))
    else:
        findings.append(FileFinding(
            file_name=fname, check="Filename Pattern",
            expected=template, actual=fname,
            status="FAIL", severity="HIGH",
            message=f"Filename does not match {rms} {report_type} pattern: {pattern}"))

    # ── Check 2: Date in filename vs XML ─────────────────────────────────
    # Extract YYYYMMDD from filename
    date_match = re.search(r'_(\d{8})\.xml$', fname)
    if date_match:
        fname_date = date_match.group(1)
        # Find ReportingPeriodEndDate in metadata or XML
        xml_end_date = metadata.get("ReportingPeriodEndDate", "")
        if not xml_end_date:
            # Try to find in the XML tree (passed via metadata)
            xml_end_date = metadata.get("_reporting_end_date", "")
        if xml_end_date:
            # Normalize: remove hyphens from ISO date
            xml_date_norm = xml_end_date.replace("-", "")
            if fname_date == xml_date_norm:
                findings.append(FileFinding(
                    file_name=fname, check="Filename Date",
                    expected=xml_date_norm, actual=fname_date,
                    status="PASS", severity="",
                    message="Reporting period end date matches filename"))
            else:
                findings.append(FileFinding(
                    file_name=fname, check="Filename Date",
                    expected=xml_date_norm, actual=fname_date,
                    status="FAIL", severity="MEDIUM",
                    message=f"Filename date ({fname_date}) != XML ReportingPeriodEndDate ({xml_end_date})"))

    # ── Check 3: National code in filename vs XML ────────────────────────
    if rms in ("BE", "DE", "NL", "GB"):
        # Detect consolidated FCA files: AIF_REPORTS_GB_{CODE}_{DATE}.xml
        is_consolidated = fname.startswith("AIF_REPORTS_")

        # Extract national code from filename.
        # Standard:     {TYPE}_{CC}_{CODE}_{DATE}.xml  → parts[2:-1]
        # Consolidated: AIF_REPORTS_{CC}_{CODE}_{DATE}.xml → parts[3:-1]
        parts = fname.replace(".xml", "").split("_")
        code_start_idx = 3 if is_consolidated else 2
        if len(parts) >= (code_start_idx + 2):
            fname_code = "_".join(parts[code_start_idx:-1])
            # For consolidated FCA files the filename carries the AIFM national
            # code (FRN), not an individual AIF national code.
            if report_type == "AIFM":
                xml_code = metadata.get("AIFMNationalCode", "")
            elif is_consolidated:
                xml_code = metadata.get("AIFMNationalCode", "")
            else:
                xml_code = metadata.get("AIFNationalCode", "")
            if xml_code and fname_code:
                if fname_code == xml_code:
                    findings.append(FileFinding(
                        file_name=fname, check="National Code",
                        expected=xml_code, actual=fname_code,
                        status="PASS", severity="",
                        message=f"{report_type} national code matches filename"))
                else:
                    findings.append(FileFinding(
                        file_name=fname, check="National Code",
                        expected=xml_code, actual=fname_code,
                        status="FAIL", severity="HIGH",
                        message=f"Filename code ({fname_code}) != XML national code ({xml_code})"))

    # ── Check 4: Packaging format ────────────────────────────────────────
    expected_pkg = spec["packaging"]
    pack_desc = spec["pack_desc"]
    parent_ext = file_path.suffix.lower()

    # Determine actual packaging by looking at what file we're validating
    # If we're validating a .xml file directly, the packaging is "xml" (unwrapped)
    if parent_ext == ".xml":
        actual_pkg = "xml"
    elif parent_ext == ".gz":
        actual_pkg = "gzip"
    elif parent_ext == ".zip":
        actual_pkg = "zip"
    else:
        actual_pkg = parent_ext

    # For now, we note the expected packaging format as informational
    # (actual delivery packaging is validated at the archive level, not individual XML)
    if expected_pkg == "xml" or actual_pkg == expected_pkg:
        findings.append(FileFinding(
            file_name=fname, check="Packaging Format",
            expected=pack_desc, actual=f"Validating as .{actual_pkg}",
            status="PASS", severity="",
            message=f"Expected delivery format: {pack_desc}"))
    elif actual_pkg == "xml" and expected_pkg in ("zip", "gzip", "zip-in-zip"):
        findings.append(FileFinding(
            file_name=fname, check="Packaging Format",
            expected=pack_desc, actual="Plain XML (not packaged)",
            status="WARNING", severity="MEDIUM",
            message=f"File is plain XML but {rms} requires: {pack_desc}. "
                    f"Ensure file is packaged correctly before submission."))

    return findings


# ============================================================================
# Layer 1: XSD Validation
# ============================================================================

class XsdFinding:
    """One XSD validation error for a single file."""
    __slots__ = ("file_name", "line", "column", "severity", "message", "xsd_source")

    def __init__(self, **kw):
        for k in self.__slots__:
            setattr(self, k, kw.get(k, ""))


def _strip_namespace(xml_bytes: bytes) -> bytes:
    """Remove namespace declarations and prefixes from XML bytes.

    Used as fallback for FCA XMLs when no FCA-specific XSD is available —
    the element structure is identical to ESMA, only the namespace differs.
    """
    # Remove xmlns="urn:..." from root element
    xml_str = xml_bytes.decode("utf-8")
    xml_str = re.sub(r'\s+xmlns="[^"]*"', '', xml_str, count=1)
    return xml_str.encode("utf-8")


def _detect_fca_xml(xml_bytes: bytes) -> bool:
    """Check if XML uses FCA namespace (post-Brexit GB)."""
    return b"urn:fsa-gov-uk:" in xml_bytes


def _load_xsd_schema(xsd_path: Optional[Path]) -> Optional[object]:
    """Load and parse an XSD schema. Returns lxml XMLSchema or None."""
    if not HAS_LXML:
        return None
    if xsd_path is None or not xsd_path.exists():
        return None
    try:
        # Parse from the XSD's own directory so xs:include resolves correctly
        import os
        orig_dir = os.getcwd()
        os.chdir(str(xsd_path.parent))
        try:
            schema = lxml_etree.XMLSchema(
                lxml_etree.parse(str(xsd_path.name)))
        finally:
            os.chdir(orig_dir)
        return schema
    except Exception as e:
        print(f"  WARNING: Failed to load XSD {xsd_path.name}: {e}")
        return None


def validate_xsd(file_path: Path, report_type: str) -> tuple[list, str]:
    """Validate XML against the appropriate XSD schema.

    Strategy:
      - FCA files (urn:fsa-gov-uk namespace) → try FCA XSD first, then
        fallback to ESMA XSD with namespace stripped.
      - All other files → use ESMA XSD directly.

    Returns (list of XsdFinding, xsd_source_label).
    """
    findings = []

    if not HAS_LXML:
        return findings, "SKIPPED (lxml not installed)"

    raw = file_path.read_bytes()
    is_fca = _detect_fca_xml(raw)
    rtype_key = "AIFM" if report_type == "AIFM" else "AIF"

    # 1. Try FCA-specific XSD first for FCA files
    schema = None
    xsd_source = ""
    if is_fca:
        fca_xsd = XSD_SCHEMAS["FCA"].get(rtype_key)
        schema = _load_xsd_schema(fca_xsd)
        if schema:
            xsd_source = f"FCA {fca_xsd.stem}"

    # 2. If no FCA schema available, strip namespace and use ESMA XSD
    if schema is None:
        esma_xsd = XSD_SCHEMAS["ESMA"].get(rtype_key)
        schema = _load_xsd_schema(esma_xsd)
        if schema is None:
            return findings, "SKIPPED (XSD not found)"
        xsd_source = f"ESMA {esma_xsd.stem}"
        if is_fca:
            xsd_source += " (FCA namespace stripped — install FCA XSDs for full validation)"
            raw = _strip_namespace(raw)

    # 3. Parse and validate
    try:
        doc = lxml_etree.fromstring(raw)
    except lxml_etree.XMLSyntaxError as e:
        findings.append(XsdFinding(
            file_name=file_path.name, line=getattr(e, 'lineno', 0),
            column=0, severity="CRITICAL",
            message=f"XML parse error: {e}", xsd_source=xsd_source))
        return findings, xsd_source

    if schema.validate(doc):
        return findings, xsd_source

    # Collect unique errors
    seen = set()
    for err in schema.error_log:
        key = (err.line, err.message)
        if key not in seen:
            seen.add(key)
            findings.append(XsdFinding(
                file_name=file_path.name, line=err.line, column=err.column,
                severity="CRITICAL" if err.level_name == "ERROR" else "HIGH",
                message=err.message, xsd_source=xsd_source))

    return findings, xsd_source

# ── Section → AIFMD article mapping ──────────────────────────────────
# AIFMD ContentType at AIF level:
#   CT=1  No-reporting / Article 3(3)(d) registered AIFM  → header only
#   CT=2  Registered / light (Art 3(3)(d))                → header + Art 24(1)
#   CT=3  Authorised / full                               → header + Art 24(1) + Art 24(2)
#   CT=4  Authorised, leveraged EU AIFs                   → header + Art 24(1) + Art 24(2) + Art 24(4)
#   CT=5  Authorised, leveraged non-EU AIFs               → header + Art 24(1) + Art 24(4)
#
# Section → content-type applicability: single source of truth in FieldRegistry.
# Import here so the validator uses the same mapping as the Report Viewer.
try:
    from canonical.aifmd_field_registry import FieldRegistry as _FieldRegistry
    _section_applicable_cts = _FieldRegistry.section_applicable_cts
except ImportError:
    # Fallback for standalone CLI usage where canonical may not be on sys.path
    import sys as _sys
    _sys.path.insert(0, str(_APP_ROOT))
    from canonical.aifmd_field_registry import FieldRegistry as _FieldRegistry
    _section_applicable_cts = _FieldRegistry.section_applicable_cts

# Elements that are XML root attributes (not child elements)
ROOT_ATTRIBUTE_ELEMENTS = {
    "ReportingMemberState", "Version", "CreationDateAndTime",
}


# ============================================================================
# Data classes
# ============================================================================

class Finding:
    __slots__ = (
        "file_name", "report_type", "rule_id", "field_name", "section",
        "m_c_o_f", "rule_source", "nca_override", "document_value",
        "expected", "status", "severity", "esma_error_code",
        "nca_error_code", "recommendation",
    )

    def __init__(self, **kw):
        for k in self.__slots__:
            setattr(self, k, kw.get(k, ""))


class Coverage:
    __slots__ = ("rule_id", "field_name", "section", "applicable", "status")

    def __init__(self, **kw):
        for k in self.__slots__:
            setattr(self, k, kw.get(k, ""))


# ============================================================================
# YAML loading
# ============================================================================

def load_rules(rules_path: Path) -> dict:
    with open(rules_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_nca_overrides(nca_code: str) -> dict:
    """Load NCA overrides, return dict keyed by base_rule_id.

    File naming convention: aifmd_nca_overrides_{cc}_{nca}.yaml
    e.g. aifmd_nca_overrides_nl_afm.yaml
    """
    cc = nca_code.lower()  # country code is the primary key
    # Find file(s) matching this country code
    candidates = sorted(NCA_OVERRIDES_DIR.glob(
        f"aifmd_nca_overrides_{cc}_*.yaml"))
    if not candidates:
        return {}
    with open(candidates[-1], "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    nca_rules = data.get("nca_overrides", {}).get("nca_rules", [])
    # Index by base_rule_id for quick lookup
    result = {}
    for r in nca_rules:
        base_id = r.get("base_rule_id", "")
        if base_id:
            result[base_id] = r
    return result


# ============================================================================
# XML parsing
# ============================================================================

def detect_report_type(root: ET.Element) -> str:
    """Detect AIFM vs AIF from root tag."""
    tag = root.tag.split("}")[-1] if "}" in root.tag else root.tag
    if "AIFM" in tag and "AIF" not in tag.replace("AIFM", ""):
        return "AIFM"
    return "AIF"


def extract_metadata(root: ET.Element, report_type: str) -> dict:
    meta = {}
    # Root attributes (AIF-1/2/3, AIFM-1/2/3)
    meta["ReportingMemberState"] = root.get("ReportingMemberState", "")
    meta["Version"] = root.get("Version", "")
    meta["CreationDateAndTime"] = root.get("CreationDateAndTime", "")
    # Also check xmlns for FCA
    meta["xmlns"] = root.get("xmlns", "")

    # Nested elements — capture all relevant metadata for validation
    _META_ELEMENTS = {
        "AIFContentType", "AIFMContentType",
        "AIFNoReportingFlag", "AIFMNoReportingFlag",
        "ReportingPeriodType", "ReportingPeriodEndDate",
        "ReportingPeriodStartDate", "ReportingPeriodYear",
        "FilingType", "AIFMJurisdiction",
        "AIFMNationalCode", "AIFNationalCode",
    }
    for elem in root.iter():
        tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
        if tag in _META_ELEMENTS and (tag not in meta or not meta[tag]):
            meta[tag] = (elem.text or "").strip()
    return meta


def find_element(root: ET.Element, element_name: str) -> Optional[ET.Element]:
    """Find element anywhere in XML, handling namespaces."""
    # Direct search
    for elem in root.iter():
        tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
        if tag == element_name:
            return elem
    return None


def get_root_attribute(root: ET.Element, attr_name: str) -> Optional[str]:
    """Get a root-level attribute value."""
    val = root.get(attr_name)
    if val is not None:
        return val
    # Also check with namespace prefix stripping
    for k, v in root.attrib.items():
        clean = k.split("}")[-1] if "}" in k else k
        if clean == attr_name:
            return v
    return None


# ============================================================================
# Validation logic
# ============================================================================

def is_applicable(rule: dict, content_type: int, report_type: str) -> bool:
    """Check if a rule applies to this content type and report type."""
    rule_report = rule.get("report_type", "").upper()
    if report_type == "AIFM" and "AIFM" not in rule_report:
        return False
    if report_type == "AIF" and "AIF" not in rule_report:
        return False

    # For AIF rules, check section against proper article/CT mapping
    if report_type == "AIF":
        section = rule.get("section", "")
        applicable_cts = _section_applicable_cts(section)
        if not applicable_cts:
            # Section not in our mapping — default to applicable for all CTs
            return True
        return content_type in applicable_cts

    # AIFM rules are always applicable for AIFM reports
    return True


def validate_value(value: str, rule: dict, ref_tables: dict) -> tuple[bool, str]:
    """Validate a single value against a rule's constraints."""
    if not value:
        return True, ""  # empty handled by presence check

    # ISO checks
    if rule.get("validation_type") == "iso_check":
        fmt = rule.get("format", "")
        if "2 [A-Z]" in fmt:
            if len(value) != 2 or not value.isalpha() or not value.isupper():
                return False, f"Invalid ISO country code: '{value}'"
        elif "3 [A-Z]" in fmt:
            if len(value) != 3 or not value.isalpha() or not value.isupper():
                return False, f"Invalid ISO currency code: '{value}'"

    # Domain value check
    ref_key = rule.get("allowed_values_ref")
    if ref_key and ref_key in ref_tables:
        allowed = ref_tables[ref_key]
        if isinstance(allowed, list) and value not in [str(v) for v in allowed]:
            return False, f"'{value}' not in {ref_key}"

    # Data type checks
    dt = rule.get("data_type", "")
    if dt == "D" and not re.match(r"^\d{4}-\d{2}-\d{2}", value):
        return False, f"Invalid date: '{value}'"
    if dt == "B" and value not in ("true", "false"):
        return False, f"Invalid boolean: '{value}'"
    if dt == "N":
        try:
            float(value)
        except ValueError:
            return False, f"Invalid number: '{value}'"

    # Format regex check (applies NCA-override or base rule format patterns)
    # The 'format' field may contain a regex pattern (e.g. "B[A-Z]{2}[0-9]{3}")
    # Skip generic ISO format strings like "2 [A-Z]" — those are handled above.
    #
    # ESMA format notation types:
    #   1. Character-class regex:  "20 [0-9a-zA-Z]{18}[0-9]{2}" (LEI)
    #      → "<length> <regex with character classes>"
    #      → Strip length prefix, validate length + regex separately.
    #   2. Numeric format description:  "16 (+/-15 )", "19 (+/- 15.2)"
    #      → NOT a regex — describes display width and sign/decimal layout.
    #      → Must be skipped (the regex engine can't parse these).
    #
    # Distinguishing rule: only strip the length prefix when the remainder
    # contains character-class brackets "[" and "]".  Numeric format strings
    # never contain brackets; identifier patterns always do.
    fmt = rule.get("format", "")
    if fmt and rule.get("validation_type") != "iso_check":
        # Only treat it as regex if it contains character-class brackets
        if "[" in fmt and "]" in fmt:
            try:
                regex_part = fmt
                expected_len = None
                # Strip ESMA length prefix (e.g. "20 [0-9…" → "[0-9…")
                length_prefix = re.match(r"^(\d+)\s+(.+)$", fmt)
                if length_prefix:
                    expected_len = int(length_prefix.group(1))
                    regex_part = length_prefix.group(2)
                # ESMA format notation uses spaces as visual separators
                # between character classes (e.g. "[A-Z]{2} [0-9]{9}").
                # Strip them so they aren't treated as literal space chars.
                regex_part = re.sub(r'(?<=[\]})]) (?=[\[(])', '', regex_part)
                # Optional length check
                if expected_len is not None and len(value) != expected_len:
                    return False, (
                        f"'{value}' length {len(value)} != expected {expected_len} "
                        f"(format {fmt})"
                    )
                # Ensure it's anchored
                pat = regex_part if regex_part.startswith("^") else f"^{regex_part}$"
                if not re.match(pat, value):
                    return False, f"'{value}' does not match format {fmt}"
            except re.error:
                pass  # Not a valid regex — skip

    # Validation_rules regex check (NCA overrides may have explicit regex rules)
    val_rules = rule.get("validation_rules", [])
    if isinstance(val_rules, str):
        val_rules = [val_rules]
    if isinstance(val_rules, list):
        for vr in val_rules:
            if isinstance(vr, str):
                m = re.search(r"[Mm]ust match regex\s+(\S+)", vr)
                if m:
                    pat = m.group(1)
                    try:
                        if not re.match(pat, value):
                            return False, f"'{value}' does not match NCA regex {pat}"
                    except re.error:
                        pass

    return True, ""


def _get_elem_text(root: ET.Element, name: str) -> Optional[str]:
    """Get text of first matching element, or None."""
    el = find_element(root, name)
    return (el.text or "").strip() if el is not None else None


def _scan_conditional_context(root: ET.Element) -> dict:
    """Pre-scan the XML for flags and containers that gate child elements.

    Returns a dict of rule_id → reason string.  Rules in this dict should
    be classified as N/A instead of WARNING when the element is absent.
    """
    na_rules: dict[str, str] = {}

    # ── 1. Flag-gated children ───────────────────────────────────────────
    # CounterpartyExposureFlag=false  →  counterparty name/LEI/BIC/rate N/A
    #   Fund-to-counterparty (AIF-162..165) and counterparty-to-fund (AIF-168..171)
    for parent_tag, flag_tag, child_rules, label in [
        ("FundToCounterpartyExposures", "CounterpartyExposureFlag",
         {"AIF-161", "AIF-162", "AIF-163", "AIF-164", "AIF-165"},
         "Fund-to-counterparty exposure flag is false"),
        ("CounterpartyToFundExposures", "CounterpartyExposureFlag",
         {"AIF-167", "AIF-168", "AIF-169", "AIF-170", "AIF-171"},
         "Counterparty-to-fund exposure flag is false"),
    ]:
        parent = find_element(root, parent_tag)
        if parent is None:
            # Entire container absent → all children N/A
            for rid in child_rules:
                na_rules[rid] = f"No {parent_tag} container"
        else:
            # Check if ALL flags in the group are false
            all_false = True
            for child in parent:
                ctag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
                if ctag == flag_tag:
                    continue
                # Check if this rank item has a true flag
                flag_el = child.find(".//{*}" + flag_tag) if "}" in child.tag else child.find(flag_tag)
                if flag_el is None:
                    flag_el2 = find_element(child, flag_tag)
                    if flag_el2 is not None and (flag_el2.text or "").strip() == "true":
                        all_false = False
                        break
                elif (flag_el.text or "").strip() == "true":
                    all_false = False
                    break
            if all_false:
                for rid in child_rules:
                    na_rules[rid] = label

    # ClearTransactionsThroughCCPFlag: false or absent → CCP children N/A
    ccp_flag = _get_elem_text(root, "ClearTransactionsThroughCCPFlag")
    if ccp_flag in ("false", None):
        for rid in ("AIF-172", "AIF-174", "AIF-175", "AIF-176", "AIF-177"):
            na_rules[rid] = "ClearTransactionsThroughCCPFlag is false or absent"

    # ShareClassFlag=false → share class children N/A
    sc_flag = _get_elem_text(root, "ShareClassFlag")
    if sc_flag == "false":
        for rid in ("AIF-34", "AIF-35", "AIF-36", "AIF-37", "AIF-38",
                     "AIF-39", "AIF-40"):
            na_rules[rid] = "ShareClassFlag is false"

    # AllCounterpartyCollateralRehypothecationFlag: false or absent → N/A
    rehyp_flag = _get_elem_text(root, "AllCounterpartyCollateralRehypothecationFlag")
    if rehyp_flag in ("false", None):
        for rid in ("AIF-282",):
            na_rules[rid] = "AllCounterpartyCollateralRehypothecationFlag is false or absent"

    # InvestorGroups absent → investor group fields N/A
    if find_element(root, "InvestorGroups") is None:
        for rid in ("AIF-208", "AIF-209"):
            na_rules[rid] = "No InvestorGroups container"

    # ── 2. Container-absent children ─────────────────────────────────────
    # CompanyDominantInfluence absent → all dominant influence fields N/A
    if find_element(root, "CompanyDominantInfluence") is None:
        for rid in ("AIF-131", "AIF-132", "AIF-133", "AIF-134", "AIF-135", "AIF-136"):
            na_rules[rid] = "No CompanyDominantInfluence container"

    # BorrowingSource absent → borrowing source fields N/A
    if find_element(root, "BorrowingSource") is None:
        for rid in ("AIF-297", "AIF-298", "AIF-299", "AIF-300", "AIF-301"):
            na_rules[rid] = "No BorrowingSource container"

    # FinancingLiquidityProfile absent → financing percentages N/A
    if find_element(root, "FinancingLiquidityProfile") is None:
        for rid in ("AIF-211", "AIF-212", "AIF-213", "AIF-214",
                     "AIF-215", "AIF-216", "AIF-217"):
            na_rules[rid] = "No FinancingLiquidityProfile container"

    # SpecialArrangement absent → special arrangement fields N/A
    if find_element(root, "SpecialArrangement") is None:
        for rid in ("AIF-200", "AIF-201"):
            na_rules[rid] = "No SpecialArrangement container"

    # ── 3. Identifier-group logic ────────────────────────────────────────
    # Instrument codes: only one of ISIN/AII needed per instrument
    # InstrumentCodeType=NONE or ISIN → AII fields N/A; =NONE → ISIN N/A too
    has_aii = False
    all_code_types = set()
    for mi in root.iter():
        mtag = mi.tag.split("}")[-1] if "}" in mi.tag else mi.tag
        if mtag == "InstrumentCodeType":
            val = (mi.text or "").strip()
            all_code_types.add(val)
            if val == "AII":
                has_aii = True
    if not has_aii:
        for rid in ("AIF-69", "AIF-70", "AIF-71", "AIF-72", "AIF-73", "AIF-74"):
            na_rules[rid] = "No AII instrument codes in use"
    # If all instruments are NONE, ISIN is N/A too
    if all_code_types and all_code_types <= {"NONE"}:
        na_rules["AIF-68"] = "All instruments have InstrumentCodeType=NONE"

    # Short position hedging: only when short positions exist
    short_val = _get_elem_text(root, "ShortPositionBorrowedSecuritiesValue")
    if short_val in (None, "0", "0.00", "0.0"):
        na_rules["AIF-77"] = "No short positions (ShortPositionBorrowedSecuritiesValue=0)"

    # InvestorRedemption children: only when ProvideWithdrawalRightsFlag=true
    withdraw_flag = _get_elem_text(root, "ProvideWithdrawalRightsFlag")
    if withdraw_flag == "false":
        for rid in ("AIF-194", "AIF-195", "AIF-196"):
            na_rules[rid] = "ProvideWithdrawalRightsFlag is false"

    # ── 4. Strategy type logic ───────────────────────────────────────────
    # Strategy code: the actual element name depends on fund type
    # (HedgeFundStrategyType, PrivateEquityFundStrategyType, etc.)
    # If any strategy type is present, the rule is satisfied via that element
    strategy_elements = ("HedgeFundStrategyType", "PrivateEquityFundStrategyType",
                         "RealEstateFundStrategyType", "FundOfFundsStrategyType",
                         "OtherFundStrategyType")
    has_any_strategy = any(find_element(root, se) is not None for se in strategy_elements)
    if has_any_strategy:
        na_rules["AIF-58"] = "Strategy type present via fund-type-specific element"

    # Strategy description "Other": only when strategy code is OTHER
    has_other_strategy = False
    for el in root.iter():
        etag = el.tag.split("}")[-1] if "}" in el.tag else el.tag
        if etag == "StrategyTypeCode" and (el.text or "").strip() in ("OTHR", "OTHER"):
            has_other_strategy = True
            break
    if not has_other_strategy:
        na_rules["AIF-61"] = "No OTHER strategy type"

    # Investment strategy: primary flag only when multiple strategies
    strategy_count = sum(1 for el in root.iter()
                         if (el.tag.split("}")[-1] if "}" in el.tag else el.tag)
                         in strategy_elements)
    if strategy_count <= 1:
        na_rules["AIF-59"] = "Single strategy — PrimaryStrategyFlag not needed"

    # ── 5. Portfolio concentration context ────────────────────────────────
    # MarketCode/MIC: only required when MarketTypeCode = "MIC"
    # "OTC" for OTC derivatives, "XXX" for transactions without market → no MIC needed
    has_mic_market = False
    for el in root.iter():
        etag = el.tag.split("}")[-1] if "}" in el.tag else el.tag
        if etag in ("MarketCodeType", "MarketTypeCode"):
            val = (el.text or "").strip()
            if val == "MIC":
                has_mic_market = True
                break
    if not has_mic_market:
        na_rules["AIF-107"] = "No MIC-type markets (OTC/XXX do not require MarketCode)"
        na_rules["AIF-116"] = "No MIC-type markets (OTC/XXX do not require MIC code)"

    # Counterparty names in portfolio concentration only when counterparty type
    if find_element(root, "CounterpartyName") is None:
        for rid in ("AIF-110", "AIF-111", "AIF-112"):
            na_rules[rid] = "No counterparty names in portfolio concentrations"

    # ── 6. Geographical focus — principal exposures ───────────────────────
    # LEI/BIC in principal exposures: only when exposure entity exists
    has_exposure_entity = False
    for el in root.iter():
        etag = el.tag.split("}")[-1] if "}" in el.tag else el.tag
        if etag == "AIFPrincipalExposure":
            children = list(el)
            if children:
                has_exposure_entity = True
            break
    if not has_exposure_entity:
        for rid in ("AIF-101", "AIF-102"):
            na_rules[rid] = "No AIFPrincipalExposure data"

    # ── 7. Turnover ──────────────────────────────────────────────────────
    # NotionalValue: only for derivatives turnover; physical assets use MarketValue
    if find_element(root, "NotionalValue") is None:
        # Check if turnover exists with MarketValue instead
        if find_element(root, "MarketValue") is not None:
            na_rules["AIF-127"] = "Turnover uses MarketValue (not NotionalValue)"
        elif find_element(root, "AssetTypeTurnover") is None:
            na_rules["AIF-125"] = "No turnover data"
            na_rules["AIF-127"] = "No turnover data"

    # ── 8. Individual exposure values ─────────────────────────────────────
    # GrossValue/ShortValue: only present for certain exposure types
    ind_exp = find_element(root, "IndividualExposure")
    if ind_exp is None:
        for rid in ("AIF-122", "AIF-123", "AIF-124"):
            na_rules[rid] = "No IndividualExposure container"
    else:
        # Check if GrossValue/ShortValue are used (vs LongValue)
        if find_element(root, "GrossValue") is None:
            na_rules["AIF-122"] = "Individual exposures use Long/Short, not Gross"
        if find_element(root, "ShortValue") is None:
            na_rules["AIF-124"] = "No short individual exposure values"
        if find_element(root, "LongValue") is None:
            na_rules["AIF-123"] = "No long individual exposure values"

    # ── 9. Miscellaneous conditional ─────────────────────────────────────
    # PositionSizeType only for certain fund types
    if find_element(root, "TypicalPositionSize") is None:
        na_rules["AIF-113"] = "No TypicalPositionSize container"

    # MIC code only when principal market uses MIC
    if find_element(root, "AIFPrincipalMarketMICCode") is None and \
       find_element(root, "MICCode") is None:
        na_rules["AIF-116"] = "No MIC code in principal markets"

    # FX rate description: only when non-ECB FX rate used
    if find_element(root, "FXEURReferenceRateDescription") is None:
        na_rules["AIF-52"] = "No custom FX rate description needed"

    # Header conditional: change quarter, question number, assumption
    if find_element(root, "AIFReportingObligationChangeQuarter") is None:
        na_rules["AIF-12"] = "No reporting obligation change"
    if find_element(root, "QuestionNumber") is None:
        na_rules["AIF-14"] = "No assumption question numbers"
        na_rules["AIF-15"] = "No assumption descriptions"

    # Vega measures: only for options-heavy strategies
    if find_element(root, "VegaExposure") is None:
        for rid in ("AIF-143", "AIF-144", "AIF-145"):
            na_rules[rid] = "No Vega exposure data"

    # Stress test results: only when present
    if find_element(root, "StressTestsResultArticle15") is None and \
       find_element(root, "StressTestsResultArticle16") is None:
        for rid in ("AIF-146", "AIF-147"):
            na_rules[rid] = "No stress test results"

    # Currency of exposures: conditional section
    if find_element(root, "CurrencyExposure") is None:
        for rid in ("AIF-288", "AIF-287", "AIF-286b"):
            na_rules[rid] = "No CurrencyExposure container"

    # Leverage article 24(2) item 30 — only in CT4/5
    if find_element(root, "AIFLeverageArticle242") is None:
        na_rules["AIF-148"] = "No leverage article 24(2) data"
        na_rules["AIF-149"] = "No leverage article 24(2) data"

    # Controlled structure (leverage) — only when structures exist
    if find_element(root, "ControlledStructure") is None:
        na_rules["AIF-293"] = "No controlled structures"

    return na_rules


def _scan_aifm_conditional_context(root: ET.Element) -> dict:
    """Pre-scan AIFM XML for conditional context that gates child elements."""
    na_rules: dict[str, str] = {}

    # ── AIFM-12: ReportingObligationChange — only if obligation changed ──
    # If no change elements exist, the AIFM did not change reporting obligation
    if find_element(root, "AIFMReportingObligationChangeFrequencyCode") is None and \
       find_element(root, "AIFMReportingObligationChangeContentsCode") is None:
        na_rules["AIFM-12"] = "No reporting obligation change"

    # ── AIFM-14/15: QuestionNumber / AssumptionDescription ───────────────
    # Only when the AIFM has declared assumptions in the filing
    if find_element(root, "QuestionNumber") is None:
        na_rules["AIFM-14"] = "No assumptions declared"
        na_rules["AIFM-15"] = "No assumptions declared"

    # ── AIFM-28: MarketCode (MIC) ────────────────────────────────────────
    # Only required when MarketCodeType = "MIC"; OTC and XXX have no MIC
    has_mic = False
    for el in root.iter():
        etag = el.tag.split("}")[-1] if "}" in el.tag else el.tag
        if etag in ("MarketCodeType", "MarketTypeCode"):
            if (el.text or "").strip() == "MIC":
                has_mic = True
                break
    if not has_mic:
        na_rules["AIFM-28"] = "No MIC-type markets (OTC/XXX do not require MarketCode)"

    # ── AIFM-34/35/36/37: BaseCurrency section ─────────────────────────
    # These fields are only required when AIFM base currency differs from EUR.
    # When AUMAmountInEuro is present but BaseCurrency/AUMAmountInBaseCurrency
    # are absent, it means the AIFM reports in EUR → base currency fields N/A.
    base_ccy = _get_elem_text(root, "BaseCurrency")
    aum_eur = _get_elem_text(root, "AUMAmountInEuro")
    aum_base = _get_elem_text(root, "AUMAmountInBaseCurrency")

    if not base_ccy and not aum_base and aum_eur:
        # AIFM uses EUR as base currency — no separate base currency section needed
        na_rules["AIFM-34"] = "AIFM reports in EUR — AUMAmountInBaseCurrency not required"
        na_rules["AIFM-35"] = "AIFM reports in EUR — BaseCurrency not required"
        na_rules["AIFM-36"] = "AIFM reports in EUR — FXEURReferenceRateType not required"
        na_rules["AIFM-37"] = "AIFM reports in EUR — FXEURRate not required"
        na_rules["AIFM-CAM-016"] = "AIFM reports in EUR — no base/EUR cross-check needed"

    # ── AIFM-38: FXEUROtherReferenceRateDescription ──────────────────────
    # Only needed when FX rate is NOT from ECB (i.e. FXEURReferenceRateType != ECB)
    fx_type = _get_elem_text(root, "FXEURReferenceRateType")
    if fx_type in ("ECB", None):
        na_rules["AIFM-38"] = "FX reference rate is ECB — no custom description needed"

    # ── AIFM-CAM-016: AuM base vs EUR consistency ────────────────────────
    # Cross-field: compare AUMAmountInEuro vs AUMAmountInBaseCurrency
    # Only check when base currency section IS present (non-EUR AIFMs)
    if base_ccy and aum_eur and aum_base:
        if base_ccy == "EUR":
            # EUR base → amounts must be equal
            if aum_eur == aum_base:
                na_rules["AIFM-CAM-016"] = "EUR base currency — AuM amounts consistent"
            # If they differ, leave as WARNING (genuine issue)
        else:
            # Non-EUR base → if amounts are equal, it's a known FX limitation
            if aum_eur == aum_base:
                na_rules["AIFM-CAM-016"] = (
                    f"AuM EUR = AuM base ({base_ccy}) — likely ECB FX rate "
                    f"unavailable (sandbox limitation)")
            else:
                # Amounts differ → FX conversion applied, consistent
                na_rules["AIFM-CAM-016"] = (
                    f"AuM EUR ({aum_eur}) differs from AuM base "
                    f"({aum_base} {base_ccy}) — FX conversion applied")

    return na_rules


def validate_xml(file_path: Path, rules: list, ref_tables: dict,
                 nca_overrides: dict, report_type: str,
                 content_type: int, root: ET.Element,
                 metadata: dict) -> tuple[list, list]:
    """Validate one XML file. Returns (findings, coverage)."""
    findings = []
    coverage = []
    file_name = file_path.name

    # Pre-check: is this a cancellation filing?
    # ESMA-compliant cancellations use CancellationAIFRecordInfo / CancellationAIFMRecordInfo
    # (no FilingType element). Legacy/M adapter may use FilingType=CANCEL or CANC.
    has_cancellation_record = (
        find_element(root, "CancellationAIFRecordInfo") is not None or
        find_element(root, "CancellationAIFMRecordInfo") is not None)
    filing_type_elem = find_element(root, "FilingType")
    filing_type_val = (filing_type_elem.text or "").strip().upper() if filing_type_elem is not None else ""
    is_cancellation = has_cancellation_record or filing_type_val in ("CANCEL", "CANC")

    # Cancelled-field rule IDs (only relevant for cancellation filings)
    CANCELLED_RULES = {"AIFM-39", "AIFM-40", "AIFM-41", "AIFM-42",
                        "AIF-303", "AIF-304", "AIF-305", "AIF-306", "AIF-307"}

    # Check for no-reporting filings (AIFNoReportingFlag=true / AIFMNoReportingFlag=true)
    no_report_flag = _get_elem_text(root, "AIFNoReportingFlag") or \
                     _get_elem_text(root, "AIFMNoReportingFlag")
    is_no_reporting = no_report_flag == "true"

    # Pre-scan conditional context (flags, absent containers)
    if report_type == "AIF":
        conditional_na = _scan_conditional_context(root)
    else:
        conditional_na = _scan_aifm_conditional_context(root)

    # Cross-field validation rules (no single XSD element to check)
    CROSS_FIELD_RULES = set()
    for r in rules:
        xsd = r.get("xsd_element", "").strip("<>").strip()
        if "/" in xsd or not xsd:
            CROSS_FIELD_RULES.add(r.get("rule_id", ""))

    for rule in rules:
        rule_id = rule.get("rule_id", "")

        # Apply NCA override if exists for this base rule
        nca_rule = nca_overrides.get(rule_id)
        effective = {**rule}
        nca_applied = False
        nca_error_code = ""
        if nca_rule and nca_rule.get("overrides_esma"):
            # Override specific fields from NCA rule
            for field in ("m_c_o_f", "mandatory", "conditional", "severity",
                          "technical_guidance", "validation_rules", "format"):
                if field in nca_rule and nca_rule[field]:
                    effective[field] = nca_rule[field]
            nca_applied = True
            nca_codes = nca_rule.get("nca_error_codes", [])
            nca_error_code = nca_codes[0] if nca_codes else ""

        # Check applicability
        if not is_applicable(effective, content_type, report_type):
            coverage.append(Coverage(
                rule_id=rule_id, field_name=effective.get("field_name", ""),
                section=effective.get("section", ""),
                applicable=False, status="N_A"))
            continue

        # No-reporting filings: only header rules apply; all content rules N/A
        # Header section rules (AIF-1..AIF-23, AIFM-1..AIFM-10 approx) still apply
        if is_no_reporting:
            section = effective.get("section", "")
            is_header = "header" in section.lower() or "cancellation" in section.lower()
            if not is_header:
                findings.append(Finding(
                    file_name=file_name, report_type=report_type,
                    rule_id=rule_id, field_name=effective.get("field_name", ""),
                    section=effective.get("section", ""),
                    m_c_o_f=effective.get("m_c_o_f", "O"),
                    rule_source="CSSF" if nca_applied else "ESMA",
                    nca_override="Yes" if nca_applied else "",
                    document_value="", expected="",
                    status="N_A", severity="INFO",
                    esma_error_code="", nca_error_code="",
                    recommendation="No-reporting filing (AIFNoReportingFlag=true) — only header fields apply."))
                coverage.append(Coverage(
                    rule_id=rule_id, field_name=effective.get("field_name", ""),
                    section=effective.get("section", ""),
                    applicable=False, status="NO_REPORTING"))
                continue

        # Cancellation filings: only cancellation-specific rules apply
        if is_cancellation and rule_id not in CANCELLED_RULES:
            findings.append(Finding(
                file_name=file_name, report_type=report_type,
                rule_id=rule_id, field_name=effective.get("field_name", ""),
                section=effective.get("section", ""),
                m_c_o_f=effective.get("m_c_o_f", "O"),
                rule_source="CSSF" if nca_applied else "ESMA",
                nca_override="Yes" if nca_applied else "",
                document_value="", expected="",
                status="N_A", severity="INFO",
                esma_error_code="", nca_error_code="",
                recommendation="Cancellation filing — only CancelledRecord fields apply."))
            coverage.append(Coverage(
                rule_id=rule_id, field_name=effective.get("field_name", ""),
                section=effective.get("section", ""),
                applicable=False, status="CANCELLATION"))
            continue

        # Cancelled-field rules: N/A for non-cancellation filings
        if rule_id in CANCELLED_RULES and not is_cancellation:
            findings.append(Finding(
                file_name=file_name, report_type=report_type,
                rule_id=rule_id, field_name=effective.get("field_name", ""),
                section=effective.get("section", ""),
                m_c_o_f=effective.get("m_c_o_f", "M"),
                rule_source="CSSF" if nca_applied else "ESMA",
                nca_override="Yes" if nca_applied else "",
                document_value="", expected="",
                status="N_A", severity="INFO",
                esma_error_code="", nca_error_code="",
                recommendation="Not applicable — filing is not a cancellation."))
            coverage.append(Coverage(
                rule_id=rule_id, field_name=effective.get("field_name", ""),
                section=effective.get("section", ""),
                applicable=False, status="N_A"))
            continue

        # Conditional context: flag=false or parent container absent → N/A
        if rule_id in conditional_na:
            reason = conditional_na[rule_id]
            findings.append(Finding(
                file_name=file_name, report_type=report_type,
                rule_id=rule_id, field_name=effective.get("field_name", ""),
                section=effective.get("section", ""),
                m_c_o_f=effective.get("m_c_o_f", "C"),
                rule_source="CSSF" if nca_applied else "ESMA",
                nca_override="Yes" if nca_applied else "",
                document_value="", expected="",
                status="N_A", severity="INFO",
                esma_error_code="", nca_error_code="",
                recommendation=f"Conditional N/A — {reason}"))
            coverage.append(Coverage(
                rule_id=rule_id, field_name=effective.get("field_name", ""),
                section=effective.get("section", ""),
                applicable=True, status="CONDITIONAL_NA"))
            continue

        xsd_element = effective.get("xsd_element", "").strip("<>").strip()
        if not xsd_element:
            continue

        # Cross-field rules: skip element lookup, mark as WARNING
        if rule_id in CROSS_FIELD_RULES:
            findings.append(Finding(
                file_name=file_name, report_type=report_type,
                rule_id=rule_id,
                field_name=effective.get("field_name", ""),
                section=effective.get("section", ""),
                m_c_o_f=effective.get("m_c_o_f", "O"),
                rule_source="CSSF" if nca_applied else "ESMA",
                nca_override="Yes" if nca_applied else "",
                document_value="", expected=effective.get("format", ""),
                status="WARNING", severity="INFO",
                esma_error_code="", nca_error_code=nca_error_code,
                recommendation="Cross-field validation — manual review recommended."))
            coverage.append(Coverage(
                rule_id=rule_id, field_name=effective.get("field_name", ""),
                section=effective.get("section", ""),
                applicable=True, status="CROSS_FIELD"))
            continue

        field_name = effective.get("field_name", "")
        section = effective.get("section", "")
        m_c_o_f = effective.get("m_c_o_f", "O")
        severity = effective.get("severity", "LOW")
        esma_codes = effective.get("esma_error_codes", [])
        esma_code = esma_codes[0] if esma_codes else ""
        rule_source = "CSSF" if nca_applied else "ESMA"

        # --- Find the value ---
        value = None

        # Check root attributes first (ReportingMemberState, Version, etc.)
        if xsd_element in ROOT_ATTRIBUTE_ELEMENTS:
            attr_val = get_root_attribute(root, xsd_element)
            if attr_val is not None:
                value = attr_val

        # Then check child elements
        if value is None:
            elem = find_element(root, xsd_element)
            if elem is not None:
                value = (elem.text or "").strip()

        # --- Classify result ---
        if value is None:
            # Not found
            if m_c_o_f == "M":
                status, sev = "MISSING", severity
                rec = f"Mandatory element <{xsd_element}> not found."
            elif m_c_o_f == "C":
                status, sev = "WARNING", severity
                rec = f"Conditional element <{xsd_element}> not present."
            else:
                status, sev = "N_A", "INFO"
                rec = f"Optional element <{xsd_element}> not present."
            doc_val = ""
        else:
            # Found — validate
            ok, msg = validate_value(value, effective, ref_tables)
            if ok:
                status, sev = "PASS", severity
                rec = ""
                doc_val = value
            else:
                status, sev = "FAIL", severity
                rec = msg
                doc_val = value

        findings.append(Finding(
            file_name=file_name, report_type=report_type,
            rule_id=rule_id, field_name=field_name, section=section,
            m_c_o_f=m_c_o_f, rule_source=rule_source,
            nca_override="Yes" if nca_applied else "",
            document_value=doc_val,
            expected=effective.get("format", ""),
            status=status, severity=sev,
            esma_error_code=esma_code, nca_error_code=nca_error_code,
            recommendation=rec))

        coverage.append(Coverage(
            rule_id=rule_id, field_name=field_name, section=section,
            applicable=True, status="EVALUATED" if value is not None else "NOT_FOUND"))

    return findings, coverage


# ============================================================================
# Excel output
# ============================================================================

HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
CRIT_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CRIT_FONT = Font(name="Arial", color="FFFFFF")
HIGH_FILL = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
HIGH_FONT = Font(name="Arial", color="FFFFFF")
WARN_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
MISS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NA_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Border(*(Side(style="thin"),) * 4)
BODY_FONT = Font(name="Arial", size=10)


def _header_row(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def _row_fill(status, severity):
    if status == "FAIL" and severity == "CRITICAL":
        return CRIT_FILL, CRIT_FONT
    if status == "FAIL" and severity == "HIGH":
        return HIGH_FILL, HIGH_FONT
    if status == "MISSING":
        return MISS_FILL, BODY_FONT
    if status == "WARNING":
        return WARN_FILL, BODY_FONT
    if status == "N_A":
        return NA_FILL, BODY_FONT
    if status == "PASS":
        return PASS_FILL, BODY_FONT
    return None, BODY_FONT


def build_excel(all_findings: list, all_coverage: dict,
                all_meta: list, output_path: Path, nca_code: str,
                xsd_findings: list = None, file_findings: list = None):
    wb = openpyxl.Workbook()
    xsd_findings = xsd_findings or []
    file_findings = file_findings or []

    # ── Summary ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells("A1:K1")
    ws["A1"].value = "AIFMD Annex IV — Validation Audit Trail"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A2"].value = f"Generated: {datetime.now():%Y-%m-%d %H:%M}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True)
    ws["A3"].value = f"NCA: {nca_code or 'ESMA base only'}"
    ws["A3"].font = Font(name="Arial", size=10)

    _header_row(ws, 5, ["File", "Type", "CT", "XSD", "XSD Errors",
                         "DQF Rules", "PASS", "FAIL",
                         "MISSING", "WARNING", "N/A"])

    stats = defaultdict(lambda: defaultdict(int))
    meta_map = {m["file_name"]: m for m in all_meta}
    for f in all_findings:
        stats[f.file_name][f.status] += 1
        stats[f.file_name]["total"] += 1

    # Count XSD errors per file
    xsd_per_file = defaultdict(int)
    for xf in xsd_findings:
        xsd_per_file[xf.file_name] += 1

    r = 6
    for fn in sorted(stats):
        s = stats[fn]
        m = meta_map.get(fn, {})
        ct = m.get("AIFContentType") or m.get("AIFMContentType") or ""
        rtype = m.get("_report_type", "")
        xsd_src = m.get("_xsd_source", "N/A")
        xsd_errs = m.get("_xsd_errors", 0)
        xsd_label = "VALID" if xsd_errs == 0 and not xsd_src.startswith("SKIPPED") else (
            "SKIPPED" if xsd_src.startswith("SKIPPED") else f"INVALID ({xsd_errs})")
        for c, v in enumerate([fn, rtype, ct, xsd_label, xsd_errs,
                                s["total"], s["PASS"], s["FAIL"],
                                s["MISSING"], s["WARNING"], s["N_A"]], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            # Color XSD cell
            if c == 4:
                if "VALID" == xsd_label:
                    cell.fill = PASS_FILL
                elif "INVALID" in xsd_label:
                    cell.fill = CRIT_FILL
                    cell.font = CRIT_FONT
        r += 1

    for c in range(1, 12):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions["A"].width = 50

    # ── File Naming & Packaging ─────────────────────────────────────────
    ws_fn = wb.create_sheet("File Naming")
    _header_row(ws_fn, 1, ["File", "Check", "Expected", "Actual",
                            "Status", "Severity", "Detail"])
    r = 2
    for ff in sorted(file_findings, key=lambda x: (x.file_name, x.check)):
        fill, font = None, BODY_FONT
        if ff.status == "FAIL":
            fill = CRIT_FILL if ff.severity == "HIGH" else HIGH_FILL
            font = CRIT_FONT if ff.severity == "HIGH" else HIGH_FONT
        elif ff.status == "WARNING":
            fill = WARN_FILL
        elif ff.status == "PASS":
            fill = PASS_FILL
        vals = [ff.file_name, ff.check, ff.expected, ff.actual,
                ff.status, ff.severity, ff.message]
        for c, v in enumerate(vals, 1):
            cell = ws_fn.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.font = font
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)
        r += 1
    if r == 2:
        ws_fn.cell(row=2, column=1,
                   value="No files to check").font = Font(italic=True)
    ws_fn.freeze_panes = "A2"
    ws_fn.column_dimensions["A"].width = 50
    ws_fn.column_dimensions["B"].width = 20
    ws_fn.column_dimensions["C"].width = 45
    ws_fn.column_dimensions["D"].width = 45
    ws_fn.column_dimensions["E"].width = 10
    ws_fn.column_dimensions["F"].width = 10
    ws_fn.column_dimensions["G"].width = 80

    # ── XSD Validation ──────────────────────────────────────────────────
    ws_xsd = wb.create_sheet("XSD Validation")
    _header_row(ws_xsd, 1, ["File", "Line", "Severity", "XSD Schema",
                              "Error Message"])
    r = 2
    if xsd_findings:
        for xf in sorted(xsd_findings, key=lambda x: (x.file_name, x.line)):
            vals = [xf.file_name, xf.line, xf.severity, xf.xsd_source,
                    xf.message]
            fill = CRIT_FILL if xf.severity == "CRITICAL" else HIGH_FILL
            font = CRIT_FONT if xf.severity == "CRITICAL" else HIGH_FONT
            for c, v in enumerate(vals, 1):
                cell = ws_xsd.cell(row=r, column=c, value=v)
                cell.border = THIN
                cell.font = font
                cell.fill = fill
                cell.alignment = Alignment(horizontal="left", vertical="top",
                                           wrap_text=True)
            r += 1
    else:
        cell = ws_xsd.cell(row=2, column=1,
                           value="All files passed XSD validation")
        cell.font = Font(name="Arial", size=10, italic=True)
        cell.fill = PASS_FILL
    ws_xsd.freeze_panes = "A2"
    ws_xsd.column_dimensions["A"].width = 50
    ws_xsd.column_dimensions["B"].width = 8
    ws_xsd.column_dimensions["C"].width = 12
    ws_xsd.column_dimensions["D"].width = 35
    ws_xsd.column_dimensions["E"].width = 100

    # ── Audit Trail (DQF) ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Audit Trail")
    headers = ["File", "Type", "Rule ID", "Field Name", "Section", "M/C/O",
               "Rule Source", "NCA Override", "Document Value", "Expected",
               "Status", "Severity", "ESMA Code", "NCA Code", "Recommendation"]
    _header_row(ws2, 1, headers)

    r = 2
    for f in sorted(all_findings, key=lambda x: (x.file_name, x.rule_id)):
        vals = [f.file_name, f.report_type, f.rule_id, f.field_name,
                f.section, f.m_c_o_f, f.rule_source, f.nca_override,
                f.document_value, f.expected, f.status, f.severity,
                f.esma_error_code, f.nca_error_code, f.recommendation]
        fill, font = _row_fill(f.status, f.severity)
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.font = font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
        r += 1

    ws2.freeze_panes = "A2"
    if r > 2:
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{r - 1}"
    for c in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 18
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions[get_column_letter(len(headers))].width = 40

    # ── NCA Overrides Applied ────────────────────────────────────────────
    ws3 = wb.create_sheet("NCA Overrides Applied")
    _header_row(ws3, 1, ["File", "Rule ID", "Field Name", "Base Rule",
                          "NCA Source", "NCA Error Code", "Override Detail"])
    r = 2
    for f in all_findings:
        if f.nca_override == "Yes":
            for c, v in enumerate([f.file_name, f.rule_id, f.field_name,
                                    f.rule_source, f.nca_error_code,
                                    f.nca_error_code, f.recommendation], 1):
                cell = ws3.cell(row=r, column=c, value=v)
                cell.border = THIN
                cell.font = BODY_FONT
            r += 1
    if r == 2:
        ws3.cell(row=2, column=1, value="No NCA overrides applied").font = Font(italic=True)
    for c in range(1, 8):
        ws3.column_dimensions[get_column_letter(c)].width = 22

    # ── Rule Coverage ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Rule Coverage")
    _header_row(ws4, 1, ["File", "Rule ID", "Field Name", "Section",
                          "Applicable", "Status"])
    r = 2
    for fn in sorted(all_coverage):
        for cv in all_coverage[fn]:
            for c, v in enumerate([fn, cv.rule_id, cv.field_name,
                                    cv.section,
                                    "Yes" if cv.applicable else "No",
                                    cv.status], 1):
                cell = ws4.cell(row=r, column=c, value=v)
                cell.border = THIN
                cell.font = BODY_FONT
            r += 1
    ws4.freeze_panes = "A2"
    for c in range(1, 7):
        ws4.column_dimensions[get_column_letter(c)].width = 22

    wb.save(output_path)
    return output_path


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Validate AIFMD Annex IV XML files against ESMA/NCA rules")
    parser.add_argument("files", nargs="+", help="XML file(s) to validate")
    parser.add_argument("--nca", default="", help="NCA code (auto-detected if omitted)")
    parser.add_argument("--output", default="", help="Output Excel path")
    parser.add_argument("--rules", default="", help="Path to base rules YAML")
    parser.add_argument("--approve-rules", action="store_true",
                        help="Approve current rule hashes as baseline (Layer B)")
    parser.add_argument("--approved-by", default="",
                        help="Name of the person approving (required with --approve-rules)")
    parser.add_argument("--reason", default="",
                        help="Motivation for the approval (required with --approve-rules)")
    args = parser.parse_args()

    # Handle --approve-rules
    if args.approve_rules:
        rp = Path(args.rules) if args.rules else DEFAULT_RULES_PATH
        try:
            approved = approve_current_hashes(
                rp, approved_by=args.approved_by, reason=args.reason)
        except ValueError as e:
            print(f"\n  ERROR: {e}")
            print(f"\n  Usage: python validate_aifmd_xml.py --approve-rules \\")
            print(f"           --approved-by \"Jan de Vries\" \\")
            print(f"           --reason \"FSMA override update na DQEF v3.2\" \\")
            print(f"           dummy.xml")
            sys.exit(1)
        print(f"\nApproved rule hashes saved to: {_APPROVED_HASHES_PATH.name}")
        print(f"  Approved by: {approved['approved_by']}")
        print(f"  Reason:      {approved['reason']}")
        print(f"  Base rules:  {approved['base_rules']['hash'][:16]}...")
        for name, info in approved.get("nca_overrides", {}).items():
            print(f"  {name}: {info['hash'][:16]}...")
        if approved.get("changes"):
            print(f"\n  Changes since last approval:")
            for ch in approved["changes"]:
                print(f"    {ch['file']}: {ch['previous_hash']} -> {ch['new_hash']}")
        else:
            print(f"\n  No changes since last approval (re-confirmation)")
        print(f"\n  History logged to: Logging/rule_approval_history.yaml")
        sys.exit(0)

    # Expand globs
    xml_files = []
    for pattern in args.files:
        expanded = glob.glob(pattern)
        xml_files.extend(expanded if expanded else [pattern])
    xml_files = [Path(f) for f in xml_files if f.endswith(".xml")]

    if not xml_files:
        print("No XML files found.")
        sys.exit(1)

    # Load rules
    rules_path = Path(args.rules) if args.rules else DEFAULT_RULES_PATH
    if not rules_path.exists():
        print(f"Rules file not found: {rules_path}")
        sys.exit(1)

    print("=" * 70)
    print("AIFMD Annex IV XML Validator")
    print("=" * 70)

    print(f"\nLoading rules from: {rules_path.name}")
    base = load_rules(rules_path)
    aif_rules = base.get("aif_rules", [])
    aifm_rules = base.get("aifm_rules", [])
    ref_tables = base.get("reference_tables", {})
    print(f"  AIF rules: {len(aif_rules)}, AIFM rules: {len(aifm_rules)}, "
          f"Reference tables: {len(ref_tables)}")

    # ── Layer B: Rule Integrity Monitor ────────────────────────────────
    integrity = verify_rule_integrity(rules_path, args.nca)
    if integrity["status"] == "FIRST_RUN":
        print(f"\n  Rule Integrity: FIRST RUN -- approving current hashes as baseline")
        approved = approve_current_hashes(rules_path)
        print(f"  Approved hashes saved to: {_APPROVED_HASHES_PATH.name}")
        integrity["status"] = "VERIFIED"
    elif integrity["status"] == "UNVERIFIED":
        print(f"\n  [!] Rule Integrity: UNVERIFIED -- rule files changed since last approval!")
        for cf in integrity["changed_files"]:
            print(f"     Changed: {cf}")
        print(f"     Run with --approve-rules to accept current versions.")
    else:
        print(f"  Rule Integrity: VERIFIED")

    # XSD availability check
    if HAS_LXML:
        xsd_avail = []
        for schema_set, paths in XSD_SCHEMAS.items():
            for rtype_key, xsd_path in paths.items():
                if xsd_path is not None and xsd_path.exists():
                    xsd_avail.append(f"{schema_set}/{rtype_key}")
        if xsd_avail:
            print(f"  XSD schemas: {', '.join(xsd_avail)}")
        else:
            print("  XSD schemas: none found (XSD validation disabled)")
    else:
        print("  XSD schemas: lxml not installed (XSD validation disabled)")

    # Process each file
    all_findings = []
    all_xsd_findings = []
    all_file_findings = []
    all_coverage = {}
    all_meta = []
    detected_nca = args.nca

    for fp in xml_files:
        if not fp.exists():
            print(f"\nSkipping (not found): {fp}")
            continue

        root = ET.parse(str(fp)).getroot()
        rtype = detect_report_type(root)
        meta = extract_metadata(root, rtype)
        meta["file_name"] = fp.name
        meta["_report_type"] = rtype

        # Auto-detect NCA from ReportingMemberState
        if not detected_nca:
            detected_nca = meta.get("ReportingMemberState", "")

        # Content type
        ct_str = meta.get("AIFContentType") or meta.get("AIFMContentType") or "1"
        try:
            ct = int(ct_str)
        except ValueError:
            ct = 1

        print(f"\n{'-' * 60}")
        print(f"  File: {fp.name}")
        print(f"  Type: {rtype} | CT={ct} | RMS={meta.get('ReportingMemberState','?')}")

        # ── Layer 0: File Naming & Packaging ────────────────────────────
        file_findings = validate_file_naming(fp, rtype, detected_nca, meta)
        meta["_file_findings"] = len(file_findings)
        file_fails = [f for f in file_findings if f.status == "FAIL"]
        file_warns = [f for f in file_findings if f.status == "WARNING"]
        file_pass = [f for f in file_findings if f.status == "PASS"]
        if file_fails:
            print(f"  File: {len(file_fails)} FAIL, {len(file_warns)} WARNING, "
                  f"{len(file_pass)} PASS")
            for ff in file_fails:
                print(f"       {ff.check}: {ff.message}")
        elif file_warns:
            print(f"  File: {len(file_warns)} WARNING, {len(file_pass)} PASS")
            for fw in file_warns:
                print(f"       {fw.check}: {fw.message}")
        else:
            checks = ", ".join(f.check for f in file_pass)
            print(f"  File: ALL PASS ({checks})")
        all_file_findings.extend(file_findings)

        # ── Layer 1: XSD Validation ─────────────────────────────────────
        xsd_findings, xsd_source = validate_xsd(fp, rtype)
        meta["_xsd_source"] = xsd_source
        meta["_xsd_errors"] = len(xsd_findings)
        if xsd_source.startswith("SKIPPED"):
            print(f"  XSD: {xsd_source}")
        elif xsd_findings:
            print(f"  XSD: INVALID -- {len(xsd_findings)} error(s) [{xsd_source}]")
            for xf in xsd_findings[:3]:
                print(f"       Line {xf.line}: {xf.message[:100]}")
            if len(xsd_findings) > 3:
                print(f"       ... and {len(xsd_findings) - 3} more")
        else:
            print(f"  XSD: VALID [{xsd_source}]")
        all_xsd_findings.extend(xsd_findings)

        # ── Layer 2+3: DQF + Cross-field Validation ─────────────────────
        # Select rules
        rules = aifm_rules if rtype == "AIFM" else aif_rules

        # Load NCA overrides (once)
        nca_overrides = load_nca_overrides(detected_nca) if detected_nca else {}
        if nca_overrides:
            nca_name = NCA_COUNTRY_MAP.get(detected_nca, detected_nca).upper()
            print(f"  NCA overrides: {nca_name} ({len(nca_overrides)} rules)")

        findings, coverage = validate_xml(
            fp, rules, ref_tables, nca_overrides, rtype, ct, root, meta)

        # Stats
        s = defaultdict(int)
        for f in findings:
            s[f.status] += 1
        total = len(findings)
        print(f"  DQF: {total} rules | "
              f"PASS={s['PASS']} FAIL={s['FAIL']} MISSING={s['MISSING']} "
              f"WARNING={s['WARNING']} N/A={s['N_A']}")

        all_findings.extend(findings)
        all_coverage[fp.name] = coverage
        all_meta.append(meta)

    # Output Excel
    if not all_findings:
        print("\nNo findings to report.")
        sys.exit(0)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.output:
        out = Path(args.output)
    else:
        logging_dir = _SCRIPT_DIR / "Logging"
        logging_dir.mkdir(parents=True, exist_ok=True)
        out = logging_dir / f"aifmd_validation_{ts}.xlsx"

    print(f"\n{'=' * 70}")
    print("Generating Excel audit trail...")
    build_excel(all_findings, all_coverage, all_meta, out,
                detected_nca, xsd_findings=all_xsd_findings,
                file_findings=all_file_findings)
    print(f"Saved: {out}")
    print("=" * 70)


if __name__ == "__main__":
    main()

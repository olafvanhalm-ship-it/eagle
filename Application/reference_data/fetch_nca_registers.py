"""Unified NCA register fetcher — downloads and stores AIFM/AIF/SUB_AIF data.

Sources (in priority order):
  - AFM  (NL): Authorised + Registered (light) managers and funds  [HIGH priority]
  - CSSF (LU): Authorised + Registered + NPPR managers, funds, sub-funds  [HIGH priority]
  - ESMA (EU): Central register of AIFMs and funds (Solr API)  [LOW priority — gap-fill]

Data model (3 tables):
  - platform.nca_aifm            — AIFM entities (manager-specific fields)
  - platform.nca_aif             — AIF + SUB_AIF entities (fund-specific fields)
  - platform.nca_registrations   — NCA registration details (shared structure)

Merge strategy:
  - fetch-all runs local NCA sources FIRST (AFM, CSSF), then ESMA.
  - Local NCA sources are the AUTHORITATIVE entity list for their countries.
    They overwrite entity names (names can change over time) and determine
    which entities exist.
  - ESMA enriches existing entities (gap-fill NULLs) but NEVER inserts new
    entities or removes entities for countries covered by a local NCA fetcher.
    For other EU countries ESMA is the primary source and does insert.
  - Matching: nca_entity_code → source_entity_id → LEI → aifm_clean_name.
  - Designed to run weekly to keep data current.

Usage:
    python fetch_nca_registers.py fetch-afm
    python fetch_nca_registers.py fetch-cssf
    python fetch_nca_registers.py fetch-esma
    python fetch_nca_registers.py fetch-all
    python fetch_nca_registers.py enrich-gleif
    python fetch_nca_registers.py enrich-lei
    python fetch_nca_registers.py report

Architecture: eagle_software_architecture.md §6.3 — platform schema
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from urllib.request import Request, urlopen

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from reference_data.config import PG_HOST, PG_PORT, PG_DBNAME, PG_USER, PG_PASSWORD
from shared.clean_name import clean_name

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)


# ── Data classes ─────────────────────────────────────────────────────

@dataclass
class AIFMRecord:
    """An AIFM entity from any source."""
    aifm_name: str
    aifm_clean_name: str = ""
    lei: Optional[str] = None
    home_country: Optional[str] = None   # ISO-2 home member state
    address: Optional[str] = None        # best available address
    address_source: Optional[str] = None # KVK, MANUAL, NCA, ESMA, GLEIF
    registered_as: Optional[str] = None  # GLEIF registration identifier

    # NCA registration info
    nca_code: str = ""
    nca_entity_code: Optional[str] = None
    auth_status: str = "AUTHORISED"      # AUTHORISED, REGISTERED, NPPR, INACTIVE, WITHDRAWN
    auth_date: Optional[date] = None     # merged: auth_date + registration_date → single date
    withdrawal_date: Optional[date] = None
    source: str = ""
    source_entity_id: Optional[str] = None

    def __post_init__(self):
        if not self.aifm_clean_name:
            self.aifm_clean_name = clean_name(self.aifm_name)


@dataclass
class AIFRecord:
    """An AIF or SUB_AIF entity from any source."""
    aif_type: str             # AIF or SUB_AIF
    aif_name: str
    aif_clean_name: str = ""
    lei: Optional[str] = None
    isin: Optional[str] = None
    home_country: Optional[str] = None   # ISO-2 home member state
    custodian: Optional[str] = None
    fund_strategy: Optional[str] = None  # was fund_category
    address: Optional[str] = None
    address_source: Optional[str] = None

    # Hierarchy hints (resolved during upsert)
    managing_aifm_name: Optional[str] = None
    parent_aif_name: Optional[str] = None

    # NCA registration info
    nca_code: str = ""
    nca_entity_code: Optional[str] = None
    auth_status: str = "AUTHORISED"
    auth_date: Optional[date] = None
    withdrawal_date: Optional[date] = None
    source: str = ""
    source_entity_id: Optional[str] = None

    def __post_init__(self):
        if not self.aif_clean_name:
            self.aif_clean_name = clean_name(self.aif_name)


# ── Address priority ────────────────────────────────────────────────

# Higher number = higher priority (overwrites lower)
ADDRESS_PRIORITY = {"GLEIF": 1, "AIFM": 2, "ESMA": 3, "NCA": 4, "MANUAL": 5, "KVK": 6}

def _should_update_address(new_source: str, existing_source: str | None) -> bool:
    """Return True if new_source has higher or equal priority than existing_source."""
    if not existing_source:
        return True
    return ADDRESS_PRIORITY.get(new_source, 0) >= ADDRESS_PRIORITY.get(existing_source, 0)


# ── Database operations ──────────────────────────────────────────────

class NCARegisterStore:
    """Store for NCA register data in platform schema (split model).

    Merge strategy:
      - Local NCA sources (AFM, CSSF) are HIGH priority — they overwrite names.
      - ESMA is LOW priority — it only fills NULL fields (gap-fill).
      - Matching order: nca_entity_code → source_entity_id → LEI → aifm_clean_name.

    enrich_only mode:
      When enrich_only=True, upsert methods only UPDATE existing entities (gap-fill).
      If no match is found, the record is skipped (returns None).
    """

    SOURCE_PRIORITY = {"ESMA": 1, "AFM": 10, "CSSF": 10}
    LOCAL_NCA_COUNTRIES = {"NL", "LU"}

    def __init__(self, conn):
        self._conn = conn
        self._conn.autocommit = False

    # ── Lineage tracking ───────────────────────────────────────────

    @staticmethod
    def _build_lineage_entry(source: str, fields: dict, today: str = None) -> str:
        if today is None:
            today = datetime.utcnow().strftime("%Y-%m-%d")
        contributed = [k for k, v in fields.items() if v is not None]
        if not contributed:
            return ""
        return f"{source}:{','.join(contributed)}@{today}"

    @staticmethod
    def _append_lineage(existing: str | None, new_entry: str) -> str:
        if not new_entry:
            return existing or ""
        if existing:
            return f"{existing}; {new_entry}"
        return new_entry

    def _is_high_priority(self, source: str) -> bool:
        return self.SOURCE_PRIORITY.get(source, 5) >= 10

    # ── AIFM upsert ────────────────────────────────────────────────

    def upsert_aifm(self, rec: AIFMRecord, enrich_only: bool = False) -> Optional[str]:
        """Upsert an AIFM and its NCA registration. Returns aifm_id."""
        cur = self._conn.cursor()
        aifm_id = None
        high_prio = self._is_high_priority(rec.source)

        # Match 1: nca_entity_code
        if rec.nca_entity_code:
            cur.execute("""
                SELECT r.entity_id FROM platform.nca_registrations r
                WHERE r.nca_entity_code = %s AND r.nca_code = %s AND r.entity_type = 'AIFM'
                LIMIT 1
            """, (rec.nca_entity_code, rec.nca_code))
            row = cur.fetchone()
            if row:
                aifm_id = row[0]

        # Match 2: source_entity_id
        if not aifm_id and rec.source_entity_id:
            cur.execute("""
                SELECT r.entity_id FROM platform.nca_registrations r
                WHERE r.source_entity_id = %s AND r.source = %s AND r.entity_type = 'AIFM'
                LIMIT 1
            """, (rec.source_entity_id, rec.source))
            row = cur.fetchone()
            if row:
                aifm_id = row[0]

        # Match 3: LEI
        if not aifm_id and rec.lei:
            cur.execute("SELECT aifm_id FROM platform.nca_aifm WHERE lei = %s LIMIT 1", (rec.lei,))
            row = cur.fetchone()
            if row:
                aifm_id = row[0]

        # Match 4: aifm_clean_name + country
        if not aifm_id:
            cur.execute("""
                SELECT aifm_id FROM platform.nca_aifm
                WHERE aifm_clean_name = %s
                  AND (home_country = %s OR home_country IS NULL OR %s IS NULL)
                LIMIT 1
            """, (rec.aifm_clean_name, rec.home_country, rec.home_country))
            row = cur.fetchone()
            if row:
                aifm_id = row[0]

        now = datetime.utcnow()
        today = now.strftime("%Y-%m-%d")

        contributed = {
            "aifm_name": rec.aifm_name, "lei": rec.lei,
            "home_country": rec.home_country, "address": rec.address,
        }
        lineage_entry = self._build_lineage_entry(rec.source, contributed, today)

        if aifm_id:
            cur.execute("""
                SELECT lineage, address_source FROM platform.nca_aifm WHERE aifm_id = %s
            """, (aifm_id,))
            current = cur.fetchone()
            current_lineage = current[0]
            current_addr_source = current[1]
            new_lineage = self._append_lineage(current_lineage, lineage_entry)

            update_addr = rec.address and _should_update_address(
                rec.address_source or rec.source, current_addr_source
            )

            if high_prio:
                cur.execute("""
                    UPDATE platform.nca_aifm SET
                        aifm_name = CASE WHEN NULLIF(%s, '') IS NOT NULL THEN %s ELSE aifm_name END,
                        aifm_clean_name = CASE WHEN NULLIF(%s, '') IS NOT NULL THEN %s ELSE aifm_clean_name END,
                        lei = COALESCE(%s, lei),
                        home_country = COALESCE(%s, home_country),
                        address = CASE WHEN %s THEN COALESCE(%s, address) ELSE address END,
                        address_source = CASE WHEN %s THEN COALESCE(%s, address_source) ELSE address_source END,
                        registered_as = COALESCE(%s, registered_as),
                        lineage = %s,
                        last_seen_at = %s, deleted_at = NULL, last_updated = %s
                    WHERE aifm_id = %s
                """, (
                    rec.aifm_name, rec.aifm_name,
                    rec.aifm_clean_name, rec.aifm_clean_name,
                    rec.lei, rec.home_country,
                    update_addr, rec.address,
                    update_addr, rec.address_source or rec.source,
                    rec.registered_as,
                    new_lineage, now, now, aifm_id,
                ))
            else:
                # Low-priority source (ESMA): only fill NULLs, never overwrite.
                # IMPORTANT: Do NOT fill home_country for entities with NPPR
                # registrations — NPPR means the entity is foreign, and ESMA's
                # ae_homeMemberState for passport notifications is the host
                # country (where notification was filed), not the actual domicile.
                has_nppr = False
                if rec.home_country:
                    cur.execute("""
                        SELECT 1 FROM platform.nca_registrations
                        WHERE entity_id = %s AND entity_type = 'AIFM'
                          AND auth_status = 'NPPR' AND deleted_at IS NULL
                        LIMIT 1
                    """, (aifm_id,))
                    has_nppr = cur.fetchone() is not None

                safe_country = None if has_nppr else rec.home_country

                cur.execute("""
                    UPDATE platform.nca_aifm SET
                        aifm_name = COALESCE(NULLIF(aifm_name, ''), %s),
                        aifm_clean_name = COALESCE(NULLIF(aifm_clean_name, ''), %s),
                        lei = COALESCE(lei, %s),
                        home_country = COALESCE(home_country, %s),
                        address = CASE WHEN %s THEN COALESCE(%s, address) ELSE COALESCE(address, %s) END,
                        address_source = CASE WHEN %s THEN COALESCE(%s, address_source) ELSE address_source END,
                        registered_as = COALESCE(registered_as, %s),
                        lineage = %s,
                        last_seen_at = %s, deleted_at = NULL, last_updated = %s
                    WHERE aifm_id = %s
                """, (
                    rec.aifm_name, rec.aifm_clean_name,
                    rec.lei, safe_country,
                    update_addr, rec.address, rec.address,
                    update_addr, rec.address_source or rec.source,
                    rec.registered_as,
                    new_lineage, now, now, aifm_id,
                ))
        else:
            if enrich_only:
                cur.close()
                return None
            cur.execute("""
                INSERT INTO platform.nca_aifm (
                    aifm_name, aifm_clean_name, lei, home_country,
                    address, address_source, registered_as,
                    lineage,
                    first_seen_at, last_seen_at, created_at, last_updated
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING aifm_id
            """, (
                rec.aifm_name, rec.aifm_clean_name, rec.lei, rec.home_country,
                rec.address, rec.address_source or rec.source if rec.address else None,
                rec.registered_as,
                lineage_entry,
                now, now, now, now,
            ))
            aifm_id = cur.fetchone()[0]

        if rec.nca_code and aifm_id:
            self._upsert_registration(cur, "AIFM", aifm_id, rec, now)

        cur.close()
        return aifm_id

    # ── AIF upsert ──────────────────────────────────────────────────

    def upsert_aif(self, rec: AIFRecord, enrich_only: bool = False) -> Optional[str]:
        """Upsert an AIF or SUB_AIF and its NCA registration. Returns aif_id."""
        cur = self._conn.cursor()
        aif_id = None
        high_prio = self._is_high_priority(rec.source)

        # Match 1: nca_entity_code
        if rec.nca_entity_code:
            cur.execute("""
                SELECT r.entity_id FROM platform.nca_registrations r
                WHERE r.nca_entity_code = %s AND r.nca_code = %s AND r.entity_type = %s
                LIMIT 1
            """, (rec.nca_entity_code, rec.nca_code, rec.aif_type))
            row = cur.fetchone()
            if row:
                aif_id = row[0]

        # Match 2: source_entity_id
        if not aif_id and rec.source_entity_id:
            cur.execute("""
                SELECT r.entity_id FROM platform.nca_registrations r
                WHERE r.source_entity_id = %s AND r.source = %s AND r.entity_type = %s
                LIMIT 1
            """, (rec.source_entity_id, rec.source, rec.aif_type))
            row = cur.fetchone()
            if row:
                aif_id = row[0]

        # Match 3: LEI
        if not aif_id and rec.lei:
            cur.execute("""
                SELECT aif_id FROM platform.nca_aif
                WHERE lei = %s AND aif_type = %s LIMIT 1
            """, (rec.lei, rec.aif_type))
            row = cur.fetchone()
            if row:
                aif_id = row[0]

        # Match 4: aif_clean_name + aif_type + country
        if not aif_id:
            cur.execute("""
                SELECT aif_id FROM platform.nca_aif
                WHERE aif_clean_name = %s AND aif_type = %s
                  AND (home_country = %s OR home_country IS NULL OR %s IS NULL)
                LIMIT 1
            """, (rec.aif_clean_name, rec.aif_type, rec.home_country, rec.home_country))
            row = cur.fetchone()
            if row:
                aif_id = row[0]

        now = datetime.utcnow()
        today = now.strftime("%Y-%m-%d")

        contributed = {
            "aif_name": rec.aif_name, "lei": rec.lei, "isin": rec.isin,
            "home_country": rec.home_country, "custodian": rec.custodian,
            "fund_strategy": rec.fund_strategy,
        }
        lineage_entry = self._build_lineage_entry(rec.source, contributed, today)

        if aif_id:
            cur.execute("""
                SELECT lineage, address_source FROM platform.nca_aif WHERE aif_id = %s
            """, (aif_id,))
            current = cur.fetchone()
            current_lineage = current[0]
            current_addr_source = current[1]
            new_lineage = self._append_lineage(current_lineage, lineage_entry)

            update_addr = rec.address and _should_update_address(
                rec.address_source or rec.source, current_addr_source
            )

            if high_prio:
                cur.execute("""
                    UPDATE platform.nca_aif SET
                        aif_name     = CASE WHEN NULLIF(%s, '') IS NOT NULL THEN %s ELSE aif_name END,
                        aif_clean_name = CASE WHEN NULLIF(%s, '') IS NOT NULL THEN %s ELSE aif_clean_name END,
                        lei = COALESCE(%s, lei),
                        isin = COALESCE(%s, isin),
                        home_country = COALESCE(%s, home_country),
                        custodian = COALESCE(%s, custodian),
                        fund_strategy = COALESCE(%s, fund_strategy),
                        address = CASE WHEN %s THEN COALESCE(%s, address) ELSE address END,
                        address_source = CASE WHEN %s THEN COALESCE(%s, address_source) ELSE address_source END,
                        lineage = %s,
                        last_seen_at = %s, deleted_at = NULL, last_updated = %s
                    WHERE aif_id = %s
                """, (
                    rec.aif_name, rec.aif_name,
                    rec.aif_clean_name, rec.aif_clean_name,
                    rec.lei, rec.isin, rec.home_country,
                    rec.custodian, rec.fund_strategy,
                    update_addr, rec.address,
                    update_addr, rec.address_source or rec.source,
                    new_lineage, now, now, aif_id,
                ))
            else:
                # Low-priority source (ESMA): only fill NULLs, never overwrite.
                # IMPORTANT: Do NOT fill home_country for entities with NPPR
                # registrations — NPPR means the entity is foreign, and ESMA's
                # ae_homeMemberState for passport notifications is the host
                # country (where notification was filed), not the actual domicile.
                has_nppr = False
                if rec.home_country:
                    cur.execute("""
                        SELECT 1 FROM platform.nca_registrations
                        WHERE entity_id = %s AND entity_type IN ('AIF', 'SUB_AIF')
                          AND auth_status = 'NPPR' AND deleted_at IS NULL
                        LIMIT 1
                    """, (aif_id,))
                    has_nppr = cur.fetchone() is not None

                safe_country = None if has_nppr else rec.home_country

                cur.execute("""
                    UPDATE platform.nca_aif SET
                        aif_name     = COALESCE(NULLIF(aif_name, ''), %s),
                        aif_clean_name = COALESCE(NULLIF(aif_clean_name, ''), %s),
                        lei = COALESCE(lei, %s),
                        isin = COALESCE(isin, %s),
                        home_country = COALESCE(home_country, %s),
                        custodian = COALESCE(custodian, %s),
                        fund_strategy = COALESCE(fund_strategy, %s),
                        address = CASE WHEN %s THEN COALESCE(%s, address) ELSE COALESCE(address, %s) END,
                        address_source = CASE WHEN %s THEN COALESCE(%s, address_source) ELSE address_source END,
                        lineage = %s,
                        last_seen_at = %s, deleted_at = NULL, last_updated = %s
                    WHERE aif_id = %s
                """, (
                    rec.aif_name, rec.aif_clean_name,
                    rec.lei, rec.isin, safe_country,
                    rec.custodian, rec.fund_strategy,
                    update_addr, rec.address, rec.address,
                    update_addr, rec.address_source or rec.source,
                    new_lineage, now, now, aif_id,
                ))
        else:
            if enrich_only:
                cur.close()
                return None
            cur.execute("""
                INSERT INTO platform.nca_aif (
                    aif_type, aif_name, aif_clean_name, lei, isin,
                    home_country, custodian, fund_strategy,
                    address, address_source,
                    lineage,
                    first_seen_at, last_seen_at, created_at, last_updated
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING aif_id
            """, (
                rec.aif_type, rec.aif_name, rec.aif_clean_name, rec.lei, rec.isin,
                rec.home_country, rec.custodian, rec.fund_strategy,
                rec.address, rec.address_source or rec.source if rec.address else None,
                lineage_entry,
                now, now, now, now,
            ))
            aif_id = cur.fetchone()[0]

        if rec.nca_code and aif_id:
            self._upsert_registration(cur, rec.aif_type, aif_id, rec, now)

        cur.close()
        return aif_id

    # ── Shared registration upsert ──────────────────────────────────

    def _upsert_registration(self, cur, entity_type: str, entity_id: str, rec, now):
        today = now.strftime("%Y-%m-%d")
        lineage_entry = self._build_lineage_entry(rec.source, {
            "auth_status": rec.auth_status, "auth_date": rec.auth_date,
        }, today)

        cur.execute("""
            INSERT INTO platform.nca_registrations (
                entity_type, entity_id, nca_code, nca_entity_code,
                auth_status, auth_date, withdrawal_date,
                source, source_entity_id, fetched_at,
                first_seen_at, last_seen_at, lineage
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (entity_type, entity_id, nca_code, nca_entity_code)
            DO UPDATE SET
                auth_status = COALESCE(EXCLUDED.auth_status, platform.nca_registrations.auth_status),
                auth_date = COALESCE(EXCLUDED.auth_date, platform.nca_registrations.auth_date),
                withdrawal_date = COALESCE(EXCLUDED.withdrawal_date, platform.nca_registrations.withdrawal_date),
                fetched_at = EXCLUDED.fetched_at,
                last_seen_at = EXCLUDED.last_seen_at,
                deleted_at = NULL,
                lineage = CASE
                    WHEN platform.nca_registrations.lineage IS NULL THEN EXCLUDED.lineage
                    WHEN EXCLUDED.lineage IS NULL THEN platform.nca_registrations.lineage
                    ELSE platform.nca_registrations.lineage || '; ' || EXCLUDED.lineage
                END
        """, (
            entity_type, entity_id, rec.nca_code, rec.nca_entity_code,
            rec.auth_status, rec.auth_date, rec.withdrawal_date,
            rec.source, rec.source_entity_id, now, now, now,
            lineage_entry,
        ))

    # ── Hierarchy linking ───────────────────────────────────────────

    def link_aif_to_aifm(self, aif_id: str, aifm_clean_name: str,
                         country_code: str = None, aifm_id: str = None):
        """Link an AIF to its managing AIFM.

        If aifm_id is provided, uses it directly (fast path).
        Otherwise falls back to clean_name lookup.
        """
        cur = self._conn.cursor()
        if not aifm_id:
            cn = clean_name(aifm_clean_name)
            if country_code:
                cur.execute("""
                    SELECT aifm_id FROM platform.nca_aifm
                    WHERE aifm_clean_name = %s AND (home_country = %s OR home_country IS NULL)
                    LIMIT 1
                """, (cn, country_code))
            else:
                cur.execute("""
                    SELECT aifm_id FROM platform.nca_aifm
                    WHERE aifm_clean_name = %s LIMIT 1
                """, (cn,))
            row = cur.fetchone()
            aifm_id = row[0] if row else None
        if aifm_id:
            cur.execute("""
                UPDATE platform.nca_aif SET managing_aifm_id = %s
                WHERE aif_id = %s AND managing_aifm_id IS NULL
            """, (aifm_id, aif_id))
        cur.close()

    def link_sub_aif_to_aif(self, sub_aif_id: str, aif_clean_name: str,
                             parent_aif_id: str = None):
        """Link a SUB_AIF to its parent AIF.

        If parent_aif_id is provided, uses it directly (fast path).
        Otherwise falls back to clean_name lookup.
        """
        cur = self._conn.cursor()
        if not parent_aif_id:
            cn = clean_name(aif_clean_name)
            cur.execute("""
                SELECT aif_id FROM platform.nca_aif
                WHERE aif_clean_name = %s AND aif_type = 'AIF' LIMIT 1
            """, (cn,))
            row = cur.fetchone()
            parent_aif_id = row[0] if row else None
        if parent_aif_id:
            cur.execute("""
                UPDATE platform.nca_aif SET parent_aif_id = %s
                WHERE aif_id = %s AND parent_aif_id IS NULL
            """, (parent_aif_id, sub_aif_id))
        cur.close()

    # ── Disappearance tracking ──────────────────────────────────────

    def mark_disappeared(self, source: str, nca_code: str, entity_type: str, seen_ids: set[str]):
        cur = self._conn.cursor()
        now = datetime.utcnow()
        cur.execute("""
            UPDATE platform.nca_registrations
            SET deleted_at = %s
            WHERE source = %s AND nca_code = %s AND entity_type = %s
              AND deleted_at IS NULL
              AND entity_id NOT IN (SELECT unnest(%s::uuid[]))
        """, (now, source, nca_code, entity_type, list(seen_ids)))
        updated = cur.rowcount
        cur.close()
        return updated

    def commit(self):
        self._conn.commit()

    # ── GLEIF enrichment ──────────────────────────────────────────

    def enrich_from_gleif(self) -> dict:
        """Enrich NCA entities from gleif_lei_cache using LEI join.

        GLEIF is AUTHORITATIVE for home_country (legal jurisdiction) — it
        overwrites ESMA's home_country which can be wrong for passport
        notifications (host MS instead of actual domicile).
        Address is updated only if GLEIF has higher priority than current source.
        """
        cur = self._conn.cursor()
        today = datetime.utcnow().strftime("%Y-%m-%d")
        result = {"aifm_enriched": 0, "aif_enriched": 0}

        # ── Enrich AIFMs ──
        # Select ALL AIFMs with a LEI in GLEIF — GLEIF is authoritative for
        # home_country (legal jurisdiction) and fills address/registered_as NULLs.
        cur.execute("""
            SELECT m.aifm_id, m.lei, m.lineage,
                   m.home_country, m.address, m.address_source, m.registered_as,
                   g.headquarters_address AS g_hq, g.legal_address AS g_legal,
                   g.registered_as AS g_reg, g.country AS g_country
            FROM platform.nca_aifm m
            JOIN platform.gleif_lei_cache g ON g.lei = m.lei
            WHERE m.lei IS NOT NULL
        """)
        rows = cur.fetchall()
        for row in rows:
            (mid, lei, lineage, hc, addr, addr_source, reg_as,
             g_hq, g_legal, g_reg, g_country) = row

            fields_filled = {}
            gleif_addr = g_hq or g_legal

            if not addr and gleif_addr:
                fields_filled["address"] = gleif_addr
            elif addr and gleif_addr and _should_update_address("GLEIF", addr_source):
                fields_filled["address"] = gleif_addr

            if not reg_as and g_reg:
                fields_filled["registered_as"] = g_reg

            # GLEIF is authoritative for home_country (legal jurisdiction).
            # Overwrites ESMA's home_country which can be wrong for passport
            # notifications (ESMA uses host MS, not actual domicile).
            if g_country and hc != g_country:
                fields_filled["home_country"] = g_country

            if not fields_filled:
                continue

            lineage_entry = self._build_lineage_entry("GLEIF", fields_filled, today)
            new_lineage = self._append_lineage(lineage, lineage_entry)
            new_addr = fields_filled.get("address")
            update_addr = new_addr is not None
            new_country = fields_filled.get("home_country")
            update_country = new_country is not None

            cur.execute("""
                UPDATE platform.nca_aifm SET
                    address = CASE WHEN %s THEN %s ELSE address END,
                    address_source = CASE WHEN %s THEN 'GLEIF' ELSE address_source END,
                    registered_as = COALESCE(registered_as, %s),
                    home_country = CASE WHEN %s THEN %s ELSE COALESCE(home_country, %s) END,
                    lineage = %s, last_updated = now()
                WHERE aifm_id = %s
            """, (update_addr, new_addr, update_addr, g_reg,
                  update_country, new_country, g_country,
                  new_lineage, mid))
            result["aifm_enriched"] += 1

        # ── Enrich AIFs ──
        # Same as AIFMs — GLEIF is authoritative for home_country.
        cur.execute("""
            SELECT f.aif_id, f.lei, f.lineage,
                   f.home_country, f.address, f.address_source,
                   g.headquarters_address AS g_hq, g.legal_address AS g_legal,
                   g.country AS g_country
            FROM platform.nca_aif f
            JOIN platform.gleif_lei_cache g ON g.lei = f.lei
            WHERE f.lei IS NOT NULL
        """)
        rows = cur.fetchall()
        for row in rows:
            (fid, lei, lineage, hc, addr, addr_source,
             g_hq, g_legal, g_country) = row

            fields_filled = {}
            gleif_addr = g_hq or g_legal

            if not addr and gleif_addr:
                fields_filled["address"] = gleif_addr
            elif addr and gleif_addr and _should_update_address("GLEIF", addr_source):
                fields_filled["address"] = gleif_addr

            # GLEIF is authoritative for home_country.
            if g_country and hc != g_country:
                fields_filled["home_country"] = g_country

            if not fields_filled:
                continue

            lineage_entry = self._build_lineage_entry("GLEIF", fields_filled, today)
            new_lineage = self._append_lineage(lineage, lineage_entry)
            new_addr = fields_filled.get("address")
            update_addr = new_addr is not None
            new_country = fields_filled.get("home_country")
            update_country = new_country is not None

            cur.execute("""
                UPDATE platform.nca_aif SET
                    address = CASE WHEN %s THEN %s ELSE address END,
                    address_source = CASE WHEN %s THEN 'GLEIF' ELSE address_source END,
                    home_country = CASE WHEN %s THEN %s ELSE COALESCE(home_country, %s) END,
                    lineage = %s, last_updated = now()
                WHERE aif_id = %s
            """, (update_addr, new_addr, update_addr,
                  update_country, new_country, g_country,
                  new_lineage, fid))
            result["aif_enriched"] += 1

        self.commit()
        cur.close()
        return result

    # ── AIF address inheritance from AIFM ──────────────────────────

    def inherit_aifm_address(self) -> int:
        """For AIFs without address, inherit from their managing AIFM.

        Address priority: KVK > MANUAL > AIFM > GLEIF.
        Only fills NULLs — never overwrites existing address.
        """
        cur = self._conn.cursor()
        today = datetime.utcnow().strftime("%Y-%m-%d")

        cur.execute("""
            SELECT f.aif_id, m.address, f.lineage
            FROM platform.nca_aif f
            JOIN platform.nca_aifm m ON m.aifm_id = f.managing_aifm_id
            WHERE f.address IS NULL
              AND m.address IS NOT NULL
              AND f.managing_aifm_id IS NOT NULL
        """)
        rows = cur.fetchall()
        count = 0
        for aif_id, aifm_addr, lineage in rows:
            lineage_entry = self._build_lineage_entry("AIFM", {"address": aifm_addr}, today)
            new_lineage = self._append_lineage(lineage, lineage_entry)
            cur.execute("""
                UPDATE platform.nca_aif SET
                    address = %s, address_source = 'AIFM',
                    lineage = %s, last_updated = now()
                WHERE aif_id = %s AND address IS NULL
            """, (aifm_addr, new_lineage, aif_id))
            count += cur.rowcount

        self.commit()
        cur.close()
        return count

    # ── LEI enrichment from GLEIF by name + country ────────────────

    def enrich_lei_from_gleif(self) -> dict:
        """For entities without LEI, look up by clean_name in GLEIF.

        Two-step matching: first name+country (precise), then name-only
        fallback (catches entities with wrong country from ESMA).
        Only assigns LEI if exactly 1 match is found.

        Uses bulk SQL JOINs instead of row-by-row queries for performance.
        """
        cur = self._conn.cursor()
        today = datetime.utcnow().strftime("%Y-%m-%d")
        lineage_suffix = f"GLEIF:lei@{today}"
        result = {"aifm_lei_found": 0, "aif_lei_found": 0,
                  "aifm_no_match": 0, "aif_no_match": 0,
                  "aifm_multi_match": 0, "aif_multi_match": 0}

        for entity_table, id_col, name_col, found_key, no_key, multi_key in [
            ("platform.nca_aifm", "aifm_id", "aifm_clean_name",
             "aifm_lei_found", "aifm_no_match", "aifm_multi_match"),
            ("platform.nca_aif", "aif_id", "aif_clean_name",
             "aif_lei_found", "aif_no_match", "aif_multi_match"),
        ]:
            # Count entities without LEI
            cur.execute(f"""
                SELECT COUNT(*) FROM {entity_table}
                WHERE lei IS NULL AND deleted_at IS NULL
            """)
            total_without = cur.fetchone()[0]
            logger.info(f"LEI enrichment: {total_without} {entity_table.split('.')[-1]}s without LEI")

            # Step 1: Match by name + country (only for entities with country)
            # Uses a CTE to find normalized_names with exactly 1 GLEIF match
            # for that name+country combination.
            cur.execute(f"""
                WITH candidates AS (
                    SELECT e.{id_col} AS eid, e.{name_col} AS cn, e.home_country AS hc, e.lineage
                    FROM {entity_table} e
                    WHERE e.lei IS NULL AND e.deleted_at IS NULL
                      AND e.home_country IS NOT NULL
                      AND e.{name_col} IS NOT NULL AND e.{name_col} != ''
                ),
                gleif_matches AS (
                    SELECT c.eid, c.lineage, g.lei
                    FROM candidates c
                    JOIN platform.gleif_lei_cache g
                        ON g.normalized_name = c.cn AND g.country = c.hc
                ),
                unique_matches AS (
                    SELECT eid, lineage, MIN(lei) AS lei
                    FROM gleif_matches
                    GROUP BY eid, lineage
                    HAVING COUNT(*) = 1
                )
                UPDATE {entity_table} e SET
                    lei = TRIM(u.lei),
                    lineage = CASE
                        WHEN e.lineage IS NULL THEN '{lineage_suffix}'
                        ELSE e.lineage || '; ' || '{lineage_suffix}'
                    END,
                    last_updated = now()
                FROM unique_matches u
                WHERE e.{id_col} = u.eid AND e.lei IS NULL
            """)
            step1_found = cur.rowcount
            result[found_key] += step1_found

            # Step 2: Name-only fallback for remaining entities without LEI
            # (includes entities with NULL country AND those where step 1
            # didn't find a match because country was wrong).
            cur.execute(f"""
                WITH candidates AS (
                    SELECT e.{id_col} AS eid, e.{name_col} AS cn, e.lineage
                    FROM {entity_table} e
                    WHERE e.lei IS NULL AND e.deleted_at IS NULL
                      AND e.{name_col} IS NOT NULL AND e.{name_col} != ''
                ),
                gleif_matches AS (
                    SELECT c.eid, c.lineage, g.lei
                    FROM candidates c
                    JOIN platform.gleif_lei_cache g
                        ON g.normalized_name = c.cn
                ),
                unique_matches AS (
                    SELECT eid, lineage, MIN(lei) AS lei
                    FROM gleif_matches
                    GROUP BY eid, lineage
                    HAVING COUNT(*) = 1
                )
                UPDATE {entity_table} e SET
                    lei = TRIM(u.lei),
                    lineage = CASE
                        WHEN e.lineage IS NULL THEN '{lineage_suffix}'
                        ELSE e.lineage || '; ' || '{lineage_suffix}'
                    END,
                    last_updated = now()
                FROM unique_matches u
                WHERE e.{id_col} = u.eid AND e.lei IS NULL
            """)
            step2_found = cur.rowcount
            result[found_key] += step2_found

            # Count remaining for stats
            cur.execute(f"""
                SELECT COUNT(*) FROM {entity_table}
                WHERE lei IS NULL AND deleted_at IS NULL
            """)
            remaining = cur.fetchone()[0]
            result[no_key] = remaining  # approximate (includes multi-match)

            logger.info(
                f"  {entity_table.split('.')[-1]}: "
                f"{step1_found} found (name+country), "
                f"{step2_found} found (name-only fallback), "
                f"{remaining} remaining"
            )

        self.commit()
        cur.close()
        return result

    # ── Schema migration ─────────────────────────────────────────

    def _ensure_terminated_at_column(self):
        """Add terminated_at column if it doesn't exist (idempotent migration)."""
        cur = self._conn.cursor()
        for table in ("platform.nca_aifm", "platform.nca_aif"):
            try:
                cur.execute(f"""
                    ALTER TABLE {table}
                    ADD COLUMN IF NOT EXISTS terminated_at TIMESTAMPTZ
                """)
            except Exception:
                self._conn.rollback()
        self.commit()
        cur.close()

    # ── Termination detection ──────────────────────────────────────

    def detect_terminations(self, staleness_days: int = 30) -> dict:
        """Detect terminated/withdrawn entities and set terminated_at.

        Strategy:
          1. CSSF entities: use withdrawal_date from nca_registrations
             (end_mgmt for managers/funds, closing_date for sub-funds).
          2. All entities: if last_seen_at < now() - staleness_days, the entity
             has disappeared from the source register → mark as terminated.
          3. Clear terminated_at for entities that reappear (last_seen_at recent
             and terminated_at was set by staleness rule, not by source data).

        Returns counts of entities marked terminated.
        """
        self._ensure_terminated_at_column()
        cur = self._conn.cursor()
        result = {"aifm_from_source": 0, "aif_from_source": 0,
                  "aifm_from_staleness": 0, "aif_from_staleness": 0,
                  "aifm_cleared": 0, "aif_cleared": 0}

        # ── Step 1: CSSF withdrawal dates → terminated_at ──
        # For AIFMs: if ALL registrations for a given AIFM have a withdrawal_date,
        # set terminated_at to the latest withdrawal_date.
        cur.execute("""
            UPDATE platform.nca_aifm m SET
                terminated_at = sub.latest_withdrawal
            FROM (
                SELECT r.entity_id,
                       MAX(r.withdrawal_date) AS latest_withdrawal
                FROM platform.nca_registrations r
                WHERE r.entity_type = 'AIFM'
                  AND r.withdrawal_date IS NOT NULL
                  AND r.deleted_at IS NULL
                GROUP BY r.entity_id
                HAVING COUNT(*) = COUNT(r.withdrawal_date)
            ) sub
            WHERE m.aifm_id = sub.entity_id
              AND m.terminated_at IS NULL
              AND m.deleted_at IS NULL
        """)
        result["aifm_from_source"] = cur.rowcount

        # For AIFs/SUB_AIFs: same logic.
        cur.execute("""
            UPDATE platform.nca_aif f SET
                terminated_at = sub.latest_withdrawal
            FROM (
                SELECT r.entity_id,
                       MAX(r.withdrawal_date) AS latest_withdrawal
                FROM platform.nca_registrations r
                WHERE r.entity_type IN ('AIF', 'SUB_AIF')
                  AND r.withdrawal_date IS NOT NULL
                  AND r.deleted_at IS NULL
                GROUP BY r.entity_id
                HAVING COUNT(*) = COUNT(r.withdrawal_date)
            ) sub
            WHERE f.aif_id = sub.entity_id
              AND f.terminated_at IS NULL
              AND f.deleted_at IS NULL
        """)
        result["aif_from_source"] = cur.rowcount

        # ── Step 2: Staleness-based termination ──
        # Entities not seen for > staleness_days are assumed terminated.
        cur.execute("""
            UPDATE platform.nca_aifm SET
                terminated_at = last_seen_at
            WHERE last_seen_at < now() - make_interval(days => %s)
              AND terminated_at IS NULL
              AND deleted_at IS NULL
        """, (staleness_days,))
        result["aifm_from_staleness"] = cur.rowcount

        cur.execute("""
            UPDATE platform.nca_aif SET
                terminated_at = last_seen_at
            WHERE last_seen_at < now() - make_interval(days => %s)
              AND terminated_at IS NULL
              AND deleted_at IS NULL
        """, (staleness_days,))
        result["aif_from_staleness"] = cur.rowcount

        # ── Step 3: Clear terminated_at for entities that reappeared ──
        # Only clear if terminated_at was set by staleness (no source withdrawal_date).
        cur.execute("""
            UPDATE platform.nca_aifm m SET
                terminated_at = NULL
            WHERE m.terminated_at IS NOT NULL
              AND m.deleted_at IS NULL
              AND m.last_seen_at >= now() - make_interval(days => %s)
              AND NOT EXISTS (
                  SELECT 1 FROM platform.nca_registrations r
                  WHERE r.entity_id = m.aifm_id
                    AND r.entity_type = 'AIFM'
                    AND r.withdrawal_date IS NOT NULL
                    AND r.deleted_at IS NULL
              )
        """, (staleness_days,))
        result["aifm_cleared"] = cur.rowcount

        cur.execute("""
            UPDATE platform.nca_aif f SET
                terminated_at = NULL
            WHERE f.terminated_at IS NOT NULL
              AND f.deleted_at IS NULL
              AND f.last_seen_at >= now() - make_interval(days => %s)
              AND NOT EXISTS (
                  SELECT 1 FROM platform.nca_registrations r
                  WHERE r.entity_id = f.aif_id
                    AND r.entity_type IN ('AIF', 'SUB_AIF')
                    AND r.withdrawal_date IS NOT NULL
                    AND r.deleted_at IS NULL
              )
        """, (staleness_days,))
        result["aif_cleared"] = cur.rowcount

        self.commit()
        cur.close()
        return result

    # ── NPPR propagation ─────────────────────────────────────────

    def propagate_nppr_status(self) -> dict:
        """Propagate NPPR status to registrations with blank auth_status.

        If an entity has at least one registration marked NPPR (from an NCA
        like CSSF), other registrations for that same entity with NULL
        auth_status (typically from ESMA, which doesn't provide status) are
        also set to NPPR.
        """
        cur = self._conn.cursor()
        result = {"aifm_propagated": 0, "aif_propagated": 0}

        cur.execute("""
            UPDATE platform.nca_registrations r
            SET auth_status = 'NPPR'
            WHERE r.auth_status IS NULL
              AND r.deleted_at IS NULL
              AND EXISTS (
                  SELECT 1 FROM platform.nca_registrations r2
                  WHERE r2.entity_id = r.entity_id
                    AND r2.entity_type = r.entity_type
                    AND r2.auth_status = 'NPPR'
                    AND r2.deleted_at IS NULL
              )
              AND r.entity_type = 'AIFM'
        """)
        result["aifm_propagated"] = cur.rowcount

        cur.execute("""
            UPDATE platform.nca_registrations r
            SET auth_status = 'NPPR'
            WHERE r.auth_status IS NULL
              AND r.deleted_at IS NULL
              AND EXISTS (
                  SELECT 1 FROM platform.nca_registrations r2
                  WHERE r2.entity_id = r.entity_id
                    AND r2.entity_type = r.entity_type
                    AND r2.auth_status = 'NPPR'
                    AND r2.deleted_at IS NULL
              )
              AND r.entity_type IN ('AIF', 'SUB_AIF')
        """)
        result["aif_propagated"] = cur.rowcount

        self.commit()
        cur.close()
        return result

    # ── Reporting ───────────────────────────────────────────────────

    def get_report(self) -> dict:
        cur = self._conn.cursor()
        report = {}

        cur.execute("""
            SELECT COUNT(*), COUNT(*) FILTER (WHERE deleted_at IS NULL),
                   COUNT(lei) FILTER (WHERE deleted_at IS NULL)
            FROM platform.nca_aifm
        """)
        row = cur.fetchone()
        report["aifm"] = {"total": row[0], "active": row[1], "with_lei": row[2]}

        cur.execute("""
            SELECT aif_type, COUNT(*), COUNT(*) FILTER (WHERE deleted_at IS NULL),
                   COUNT(lei) FILTER (WHERE deleted_at IS NULL)
            FROM platform.nca_aif GROUP BY aif_type ORDER BY aif_type
        """)
        report["aif"] = {r[0]: {"total": r[1], "active": r[2], "with_lei": r[3]} for r in cur.fetchall()}

        cur.execute("""
            SELECT source, nca_code, entity_type,
                   COUNT(*), COUNT(*) FILTER (WHERE deleted_at IS NULL)
            FROM platform.nca_registrations
            GROUP BY source, nca_code, entity_type
            ORDER BY source, nca_code, entity_type
        """)
        report["registrations"] = [
            {"source": r[0], "nca_code": r[1], "entity_type": r[2],
             "total": r[3], "active": r[4]}
            for r in cur.fetchall()
        ]

        cur.execute("""
            SELECT m.aifm_name, m.home_country, COUNT(f.aif_id) AS aif_count
            FROM platform.nca_aifm m
            LEFT JOIN platform.nca_aif f ON f.managing_aifm_id = m.aifm_id AND f.deleted_at IS NULL
            WHERE m.deleted_at IS NULL
            GROUP BY m.aifm_id, m.aifm_name, m.home_country
            ORDER BY aif_count DESC LIMIT 20
        """)
        report["top_aifm"] = [
            {"name": r[0], "country": r[1], "aif_count": r[2]}
            for r in cur.fetchall()
        ]

        cur.execute("""
            SELECT entity_type, entity_id, COUNT(DISTINCT nca_code) AS nca_count
            FROM platform.nca_registrations
            WHERE deleted_at IS NULL
            GROUP BY entity_type, entity_id
            HAVING COUNT(DISTINCT nca_code) > 1
        """)
        report["multi_nca_count"] = cur.rowcount

        cur.close()
        return report


# ── Utility ──────────────────────────────────────────────────────────

def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s in ("", "n.v.t.", "n/a"):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%dT%H:%M:%SZ"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fetch_url(url: str, timeout: int = 60) -> bytes:
    req = Request(url, headers={"User-Agent": "Eagle/1.0 (NCA register fetch)"})
    with urlopen(req, timeout=timeout) as resp:
        return resp.read()


# ── AFM fetcher ──────────────────────────────────────────────────────

def fetch_afm(store: NCARegisterStore) -> dict:
    """Fetch AFM authorised and registered (light) registers."""
    import openpyxl

    result = {"aifm": 0, "aif": 0, "sub_aif": 0}
    seen_aifm_ids = set()
    seen_aif_ids = set()

    # ── AFM Authorised ───────────────────────────────────────────
    logger.info("Fetching AFM authorised register...")
    data = _fetch_url("https://www.afm.nl/~/profmedia/files/registers/register-aifm.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    header_row = None
    for r in range(1, 20):
        vals = [str(c.value or "") for c in ws[r]]
        if any("vergunningnummer" in v.lower() for v in vals):
            header_row = r
            break
    if not header_row:
        raise ValueError("Could not find header row in AFM authorised register")

    headers = [str(c.value or "").strip() for c in ws[header_row]]
    logger.info(f"AFM authorised: header row {header_row}, columns: {headers}")

    col = {h: i for i, h in enumerate(headers)}
    current_aifm_name = None
    current_aifm_license = None

    for r in range(header_row + 1, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        if not vals or all(v is None for v in vals):
            continue

        license_nr = vals[col.get("Vergunningnummer", 0)]
        aifm_name = vals[col.get("Naam Beheerder", 2)]

        if license_nr and aifm_name:
            current_aifm_license = str(int(license_nr)) if isinstance(license_nr, (int, float)) else str(license_nr)
            current_aifm_name = str(aifm_name).strip()

            rec = AIFMRecord(
                aifm_name=current_aifm_name,
                home_country="NL",
                nca_code="AFM",
                auth_status="AUTHORISED",
                source="AFM",
                source_entity_id=current_aifm_license,
            )
            mid = store.upsert_aifm(rec)
            seen_aifm_ids.add(mid)
            result["aifm"] += 1

        fund_name_idx = col.get("Naam Beleggingsinstelling", 4)
        fund_id_idx = col.get("AFM Fonds ID", 5)
        subfund_name_idx = col.get("Naam Subfonds", 6)
        subfund_id_idx = col.get("AFM Subfonds ID", 7)
        custodian_idx = col.get("Bewaarder", 9)
        date_idx = col.get("Datum", 3)

        fund_name = vals[fund_name_idx] if fund_name_idx is not None and fund_name_idx < len(vals) else None
        fund_id = vals[fund_id_idx] if fund_id_idx is not None and fund_id_idx < len(vals) else None
        subfund_name = vals[subfund_name_idx] if subfund_name_idx is not None and subfund_name_idx < len(vals) else None
        subfund_id = vals[subfund_id_idx] if subfund_id_idx is not None and subfund_id_idx < len(vals) else None
        custodian = vals[custodian_idx] if custodian_idx is not None and custodian_idx < len(vals) else None
        reg_date = _parse_date(vals[date_idx] if date_idx is not None and date_idx < len(vals) else None)

        if fund_name and str(fund_name).strip():
            fund_id_str = str(int(fund_id)) if isinstance(fund_id, (int, float)) else str(fund_id or "")
            aif_rec = AIFRecord(
                aif_type="AIF",
                aif_name=str(fund_name).strip(),
                home_country="NL",
                custodian=str(custodian).strip() if custodian else None,
                nca_code="AFM",
                nca_entity_code=fund_id_str if fund_id_str else None,
                auth_status="AUTHORISED",
                auth_date=reg_date,
                source="AFM",
                source_entity_id=fund_id_str,
                managing_aifm_name=current_aifm_name,
            )
            fid = store.upsert_aif(aif_rec)
            seen_aif_ids.add(fid)
            result["aif"] += 1

            if current_aifm_name:
                store.link_aif_to_aifm(fid, current_aifm_name, "NL")

            if subfund_name and str(subfund_name).strip() not in ("n.v.t.", "", "n/a"):
                subfund_id_str = str(int(subfund_id)) if isinstance(subfund_id, (int, float)) else str(subfund_id or "")
                sub_rec = AIFRecord(
                    aif_type="SUB_AIF",
                    aif_name=str(subfund_name).strip(),
                    home_country="NL",
                    custodian=str(custodian).strip() if custodian else None,
                    nca_code="AFM",
                    nca_entity_code=subfund_id_str if subfund_id_str else None,
                    auth_status="AUTHORISED",
                    auth_date=reg_date,
                    source="AFM",
                    source_entity_id=subfund_id_str,
                    managing_aifm_name=current_aifm_name,
                    parent_aif_name=str(fund_name).strip(),
                )
                sid = store.upsert_aif(sub_rec)
                seen_aif_ids.add(sid)
                store.link_sub_aif_to_aif(sid, str(fund_name).strip())
                store.link_aif_to_aifm(sid, current_aifm_name, "NL")
                result["sub_aif"] += 1

    # ── AFM Light (Registered) ───────────────────────────────────
    logger.info("Fetching AFM light (registered) register...")
    data = _fetch_url("https://www.afm.nl/~/profmedia/files/registers/register-aifmd-light.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    header_row = None
    for r in range(1, 20):
        vals = [str(c.value or "").lower() for c in ws[r]]
        if any("naam beheerder" in v for v in vals):
            header_row = r
            break
    if not header_row:
        raise ValueError("Could not find header row in AFM light register")

    headers_nl = [str(c.value or "").strip() for c in ws[header_row]]
    col = {h: i for i, h in enumerate(headers_nl)}
    current_aifm_name = None

    for r in range(header_row + 2, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        if not vals or all(v is None for v in vals):
            continue

        mgr_name = vals[col.get("Naam beheerder", 0)]
        fund_name = vals[col.get("Naam beleggingsinstelling", 1)]
        fund_id = vals[col.get("Fonds ID", 2)]
        start_date = _parse_date(vals[col.get("Ingangsdatum", 3)])

        if mgr_name and str(mgr_name).strip():
            mgr_name_str = str(mgr_name).strip()
            if mgr_name_str != current_aifm_name:
                current_aifm_name = mgr_name_str
                rec = AIFMRecord(
                    aifm_name=current_aifm_name,
                    home_country="NL",
                    nca_code="AFM",
                    auth_status="REGISTERED",
                    source="AFM",
                )
                mid = store.upsert_aifm(rec)
                seen_aifm_ids.add(mid)
                result["aifm"] += 1

        if fund_name and str(fund_name).strip():
            fund_id_str = str(int(fund_id)) if isinstance(fund_id, (int, float)) else str(fund_id or "")
            aif_rec = AIFRecord(
                aif_type="AIF",
                aif_name=str(fund_name).strip(),
                home_country="NL",
                nca_code="AFM",
                nca_entity_code=fund_id_str if fund_id_str else None,
                auth_status="REGISTERED",
                auth_date=start_date,
                source="AFM",
                source_entity_id=fund_id_str,
                managing_aifm_name=current_aifm_name,
            )
            fid = store.upsert_aif(aif_rec)
            seen_aif_ids.add(fid)
            result["aif"] += 1

            if current_aifm_name:
                store.link_aif_to_aifm(fid, current_aifm_name, "NL")

    store.commit()
    logger.info(f"AFM fetch complete: {result}")
    return result


# ── CSSF fetcher ─────────────────────────────────────────────────────

CSSF_STATUS_MAP = {"AUT": "AUTHORISED", "REG": "REGISTERED", "A42": "NPPR"}

def fetch_cssf(store: NCARegisterStore) -> dict:
    logger.info("Fetching CSSF IDENTIFIANTS_AIFM.zip...")
    data = _fetch_url("https://www.cssf.lu/wp-content/uploads/IDENTIFIANTS_AIFM.zip")
    result = {"aifm": 0, "aif": 0, "sub_aif": 0}
    seen_aifm_ids = set()
    seen_aif_ids = set()

    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        with zf.open("IDENTIFIANTS_AIFM.csv") as f:
            text = f.read().decode("utf-16")

    lines = text.strip().split("\n")
    reader = csv.reader(lines, delimiter="\t")
    rows = list(reader)

    headers = [h.strip() for h in rows[0]]
    logger.info(f"CSSF register: {len(rows)-2} data rows, headers: {headers}")

    aifm_cache = {}    # nca_code → aifm_id
    aif_cache = {}     # nca_code → aif_id
    aif_seen = set()
    row_count = 0

    for row in rows[2:]:
        if len(row) < 13:
            continue

        aifm_prefix = row[0].strip()
        aifm_number = row[1].strip()
        aifm_name = row[2].strip()
        status_raw = row[3].strip()
        aif_prefix = row[4].strip()
        aif_number = row[5].strip()
        aif_name = row[6].strip()
        subfund_number = row[7].strip()
        subfund_name = row[8].strip()
        start_mgmt = _parse_date(row[9].strip())
        end_mgmt = _parse_date(row[10].strip())
        reg_date = _parse_date(row[11].strip())
        closing_date = _parse_date(row[12].strip())

        auth_status = CSSF_STATUS_MAP.get(status_raw, "AUTHORISED")
        aifm_nca_code = f"{aifm_prefix}{aifm_number}"
        aif_nca_code = f"{aif_prefix}{aif_number}"
        is_ended = end_mgmt and end_mgmt.year < 2400
        is_closed = closing_date and closing_date.year < 2400

        if aifm_nca_code not in aifm_cache:
            rec = AIFMRecord(
                aifm_name=aifm_name,
                home_country="LU" if auth_status != "NPPR" else None,
                nca_code="CSSF",
                nca_entity_code=aifm_nca_code,
                auth_status=auth_status,
                source="CSSF",
                source_entity_id=aifm_nca_code,
            )
            mid = store.upsert_aifm(rec)
            aifm_cache[aifm_nca_code] = mid
            seen_aifm_ids.add(mid)
            result["aifm"] += 1

        aifm_id = aifm_cache[aifm_nca_code]
        has_subfund = subfund_name and subfund_number and subfund_number != "00000000"

        if has_subfund:
            subfund_nca_code = f"{aif_prefix}{aif_number}_{subfund_number}"

            # Parent AIF — cache to avoid re-lookup
            if aif_nca_code not in aif_cache:
                aif_rec = AIFRecord(
                    aif_type="AIF",
                    aif_name=aif_name,
                    home_country="LU" if auth_status != "NPPR" else None,
                    nca_code="CSSF",
                    nca_entity_code=aif_nca_code,
                    auth_status="INACTIVE" if is_ended else auth_status,
                    auth_date=start_mgmt,
                    withdrawal_date=end_mgmt if is_ended else None,
                    source="CSSF",
                    source_entity_id=aif_nca_code,
                    managing_aifm_name=aifm_name,
                )
                fid = store.upsert_aif(aif_rec)
                aif_cache[aif_nca_code] = fid
                seen_aif_ids.add(fid)
                store.link_aif_to_aifm(fid, aifm_name, "LU", aifm_id=aifm_id)
                if aif_nca_code not in aif_seen:
                    aif_seen.add(aif_nca_code)
                    result["aif"] += 1

            parent_aif_id = aif_cache[aif_nca_code]

            sub_rec = AIFRecord(
                aif_type="SUB_AIF",
                aif_name=subfund_name,
                home_country="LU" if auth_status != "NPPR" else None,
                nca_code="CSSF",
                nca_entity_code=subfund_nca_code,
                auth_status="INACTIVE" if is_closed else auth_status,
                auth_date=reg_date,
                withdrawal_date=closing_date if is_closed else None,
                source="CSSF",
                source_entity_id=subfund_nca_code,
                managing_aifm_name=aifm_name,
                parent_aif_name=aif_name,
            )
            sid = store.upsert_aif(sub_rec)
            seen_aif_ids.add(sid)
            store.link_sub_aif_to_aif(sid, aif_name, parent_aif_id=parent_aif_id)
            store.link_aif_to_aifm(sid, aifm_name, "LU", aifm_id=aifm_id)
            result["sub_aif"] += 1
        else:
            aif_rec = AIFRecord(
                aif_type="AIF",
                aif_name=aif_name,
                home_country="LU" if auth_status != "NPPR" else None,
                nca_code="CSSF",
                nca_entity_code=aif_nca_code,
                auth_status="INACTIVE" if is_ended else auth_status,
                auth_date=reg_date or start_mgmt,
                withdrawal_date=end_mgmt if is_ended else None,
                source="CSSF",
                source_entity_id=aif_nca_code,
                managing_aifm_name=aifm_name,
            )
            fid = store.upsert_aif(aif_rec)
            seen_aif_ids.add(fid)
            aif_cache[aif_nca_code] = fid
            store.link_aif_to_aifm(fid, aifm_name, "LU", aifm_id=aifm_id)
            result["aif"] += 1

        row_count += 1
        if row_count % 5000 == 0:
            store.commit()
            logger.info(f"  ... processed {row_count:,}/{len(rows)-2:,} CSSF rows")

    store.commit()
    logger.info(f"CSSF fetch complete: {result}")
    return result


# ── ESMA fetcher ─────────────────────────────────────────────────────

ESMA_AIFM_URL = "https://registers.esma.europa.eu/solr/esma_registers_upreg/select"
ESMA_FUND_URL = "https://registers.esma.europa.eu/solr/esma_registers_funds/select"

def fetch_esma(store: NCARegisterStore) -> dict:
    result = {"aifm": 0, "aif": 0}
    page_size = 500
    start = 0
    total = None

    logger.info("Fetching ESMA AIFM register (all EU, head offices only)...")
    while total is None or start < total:
        params = (
            f"q=*:*&fq=type_s:parent&fq=ae_entityTypeCode:aif"
            f"&fq=ae_officeType:%22head%20office%22"
            f"&wt=json&rows={page_size}&start={start}&fl=*"
        )
        url = f"{ESMA_AIFM_URL}?{params}"
        raw = _fetch_url(url)
        data = json.loads(raw)
        total = data["response"]["numFound"]
        docs = data["response"]["docs"]

        if start == 0:
            logger.info(f"ESMA AIFMs: {total} total entities")

        for doc in docs:
            name = doc.get("ae_entityName", "").strip()
            if not name:
                continue

            lei = doc.get("ae_lei", "").strip() or None
            if lei and len(lei) < 10:
                lei = None

            nca_name = doc.get("ae_competentAuthority", "")
            nca_code = _esma_nca_to_code(nca_name)
            country = _esma_country_to_code(doc.get("ae_homeMemberState", ""))
            status = doc.get("ae_status", "Active")
            # ESMA does not provide the actual auth status (authorised,
            # registered, NPPR).  Only mark clearly inactive entities;
            # leave active ones as None so NCA-provided statuses are
            # not overwritten.
            auth_status = "INACTIVE" if status == "Inactive" else None
            esma_address = doc.get("ae_headOfficeAddress", "").strip() or None

            enrich = country in store.LOCAL_NCA_COUNTRIES

            esma_id = doc.get("id", "").strip()
            rec = AIFMRecord(
                aifm_name=name,
                lei=lei,
                home_country=country if country else None,
                address=esma_address,
                address_source="ESMA" if esma_address else None,
                nca_code=nca_code,
                nca_entity_code=esma_id or None,
                auth_status=auth_status,
                auth_date=_parse_date(doc.get("ae_authorisationNotificationDateStr")),
                withdrawal_date=_parse_date(doc.get("ae_authorisationWithdrawalEndDateStr")),
                source="ESMA",
                source_entity_id=esma_id or None,
            )
            mid = store.upsert_aifm(rec, enrich_only=enrich)
            if mid:
                result["aifm"] += 1
            else:
                result.setdefault("aifm_skipped", 0)
                result["aifm_skipped"] += 1

        start += page_size
        if start % 2000 == 0:
            logger.info(f"  ... processed {start}/{total} AIFMs")
            store.commit()

    store.commit()
    logger.info(f"ESMA AIFMs complete: {result['aifm']}")

    # ESMA fund fetching — active parent AIFs only; enrich-only for local NCA countries
    logger.info("Fetching ESMA AIF register (all EU, active parents only)...")
    start = 0
    total = None
    while total is None or start < total:
        params = (
            f"q=*:*&fq=funds_legal_framework_name:AIF"
            f"&fq=type_s:parent&fq=funds_status_code:ACTV"
            f"&wt=json&rows={page_size}&start={start}&fl=*"
        )
        url = f"{ESMA_FUND_URL}?{params}"
        raw = _fetch_url(url)
        data = json.loads(raw)
        total = data["response"]["numFound"]
        docs = data["response"]["docs"]

        if start == 0:
            logger.info(f"ESMA AIFs: {total} total entities")

        for doc in docs:
            fund_name = doc.get("funds_national_name", "").strip()
            if not fund_name:
                continue
            mgr_name = doc.get("funds_manager_nat_name", "").strip()
            nca_name = doc.get("funds_manager_comp_cou_code", "")
            nca_code = _esma_nca_to_code(nca_name)
            status = doc.get("funds_status_code_name", "Active")
            # ESMA does not provide actual auth status; leave as None.
            auth_status = "INACTIVE" if status == "Inactive" else None
            home = (
                _esma_country_to_code(doc.get("funds_home_member_state", ""))
                or doc.get("funds_ca_cou_code", "").strip().upper() or None
            )

            enrich = home in store.LOCAL_NCA_COUNTRIES

            esma_id = doc.get("id", "").strip()
            fund_rec = AIFRecord(
                aif_type="AIF",
                aif_name=fund_name,
                home_country=home if home else None,
                nca_code=nca_code,
                nca_entity_code=esma_id or None,
                auth_status=auth_status,
                source="ESMA",
                source_entity_id=esma_id or None,
                managing_aifm_name=mgr_name,
            )
            fid = store.upsert_aif(fund_rec, enrich_only=enrich)
            if fid:
                result["aif"] += 1
                if mgr_name:
                    store.link_aif_to_aifm(fid, mgr_name, home)
            else:
                result.setdefault("aif_skipped", 0)
                result["aif_skipped"] += 1

        start += page_size
        if start % 2000 == 0:
            logger.info(f"  ... processed {start}/{total} AIFs")
            store.commit()

    store.commit()
    logger.info(f"ESMA AIFs complete: {result['aif']}")

    logger.info(f"ESMA fetch complete: {result}")
    return result


def cleanup_esma_branches(store: NCARegisterStore) -> dict:
    """Remove ESMA branch records from nca_aifm and nca_registrations.

    Strategy:
      1. Fetch head-office ae_entityName set from ESMA (quoted Solr value).
      2. Fetch branch ae_entityName set from ESMA.
      3. branch_only = branch_names − head_names.
      4. Delete from nca_aifm where aifm_name matches a branch_only name.
    """

    def _collect_names(office_type: str) -> tuple[set[str], int]:
        """Paginate ESMA and collect ae_entityName values for an office type."""
        names: set[str] = set()
        start, total, page_size = 0, None, 500
        # Solr needs quoted multi-word values: "head office" → %22head%20office%22
        fq_val = f"%22{office_type.replace(' ', '%20')}%22"
        while total is None or start < total:
            params = (
                f"q=*:*&fq=type_s:parent&fq=ae_entityTypeCode:aif"
                f"&fq=ae_officeType:{fq_val}"
                f"&wt=json&rows={page_size}&start={start}&fl=ae_entityName"
            )
            url = f"{ESMA_AIFM_URL}?{params}"
            raw = _fetch_url(url)
            data = json.loads(raw)
            total = data["response"]["numFound"]
            for doc in data["response"]["docs"]:
                name = (doc.get("ae_entityName") or "").strip()
                if name:
                    names.add(name)
            start += page_size
        return names, total or 0

    # ── Step 1 + 2: collect names ────────────────────────────────────
    logger.info("Collecting ESMA head-office names...")
    head_names, head_total = _collect_names("head office")
    logger.info("  %d unique names (from %d records)", len(head_names), head_total)

    logger.info("Collecting ESMA branch names...")
    branch_names, branch_total = _collect_names("branch")
    logger.info("  %d unique names (from %d records)", len(branch_names), branch_total)

    # ── Step 3: branch-only names ────────────────────────────────────
    branch_only = branch_names - head_names
    shared = branch_names & head_names
    logger.info("  %d branch-only names, %d shared with head offices (safe)",
                len(branch_only), len(shared))

    if not branch_only:
        logger.info("No branch-only names found — nothing to delete")
        return {"head_offices": len(head_names), "branches": len(branch_names),
                "branch_only": 0, "aifm_deleted": 0, "registrations_deleted": 0,
                "aif_deleted": 0}

    # ── Step 4: match against DB and delete ──────────────────────────
    cur = store._conn.cursor()

    # Diagnostic: verify table is accessible
    cur.execute("SELECT COUNT(*) FROM platform.nca_aifm")
    total_aifm = cur.fetchone()[0]
    logger.info("  DB check: nca_aifm has %d records", total_aifm)

    cur.execute("SELECT aifm_name FROM platform.nca_aifm LIMIT 3")
    sample_db = [r[0] for r in cur.fetchall()]
    logger.info("  DB sample names: %s", sample_db)

    # Try one specific branch name to debug matching
    test_name = sorted(branch_only)[0]
    logger.info("  Testing exact match for: %r", test_name)
    cur.execute(
        "SELECT aifm_id, aifm_name FROM platform.nca_aifm WHERE aifm_name = %s",
        (test_name,),
    )
    test_rows = cur.fetchall()
    logger.info("  Result: %d rows", len(test_rows))
    if not test_rows:
        # Try LIKE to see if encoding differs
        cur.execute(
            "SELECT aifm_id, aifm_name FROM platform.nca_aifm WHERE aifm_name LIKE %s LIMIT 3",
            (f"%{test_name[:20]}%",),
        )
        like_rows = cur.fetchall()
        logger.info("  LIKE '%s%%' found %d rows", test_name[:20], len(like_rows))
        for r in like_rows:
            logger.info("    %r", r[1])

    aifm_ids_to_delete = []

    for name in branch_only:
        cur.execute("""
            SELECT aifm_id, aifm_name, home_country
            FROM platform.nca_aifm WHERE aifm_name = %s
        """, (name,))
        for row in cur.fetchall():
            aifm_ids_to_delete.append(row)

    logger.info("Matched %d AIFM records in database", len(aifm_ids_to_delete))
    for aifm_id, aname, country in aifm_ids_to_delete[:20]:
        logger.info("  [%s] %s", country or "??", aname)
    if len(aifm_ids_to_delete) > 20:
        logger.info("  ... and %d more", len(aifm_ids_to_delete) - 20)

    regs_deleted = 0
    aif_deleted = 0
    aifm_deleted = 0

    for aifm_id, aifm_name, _ in aifm_ids_to_delete:
        cur.execute("""
            DELETE FROM platform.nca_registrations
            WHERE entity_type = 'AIFM' AND entity_id = %s
        """, (aifm_id,))
        regs_deleted += cur.rowcount

        cur.execute(
            "DELETE FROM platform.nca_aif WHERE managing_aifm_id = %s",
            (aifm_id,),
        )
        aif_deleted += cur.rowcount

        cur.execute(
            "DELETE FROM platform.nca_aifm WHERE aifm_id = %s",
            (aifm_id,),
        )
        aifm_deleted += cur.rowcount

    store.commit()
    result = {
        "head_offices": len(head_names),
        "branches": len(branch_names),
        "branch_only": len(branch_only),
        "db_matches": len(aifm_ids_to_delete),
        "aifm_deleted": aifm_deleted,
        "registrations_deleted": regs_deleted,
        "aif_deleted": aif_deleted,
    }
    logger.info("Branch cleanup complete: %s", result)
    return result


# ── NCA / country code mapping helpers ───────────────────────────────

_NCA_MAP = {
    "Netherlands Authority for the Financial Markets (AFM)": "AFM",
    "Autoriteit Financiële Markten (AFM)": "AFM",
    "De Nederlandsche Bank": "DNB",
    "Commission de Surveillance du Secteur Financier (CSSF)": "CSSF",
    "Autorité des Marchés Financiers (AMF)": "AMF",
    "Autorité de Contrôle Prudentiel et de Résolution (ACPR)": "ACPR",
    "Autorité de contrôle prudentiel": "ACPR",
    "Federal Financial Supervisory Authority (BaFin)": "BAFIN",
    "Bundesanstalt für Finanzdienstleistungsaufsicht (BaFin)": "BAFIN",
    "Central Bank of Ireland (CBI)": "CBI",
    "Financial Conduct Authority (FCA)": "FCA",
    "Finansinspektionen (FI)": "FI_SE",
    "Austrian Financial Market Authority (FMA)": "FMA_AT",
    "Finanzmarktaufsicht (FMA)": "FMA_AT",
    "Malta Financial Services Authority (MFSA)": "MFSA",
    "Commissione Nazionale per le Società e la Borsa (CONSOB)": "CONSOB",
    "Commissione Nazionale per le Societa e la Borsa (CONSOB)": "CONSOB",
    "Banca d'Italia": "BANCA_IT",
    "BANCA D'ITALIA": "BANCA_IT",
    "Comisión Nacional del Mercado de Valores (CNMV)": "CNMV",
    "Bank of Spain": "BDE",
    "Danish Financial Supervisory Authority (Finanstilsynet)": "DFSA",
    "Finanstilsynet": "DFSA",
    "Financial Supervisory Authority (FIN-FSA)": "FIN_FSA",
    "Finanssivalvonta (FSA)": "FIN_FSA",
    "Hellenic Capital Market Commission (HCMC)": "HCMC",
    "Czech National Bank (CNB)": "CNB",
    "National Bank of Belgium (NBB)": "NBB",
    "National Bank of Belgium": "NBB",
    "Financial Services and Markets Authority (FSMA)": "FSMA",
    "Cyprus Securities and Exchange Commission (CySEC)": "CYSEC",
    "Comissão do Mercado de Valores Mobiliários (CMVM)": "CMVM",
    "Banco de Portugal": "BDP",
    "Komisja Nadzoru Finansowego (KNF)": "KNF",
    "Polish Financial Supervisory Authority (KNF)": "KNF",
    "Hungarian National Bank (MNB)": "MNB",
    "Central Bank of Hungary": "MNB",
    "Hungarian Financial Supervisory Authority (PSZAF)": "MNB",
    "Financial Supervisory Authority of Norway (Finanstilsynet)": "NFSA",
    "Norwegian Financial Supervisory Authority": "NFSA",
    "National Bank of Slovakia (NBS)": "NBS",
    "Croatian Financial Services Supervisory Agency (HANFA)": "HANFA",
    "Romanian Financial Supervisory Authority": "ASF_RO",
    "Estonian Financial Supervision Authority (EFSA)": "EFSA",
    "Iceland Financial Supervisory Authority (FME)": "FME",
    "Bank of Lithuania (LSC)": "LB_LT",
    "Latvijas Banka": "LB_LV",
    "Financial Supervision Commission (FSC)": "FSC_BG",
    "Securities Market Agency (ATVP)": "ATVP",
}

def _esma_nca_to_code(nca_name: str) -> str:
    nca_name = nca_name.strip()
    if not nca_name:
        return "UNKNOWN"
    if nca_name in _NCA_MAP:
        return _NCA_MAP[nca_name]
    nca_lower = nca_name.lower()
    for key, code in _NCA_MAP.items():
        if key.lower() == nca_lower:
            return code
    for key, code in _NCA_MAP.items():
        if key.lower() in nca_lower or nca_lower in key.lower():
            return code
    m = re.search(r"\(([A-Za-z]{2,10})\)", nca_name)
    if m:
        return m.group(1).upper()
    logger.warning(f"Unmapped NCA authority: '{nca_name}' — using truncated fallback")
    return nca_name[:10] if nca_name else "UNKNOWN"

_COUNTRY_MAP = {
    "AUSTRIA": "AT", "BELGIUM": "BE", "BULGARIA": "BG", "CROATIA": "HR",
    "CYPRUS": "CY", "CZECHIA": "CZ", "CZECH REPUBLIC": "CZ",
    "DENMARK": "DK", "ESTONIA": "EE", "FINLAND": "FI", "FRANCE": "FR",
    "GERMANY": "DE", "GREECE": "GR", "HUNGARY": "HU", "ICELAND": "IS",
    "IRELAND": "IE", "ITALY": "IT", "LATVIA": "LV", "LIECHTENSTEIN": "LI",
    "LITHUANIA": "LT", "LUXEMBOURG": "LU", "MALTA": "MT",
    "NETHERLANDS": "NL", "NORWAY": "NO", "POLAND": "PL", "PORTUGAL": "PT",
    "ROMANIA": "RO", "SLOVAKIA": "SK", "SLOVENIA": "SI", "SPAIN": "ES",
    "SWEDEN": "SE", "UNITED KINGDOM": "GB",
    "Nederland": "NL", "Luxemburg": "LU",
}

def _esma_country_to_code(country_name: str) -> Optional[str]:
    if not country_name:
        return None
    country_name = country_name.strip()
    if len(country_name) == 2:
        return country_name.upper()
    return _COUNTRY_MAP.get(country_name.upper(), _COUNTRY_MAP.get(country_name))

def _country_code(name: str) -> Optional[str]:
    return _esma_country_to_code(name)


# ── CLI ──────────────────────────────────────────────────────────────

def main():
    import psycopg2

    if len(sys.argv) < 2:
        print("Usage: python fetch_nca_registers.py <command>")
        print("Commands: fetch-afm, fetch-cssf, fetch-esma, fetch-all, enrich-gleif, enrich-lei, detect-terminations, diagnose-branches, cleanup-branches, report")
        sys.exit(1)

    cmd = sys.argv[1].lower()

    conn = psycopg2.connect(
        host=PG_HOST, port=PG_PORT, dbname=PG_DBNAME,
        user=PG_USER, password=PG_PASSWORD,
    )
    store = NCARegisterStore(conn)

    try:
        if cmd == "fetch-afm":
            result = fetch_afm(store)
            print(f"\nAFM: {result}")
        elif cmd == "fetch-cssf":
            result = fetch_cssf(store)
            print(f"\nCSSF: {result}")
        elif cmd == "fetch-esma":
            result = fetch_esma(store)
            print(f"\nESMA: {result}")
        elif cmd == "fetch-all":
            print("=" * 60)
            print("  Fetching all NCA registers")
            print("  Order: local NCA first (high prio), ESMA last (gap-fill)")
            print("  Then: LEI enrichment → GLEIF enrichment → termination detection")
            print("=" * 60)

            # Phase 1: Fetch from all sources
            r1 = fetch_afm(store)
            r2 = fetch_cssf(store)
            r3 = fetch_esma(store)
            print(f"\nAFM:  {r1}")
            print(f"CSSF: {r2}")
            print(f"ESMA: {r3}")

            # Phase 2: LEI enrichment (name→LEI lookup for entities without LEI)
            print("\n" + "-" * 60)
            print("  Phase 2: LEI enrichment from GLEIF cache")
            print("-" * 60)
            lei_result = store.enrich_lei_from_gleif()
            print(f"  AIFM: {lei_result['aifm_lei_found']} found, {lei_result['aifm_no_match']} no match, {lei_result['aifm_multi_match']} multi-match")
            print(f"  AIF:  {lei_result['aif_lei_found']} found, {lei_result['aif_no_match']} no match, {lei_result['aif_multi_match']} multi-match")

            # Phase 3: GLEIF enrichment (address, home_country, registered_as)
            print("\n" + "-" * 60)
            print("  Phase 3: GLEIF enrichment (address, home_country, registered_as)")
            print("-" * 60)
            gleif_result = store.enrich_from_gleif()
            print(f"  GLEIF enrichment: {gleif_result}")
            print("  Inheriting AIFM address to AIFs...")
            inherit_count = store.inherit_aifm_address()
            print(f"  AIFs with address inherited from AIFM: {inherit_count}")

            # Phase 4: NPPR status propagation
            print("\n" + "-" * 60)
            print("  Phase 4: NPPR status propagation")
            print("-" * 60)
            nppr_result = store.propagate_nppr_status()
            print(f"  AIFM registrations set to NPPR: {nppr_result['aifm_propagated']}")
            print(f"  AIF  registrations set to NPPR: {nppr_result['aif_propagated']}")

            # Phase 5: Termination detection
            print("\n" + "-" * 60)
            print("  Phase 5: Termination detection")
            print("-" * 60)
            term_result = store.detect_terminations()
            print(f"  AIFM terminated (source data): {term_result['aifm_from_source']}")
            print(f"  AIF  terminated (source data): {term_result['aif_from_source']}")
            print(f"  AIFM terminated (staleness):   {term_result['aifm_from_staleness']}")
            print(f"  AIF  terminated (staleness):   {term_result['aif_from_staleness']}")
            print(f"  AIFM cleared (reappeared):     {term_result['aifm_cleared']}")
            print(f"  AIF  cleared (reappeared):     {term_result['aif_cleared']}")
        elif cmd == "enrich-gleif":
            print("Enriching NCA entities from GLEIF LEI cache...")
            result = store.enrich_from_gleif()
            print(f"\nGLEIF enrichment: {result}")
            print("\nInheriting AIFM address to AIFs...")
            count = store.inherit_aifm_address()
            print(f"AIFs with address inherited from AIFM: {count}")
        elif cmd == "enrich-lei":
            print("Looking up LEI by name + country in GLEIF cache...")
            result = store.enrich_lei_from_gleif()
            print(f"\nLEI enrichment:")
            print(f"  AIFM: {result['aifm_lei_found']} found, {result['aifm_no_match']} no match, {result['aifm_multi_match']} multi-match")
            print(f"  AIF:  {result['aif_lei_found']} found, {result['aif_no_match']} no match, {result['aif_multi_match']} multi-match")
        elif cmd == "detect-terminations":
            staleness = 30
            if len(sys.argv) > 2:
                staleness = int(sys.argv[2])
            print(f"Detecting terminations (staleness threshold: {staleness} days)...")
            result = store.detect_terminations(staleness_days=staleness)
            print(f"\nTermination detection:")
            print(f"  AIFM terminated (source data): {result['aifm_from_source']}")
            print(f"  AIF  terminated (source data): {result['aif_from_source']}")
            print(f"  AIFM terminated (staleness):   {result['aifm_from_staleness']}")
            print(f"  AIF  terminated (staleness):   {result['aif_from_staleness']}")
            print(f"  AIFM cleared (reappeared):     {result['aifm_cleared']}")
            print(f"  AIF  cleared (reappeared):     {result['aif_cleared']}")
        elif cmd == "purge-esma-regs":
            cur = conn.cursor()
            cur.execute("DELETE FROM platform.nca_registrations WHERE source = 'ESMA'")
            print(f"Deleted {cur.rowcount:,} ESMA registrations")
            conn.commit()

        elif cmd == "diagnose-branches":
            print("=== BRANCH DIAGNOSIS ===\n")
            cur = conn.cursor()

            # 1. What sources exist in nca_registrations?
            cur.execute("""
                SELECT source, entity_type, COUNT(*)
                FROM platform.nca_registrations
                GROUP BY source, entity_type
                ORDER BY source, entity_type
            """)
            rows = cur.fetchall()
            print("1. nca_registrations breakdown:")
            for src, etype, cnt in rows:
                print(f"   {src:8s} / {etype:8s}: {cnt:,}")

            # 2. Sample source_entity_ids per source
            for src in set(r[0] for r in rows):
                cur.execute("""
                    SELECT source_entity_id FROM platform.nca_registrations
                    WHERE source = %s AND source_entity_id IS NOT NULL
                    LIMIT 5
                """, (src,))
                ids = [r[0] for r in cur.fetchall()]
                print(f"   Sample IDs for {src}: {ids}")

            # 3. Total AIFMs and sample names
            cur.execute("SELECT COUNT(*) FROM platform.nca_aifm")
            print(f"\n2. Total AIFMs: {cur.fetchone()[0]:,}")

            # 4. AIFMs with branch-like words in aifm_name (not clean_name)
            cur.execute("""
                SELECT aifm_id, aifm_name, aifm_clean_name, home_country, source
                FROM platform.nca_aifm
                WHERE UPPER(aifm_name) LIKE '%BRANCH%'
                   OR UPPER(aifm_name) LIKE '%SUCCURSALE%'
                   OR UPPER(aifm_name) LIKE '%SUCURSAL%'
                   OR UPPER(aifm_name) LIKE '%ZWEIGNIEDERLASSUNG%'
                   OR UPPER(aifm_name) LIKE '%FILIALE%'
                   OR UPPER(aifm_name) LIKE '%BIJKANTOOR%'
                LIMIT 20
            """)
            branch_rows = cur.fetchall()
            print(f"\n3. AIFMs with branch keywords in aifm_name: {len(branch_rows)}")
            for aid, name, cn, country, src in branch_rows:
                print(f"   [{country or '??'}] id={aid} src={src}")
                print(f"         name: {name}")
                print(f"   clean_name: {cn}")

            # 5. AIFMs with branch-like words in aifm_clean_name
            cur.execute("""
                SELECT COUNT(*) FROM platform.nca_aifm
                WHERE aifm_clean_name LIKE '%BRANCH%'
                   OR aifm_clean_name LIKE '%SUCCURSALE%'
                   OR aifm_clean_name LIKE '%SUCURSAL%'
                   OR aifm_clean_name LIKE '%ZWEIGNIEDERLASSUNG%'
                   OR aifm_clean_name LIKE '%FILIALE%'
                   OR aifm_clean_name LIKE '%BIJKANTOOR%'
            """)
            print(f"\n4. AIFMs with branch keywords in aifm_clean_name: {cur.fetchone()[0]}")

            # 6. Check if 'source' column exists on nca_aifm
            cur.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_schema = 'platform' AND table_name = 'nca_aifm'
                ORDER BY ordinal_position
            """)
            cols = [r[0] for r in cur.fetchall()]
            print(f"\n5. nca_aifm columns: {cols}")

            # 7. Sample of Tikehau records specifically
            cur.execute("""
                SELECT aifm_id, aifm_name, aifm_clean_name, home_country
                FROM platform.nca_aifm
                WHERE aifm_clean_name LIKE '%TIKEHAU%'
            """)
            tik = cur.fetchall()
            print(f"\n6. Tikehau records ({len(tik)}):")
            for aid, name, cn, country in tik:
                print(f"   [{country or '??'}] id={aid}")
                print(f"         name: {name}")
                print(f"   clean_name: {cn}")

        elif cmd == "cleanup-branches":
            print("Removing ESMA branch records from database...")
            result = cleanup_esma_branches(store)
            print(f"\nBranch cleanup: {result}")
        elif cmd == "report":
            report = store.get_report()
            print("\n=== NCA Register Report ===")
            print(f"\nAIFMs: {report['aifm']['total']:,} total, "
                  f"{report['aifm']['active']:,} active, "
                  f"{report['aifm']['with_lei']:,} with LEI")
            print("\nAIFs:")
            for ftype, counts in report["aif"].items():
                print(f"  {ftype:10s}: {counts['total']:>6,} total, {counts['active']:>6,} active, {counts['with_lei']:>6,} with LEI")
            print("\nRegistrations by source:")
            for reg in report["registrations"]:
                print(f"  {reg['source']:6s} / {reg['nca_code']:8s} / {reg['entity_type']:8s}: "
                      f"{reg['total']:>6,} total, {reg['active']:>6,} active")
            print(f"\nEntities with multiple NCAs: {report['multi_nca_count']}")
            print("\nTop 20 AIFMs by AIF count:")
            for mgr in report["top_aifm"]:
                print(f"  {mgr['country'] or '??':2s}  {mgr['aif_count']:>4d}  {mgr['name']}")
        else:
            print(f"Unknown command: {cmd}")
            sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()

"""Export NCA register tables to Excel.

Produces one workbook with sheets:
  - AIFM           — platform.nca_aifm (all columns)
  - AIF            — platform.nca_aif (all columns, managing AIFM ID + parent AIF ID)
  - NCA Registrations — platform.nca_registrations (all columns, entity_id for lookup)
  - Summary        — quick stats

Usage:
    python Application/reference_data/export_nca_register.py
    python Application/reference_data/export_nca_register.py --output my_export.xlsx
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from reference_data.config import PG_HOST, PG_PORT, PG_DBNAME, PG_USER, PG_PASSWORD

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))

QUERIES = {
    "AIFM": """
        SELECT m.aifm_id, m.aifm_name, m.aifm_clean_name, m.lei,
               m.home_country,
               m.address, m.address_source,
               m.registered_as,
               m.first_seen_at, m.last_seen_at,
               m.created_at, m.last_updated, m.deleted_at,
               m.lineage
        FROM platform.nca_aifm m
        ORDER BY m.home_country, m.aifm_name
    """,
    "AIF": """
        SELECT f.aif_id, f.aif_type, f.aif_name, f.aif_clean_name, f.lei, f.isin,
               f.home_country,
               f.custodian, f.fund_strategy,
               f.address, f.address_source,
               f.managing_aifm_id,
               f.parent_aif_id,
               f.first_seen_at, f.last_seen_at,
               f.created_at, f.last_updated, f.deleted_at,
               f.lineage
        FROM platform.nca_aif f
        ORDER BY f.home_country, f.aif_type, f.aif_name
    """,
    "NCA Registrations": """
        SELECT r.registration_id, r.entity_type, r.entity_id,
               r.nca_code, r.nca_entity_code,
               r.auth_status, r.auth_date, r.withdrawal_date,
               r.source, r.source_entity_id,
               r.fetched_at, r.first_seen_at, r.last_seen_at, r.deleted_at,
               r.lineage
        FROM platform.nca_registrations r
        ORDER BY r.source, r.nca_code, r.entity_type
    """,
}

FRIENDLY_HEADERS = {
    "AIFM": [
        "AIFM ID", "AIFM Name", "AIFM Clean Name", "LEI",
        "Home Country",
        "Address", "Address Source",
        "Registered As",
        "First Seen", "Last Seen",
        "Created", "Last Updated", "Deleted",
        "Lineage",
    ],
    "AIF": [
        "AIF ID", "Type", "AIF Name", "AIF Clean Name", "LEI", "ISIN",
        "Home Country",
        "Custodian", "Fund Strategy",
        "Address", "Address Source",
        "Managing AIFM ID",
        "Parent AIF ID",
        "First Seen", "Last Seen",
        "Created", "Last Updated", "Deleted",
        "Lineage",
    ],
    "NCA Registrations": [
        "Registration ID", "Entity Type", "Entity ID",
        "NCA Code", "NCA Entity Code",
        "Auth Status", "Auth Date", "Withdrawal Date",
        "Source", "Source Entity ID",
        "Fetched At", "First Seen", "Last Seen", "Deleted",
        "Lineage",
    ],
}

COL_WIDTHS = {
    "AIFM": {1: 12, 2: 45, 3: 40, 4: 24, 5: 8, 6: 50, 7: 12, 8: 20, 14: 60},
    "AIF": {1: 12, 2: 10, 3: 45, 4: 40, 5: 24, 6: 14, 7: 8, 8: 35, 9: 14, 10: 50, 11: 12, 12: 12, 13: 12, 19: 60},
    "NCA Registrations": {1: 12, 2: 10, 3: 12, 4: 10, 5: 22, 6: 14, 7: 14, 15: 60},
    "Summary": {1: 30, 2: 12, 3: 3, 4: 12, 5: 12, 6: 12, 7: 14, 8: 12},
}


def _style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

def _style_data(ws, row_count, col_count):
    for row in range(2, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            if isinstance(cell.value, datetime):
                cell.number_format = "YYYY-MM-DD HH:MM"

def _set_widths(ws, sheet_name):
    widths = COL_WIDTHS.get(sheet_name, {})
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

def _add_autofilter(ws, col_count):
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}1"
    ws.freeze_panes = "A2"


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Export NCA register to Excel")
    parser.add_argument("--output", "-o", default=None, help="Output file path")
    args = parser.parse_args()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output = args.output or f"nca_register_export_{timestamp}.xlsx"

    print(f"Connecting to PostgreSQL ({PG_HOST}:{PG_PORT}/{PG_DBNAME})...")
    conn = psycopg2.connect(
        host=PG_HOST, port=PG_PORT, dbname=PG_DBNAME,
        user=PG_USER, password=PG_PASSWORD,
    )
    cur = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)

    stats = {}

    for sheet_name, query in QUERIES.items():
        print(f"  Exporting {sheet_name}...")
        cur.execute(query)
        rows = cur.fetchall()
        headers = FRIENDLY_HEADERS[sheet_name]

        ws = wb.create_sheet(sheet_name)
        ws.append(headers)

        for row in rows:
            ws.append([
                v.replace(tzinfo=None) if isinstance(v, datetime) else v
                for v in row
            ])

        col_count = len(headers)
        row_count = len(rows) + 1
        _style_header(ws, col_count)
        _style_data(ws, row_count, col_count)
        _set_widths(ws, sheet_name)
        _add_autofilter(ws, col_count)
        stats[sheet_name] = len(rows)
        print(f"    {len(rows):,} rows")

    # Summary sheet
    ws = wb.create_sheet("Summary", 0)
    ws.append(["NCA Register Export", ""])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append(["", ""])
    ws.append(["Table", "Row Count"])
    for name, count in stats.items():
        ws.append([name, count])
    ws.append(["", ""])

    # ── Country summary by NCA country with AIFM + AIF breakdown ──
    # Uses nca_code to determine country (the NCA's jurisdiction),
    # not home_country (the entity's legal domicile).
    # Each entity is counted once per NCA, using its highest-priority
    # registration status (AUTHORISED > REGISTERED > NPPR > unknown).
    cur.execute("""
        WITH nca_country_map(nca_code, country) AS (VALUES
            ('AFM', 'NL'), ('CSSF', 'LU'), ('BAFIN', 'DE'), ('AMF', 'FR'),
            ('CONSOB', 'IT'), ('CNMV', 'ES'), ('FCA', 'GB'), ('CBI', 'IE'),
            ('FMA', 'AT'), ('FINMA', 'CH'), ('FSMA', 'BE'), ('CMVM', 'PT'),
            ('FIN-FSA', 'FI'), ('DFSA', 'DK'), ('FI', 'SE'), ('FSA', 'NO'),
            ('PFSA', 'PL'), ('CNB', 'CZ'), ('MNB', 'HU'), ('MFSA', 'MT'),
            ('CYSEC', 'CY'), ('FSC', 'BG'), ('HANFA', 'HR'), ('ASF', 'RO'),
            ('FMA-LI', 'LI'), ('GFSC', 'GG'), ('JFSC', 'JE'), ('FSA-IS', 'IS'),
            ('ATVP', 'SI'), ('NBS', 'SK'), ('FCMC', 'LV'), ('LB', 'LT'),
            ('EFSA', 'EE'), ('FI-IS', 'IS'), ('IOMFSA', 'IM')
        ),
        -- Determine best status per entity per NCA country
        reg_best AS (
            SELECT r.entity_id, r.entity_type,
                   COALESCE(nc.country, r.nca_code) AS nca_country,
                   -- Pick highest-priority status per entity+NCA
                   MIN(CASE r.auth_status
                       WHEN 'AUTHORISED' THEN 1
                       WHEN 'REGISTERED' THEN 2
                       WHEN 'NPPR'       THEN 3
                       ELSE 4
                   END) AS status_rank
            FROM platform.nca_registrations r
            LEFT JOIN nca_country_map nc ON nc.nca_code = r.nca_code
            WHERE r.deleted_at IS NULL
            GROUP BY r.entity_id, r.entity_type,
                     COALESCE(nc.country, r.nca_code)
        ),
        aifm_counts AS (
            SELECT rb.nca_country AS country,
                   COUNT(DISTINCT rb.entity_id) AS aifm_total
            FROM reg_best rb
            WHERE rb.entity_type = 'AIFM'
            GROUP BY rb.nca_country
        ),
        aif_counts AS (
            SELECT rb.nca_country AS country,
                   COUNT(DISTINCT rb.entity_id) FILTER (WHERE rb.status_rank = 1) AS aif_authorised,
                   COUNT(DISTINCT rb.entity_id) FILTER (WHERE rb.status_rank = 2) AS aif_registered,
                   COUNT(DISTINCT rb.entity_id) FILTER (WHERE rb.status_rank = 3) AS aif_nppr,
                   COUNT(DISTINCT rb.entity_id) FILTER (WHERE rb.status_rank = 4) AS aif_unknown,
                   COUNT(DISTINCT rb.entity_id) AS aif_total
            FROM reg_best rb
            WHERE rb.entity_type IN ('AIF', 'SUB_AIF')
            GROUP BY rb.nca_country
        )
        SELECT COALESCE(a.country, f.country) AS country,
               COALESCE(a.aifm_total, 0),
               COALESCE(f.aif_authorised, 0),
               COALESCE(f.aif_registered, 0),
               COALESCE(f.aif_nppr, 0),
               COALESCE(f.aif_unknown, 0),
               COALESCE(f.aif_total, 0)
        FROM aifm_counts a
        FULL OUTER JOIN aif_counts f ON a.country = f.country
        ORDER BY COALESCE(a.aifm_total, 0) DESC
    """)
    country_header_row = ws.max_row + 1
    ws.append(["Country", "AIFMs", "",
               "Auth AIFs", "Reg AIFs", "NPPR AIFs", "Unknown AIFs", "Total AIFs"])
    for row_data in cur.fetchall():
        ws.append(list(row_data))
    ws.append(["", ""])

    cur.execute("""
        SELECT aif_type, COUNT(*) FROM platform.nca_aif
        WHERE deleted_at IS NULL GROUP BY aif_type ORDER BY aif_type
    """)
    ws.append(["AIFs by Type", "Count"])
    for ftype, cnt in cur.fetchall():
        ws.append([ftype, cnt])
    ws.append(["", ""])

    cur.execute("""
        SELECT source, COUNT(*) FROM platform.nca_registrations
        GROUP BY source ORDER BY source
    """)
    ws.append(["Registrations by Source", "Count"])
    for src, cnt in cur.fetchall():
        ws.append([src, cnt])

    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
    for row in ws.iter_rows(min_row=4, max_row=4, max_col=2):
        for cell in row:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
    # Style country summary header
    for row in ws.iter_rows(min_row=country_header_row, max_row=country_header_row, max_col=8):
        for cell in row:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
    _set_widths(ws, "Summary")

    cur.close()
    conn.close()

    wb.save(output)
    print(f"\nExported to: {output}")
    print(f"  AIFM:               {stats.get('AIFM', 0):>7,}")
    print(f"  AIF:                {stats.get('AIF', 0):>7,}")
    print(f"  NCA Registrations:  {stats.get('NCA Registrations', 0):>7,}")


if __name__ == "__main__":
    main()

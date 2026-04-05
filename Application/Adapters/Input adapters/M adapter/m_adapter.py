"""
M adapter AIFMD Annex IV Adapter - Light & Full Templates
==========================================================
Reads M adapter Excel templates and generates AIFMD Annex IV
XML reports (AIFM-level and AIF-level) conforming to XSD v1.2.

Supported template types:
  - AIFM_LIGHT / AIF_LIGHT   → registered (below-threshold)
  - AIFM / AIF               → authorised (full templates)

Usage:
    from m_adapter import MAdapter
    adapter = MAdapter("template.xlsx")
    adapter.generate_aifm_xml("output_aifm.xml")
    adapter.generate_aif_xml("output_aif.xml")
    adapter.generate_all()  # or with reporting_member_state="NL"
"""

import csv
import gzip
import logging
import math
import re
import urllib.request
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional
from xml.dom.minidom import parseString
from xml.etree.ElementTree import Element, SubElement, tostring

import openpyxl
import yaml

# ── Ensure shared module can be imported from parent directory ────────────────
import sys
from pathlib import Path as _Path
_app_root = _Path(__file__).resolve().parent.parent.parent.parent  # Application/
if str(_app_root) not in sys.path:
    sys.path.insert(0, str(_app_root))
# Also add M adapter dir itself for m_parser subpackage
_m_adapter_dir = _Path(__file__).resolve().parent
if str(_m_adapter_dir) not in sys.path:
    sys.path.insert(0, str(_m_adapter_dir))

from shared.constants import EEA_COUNTRIES, REGION_MAP
from shared.aifmd_constants import (
    STRATEGY_TO_AIF_TYPE, STRATEGY_ELEMENT_MAP,
    XSD_VERSION, AIFM_XSD, AIF_XSD, XSI_NS,
    FCA_VERSION, FCA_AIFM_NS, FCA_AIF_NS, FCA_REGION_MAP,
    _ESMA_TO_FCA_FIELD, FCA_AGG_VALUE_MIN_PCT, FCA_SOVEREIGN_TURNOVER_MAP,
)
from shared.reference_data import _COUNTRY_TO_REGION
from shared.aifmd_reference_data import _SUBASSET_TO_TURNOVER
from shared.formatting import (
    _str, _int_round, _float_val, _rate_fmt, _bool_str, _date_str,
    _is_eea, _sub, _macro_type, _asset_type, _turnover_sub_asset_type,
    _predominant_type, _pretty_xml, _reporting_period_dates, _fca_sovereign_sub_asset,
)

log = logging.getLogger(__name__)

# ── M-adapter-specific parser models ────────────────────────────────────────
from m_parser.schema_loader import _load_column_schema, _COLUMN_SCHEMA
from m_parser.record import MRecord


# ── Constants ────────────────────────────────────────────────────────────────
# Extracted to shared/constants.py, shared/aifmd_constants.py,
# shared/reference_data.py, and shared/aifmd_reference_data.py


# ── Helper functions ─────────────────────────────────────────────────────────
# Extracted to shared/formatting.py

# ── Derivation engine (period derivation, FX rates, portfolio, turnover, investors)
from derivation.aifmd_period import _derive_aif_period
from derivation.fx_service import (
    _fetch_ecb_fx_rate, _get_ecb_fallback_rate,
    _ECB_FX_CACHE, _ECB_FALLBACK_RATES,
)
from derivation.aifmd_portfolio import (
    pos_val as _pos_val_fn,
    calc_aum as _calc_aum_fn,
    calc_nav as _calc_nav_fn,
    long_positions_sorted as _long_positions_sorted_fn,
    aggregate_by_sub_asset as _aggregate_by_sub_asset_fn,
    aggregate_by_asset_type as _aggregate_by_asset_type_fn,
    geo_focus as _geo_focus_fn,
    GeoFocusResult,
)
from derivation.aifmd_turnover import (
    derive_turnovers_from_transactions as _derive_turnovers_fn,
)
from derivation.aifmd_investor import (
    derive_investor_pcts_from_amounts as _derive_investor_pcts_fn,
)

# ── AIFMD XML packaging mixins ──────────────────────────────────────────────
from aifmd_packaging.aifm_builder import AifmBuilderMixin
from aifmd_packaging.aif_builder import AifBuilderMixin
from aifmd_packaging.orchestrator import OrchestratorMixin

# ── Canonical model imports ──────────────────────────────────────────────────
from canonical.model import CanonicalAIFMReport, CanonicalAIFReport
from canonical.provenance import SourcePriority
from canonical.aifmd_source_entities import (
    SourceCanonical, ManagerStatic, FundStatic, FundDynamic,
    Position, Transaction, Instrument, ShareClass, Counterparty,
    Strategy, Investor, RiskMeasure, MonthlyData, BorrowingSource,
    ControlledCompany, ControlledStructure,
)
from canonical.aifmd_projection import project_aifm, project_aif


# ── Data extraction ──────────────────────────────────────────────────────────

class MAdapter(AifmBuilderMixin, AifBuilderMixin, OrchestratorMixin):
    """Reads a M adapter AIFMD Annex IV template and produces XML."""

    def __init__(self, *paths: str, csv_paths: Optional[list[str]] = None):
        """Create a MAdapter from one or more input files.

        Accepts any mix of Excel (.xlsx/.xls) and CSV (.csv) files.
        Examples:
            MAdapter("template.xlsx")
            MAdapter("template.xlsx", "positions.csv")
            MAdapter("all_data.csv")
            MAdapter("aifm.csv", "aif.csv", "positions.csv")
            MAdapter("template.xlsx", csv_paths=["pos.csv"])  # legacy
        """
        # Merge legacy csv_paths kwarg into paths
        all_paths = list(paths) + (csv_paths or [])
        if not all_paths:
            raise ValueError("At least one input file path is required")

        self.excel_paths: list[Path] = []
        self.csv_paths: list[Path] = []
        for p in all_paths:
            pp = Path(p)
            if pp.suffix.lower() in (".xlsx", ".xls"):
                self.excel_paths.append(pp)
            elif pp.suffix.lower() == ".csv":
                self.csv_paths.append(pp)
            else:
                raise ValueError(f"Unsupported file type: {pp.suffix} (expected .xlsx, .xls, or .csv)")

        self.path = self.excel_paths[0] if self.excel_paths else self.csv_paths[0]

        # ── Load Excel workbooks ────────────────────────────────────────
        # Rules:
        #   (1) Tabs starting with # are always ignored
        #   (2) Tab names can be anything — classification is by content only
        #   (3) Excel/CSV file names can be anything
        #   (4) Any combination of 1+ tabs, 1+ excels, 1+ csvs
        self._workbooks: list = []     # keep refs alive for worksheet access
        self.wb = None                 # primary workbook (backward compat)
        self.ws = None                 # primary data sheet (contains AIFM record)
        self.positions_ws = None
        self.extra_data_sheets = []    # additional sheets to parse

        for excel_path in self.excel_paths:
            wb = openpyxl.load_workbook(str(excel_path), data_only=True)
            self._workbooks.append(wb)

            for name in wb.sheetnames:
                # Rule 1: always skip tabs that start with #
                if name.startswith("#"):
                    continue

                ws_candidate = wb[name]
                sheet_role = self._classify_sheet(ws_candidate)

                if sheet_role == "primary":
                    if self.ws is None:
                        self.wb = wb
                        self.ws = ws_candidate
                    else:
                        # Already have a primary — add as extra data
                        self.extra_data_sheets.append(ws_candidate)
                elif sheet_role == "data":
                    self.extra_data_sheets.append(ws_candidate)
                # sheet_role == "unknown" → skip

        self._parse()

    # ── Sheet classification ───────────────────────────────────────────

    def _classify_sheet(self, ws) -> str:
        """Classify a worksheet by scanning its content (not its name).

        Returns:
            "primary" — contains AIFM / AIFM_LIGHT record (the main data sheet)
            "data"    — contains known M adapter section tags or record types
            "unknown" — no recognisable M adapter content
        """
        has_data_content = False
        scan_limit = min(30, ws.max_row or 0)

        for r in range(1, scan_limit + 1):
            cell_val = _str(ws.cell(r, 1).value or "").strip()
            if not cell_val:
                continue

            if cell_val.startswith("#"):
                # Check for section header (skip decorative lines)
                stripped = cell_val.lstrip("#").strip()
                if not stripped:
                    continue
                tag = stripped.split()[0].upper()
                # Skip non-data markers
                if any(x in cell_val.lower()
                       for x in ("fieldnum", "copyright", "start", "ignore",
                                 "end of", "doc ", "validat", "overview")):
                    continue
                dispatch_key = self._dispatch_tag(tag)
                if dispatch_key in self._KNOWN_SECTION_TAGS:
                    has_data_content = True
            else:
                # Data row — check for record types
                tag_upper = cell_val.strip().upper()
                if tag_upper in ("AIFM", "AIFM_LIGHT"):
                    return "primary"
                dispatch_key = self._dispatch_tag(cell_val)
                if dispatch_key in self._RECORD_DISPATCH:
                    has_data_content = True

        return "data" if has_data_content else "unknown"

    # ── Parsing ──────────────────────────────────────────────────────────

    def _detect_template_type(self) -> str:
        """Detect if template is LIGHT or FULL by scanning all sources."""

        def _scan_excel_ws(ws) -> str | None:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
                a_val = row[0].value
                if a_val:
                    a_str = _str(a_val)
                    if a_str == "AIFM" and not a_str.endswith("_LIGHT"):
                        return "FULL"
                    elif a_str == "AIFM_LIGHT":
                        return "LIGHT"
            return None

        # Check primary Excel sheet + all extra data sheets
        for ws in [self.ws] + self.extra_data_sheets:
            if ws is not None:
                result = _scan_excel_ws(ws)
                if result:
                    return result
        # Check CSV files
        for csv_path in self.csv_paths:
            try:
                with open(csv_path, newline="", encoding="utf-8-sig") as fh:
                    sample = fh.read(4096)
                    fh.seek(0)
                    delim = ";" if sample.count(";") >= sample.count(",") else ","
                    for line_raw in fh:
                        tag = line_raw.split(delim)[0].strip()
                        if tag == "AIFM":
                            return "FULL"
                        elif tag == "AIFM_LIGHT":
                            return "LIGHT"
            except Exception:
                pass
        return "LIGHT"

    def _parse(self):
        """Extract all record types from all input files."""
        self.template_type = self._detect_template_type()

        self.aifm = {}
        self.aifs = []
        self.positions = []
        self.aifm_assumptions = []
        self.aif_assumptions = []
        self.share_classes = []

        # Full template only
        self.aifm_national_codes = []
        self.aifm_national_data = []
        self.aif_national_codes = []
        self.aif_national_data = []
        self.historical_risk_profiles = []
        self.finance_records = []
        self.strategies = []
        self.turnovers = []
        self.transactions = []
        self.risks = []
        self.counterparty_risks = []
        self.investor_liquidity_profiles = []
        self.investor_redemptions = []
        self.investor_groups = []
        self.investor_amounts = []
        self.stress_tests = []
        self.portfolio_liquidity_profiles = []
        self.portfolio_positions = []
        self.custom_fx_rates = []
        self.master_records = []
        self.special_arrangements = []
        self.dominant_influences = []
        self.prime_brokers = []
        self.controlled_structures = []
        self.borrow_sources = []
        self.financing_liquidity = []
        self.instruments = []
        self.risk_defaults = []
        self._ecb_fx_rate_cache: dict[str, float] = {}  # per-instance cache

        # Parse Excel sources (if any)
        if self.ws:
            self._parse_sheet(self.ws)
            for extra_ws in self.extra_data_sheets:
                self._parse_sheet(extra_ws)
            if self.positions_ws and self.template_type == "FULL":
                self._parse_positions_sheet()

        # Parse CSV sources (primary or supplementary)
        for csv_path in self.csv_paths:
            self._parse_csv(csv_path)

        # Derive shared fields
        # For full template, reporting member state might be in AIFM_NATIONAL_CODE records
        self.reporting_member_state = _str(self.aifm.get("AIF(M) Reporting Member State", "") or
                                           self.aifm.get("Reporting Member State", ""))
        if not self.reporting_member_state and self.aifm_national_codes:
            # Use the first AIFM national code's reporting member state
            self.reporting_member_state = _str(self.aifm_national_codes[0].get("AIFM Reporting Member State", "") or
                                              self.aifm_national_codes[0].get("Reporting Member State", ""))

        self.reporting_year = int(self.aifm.get("Reporting Period Year", 0))
        self.reporting_period_type = _str(self.aifm.get("Reporting Period Type", "")) or "Y1"
        self.aifm_national_code = _str(self.aifm.get("AIFM National Code", ""))
        self.aifm_name = _str(self.aifm.get("AIFM Name", ""))
        self.aifm_jurisdiction = _str(
            self.aifm.get("AIFM Jurisdiction", "") or self.reporting_member_state
        )
        self.filing_type = _str(self.aifm.get("Filing Type", "")) or "INIT"
        self.aifm_lei = _str(self.aifm.get("AIFM LEI code", "") or self.aifm.get("AIFM LEI Code", ""))
        self.aifm_bic = _str(self.aifm.get("AIFM BIC code", "") or self.aifm.get("AIFM BIC Code", ""))

        # ── Multi-NCA: AIFM_NATIONAL_DATA extends AIFM_NATIONAL_CODE ────────
        # AIFM_NATIONAL_DATA adds per-NCA Content Type, Reporting Period Type,
        # and Reporting Code on top of the base national code registration.
        #
        # Three scenarios:
        #   1. Only AIFM_NATIONAL_CODE present → works as before (no change)
        #   2. Only AIFM_NATIONAL_DATA present → also serves as national codes
        #      (backward compatible: NATIONAL_DATA is a superset of NATIONAL_CODE)
        #   3. Both present → NATIONAL_CODE for registration, NATIONAL_DATA
        #      for per-NCA overrides (accessed via get_aifm_nca_data())
        if self.aifm_national_data and not self.aifm_national_codes:
            # Scenario 2: promote NATIONAL_DATA to also serve as national codes
            for nd in self.aifm_national_data:
                self.aifm_national_codes.append(nd)
            if not self.reporting_member_state:
                self.reporting_member_state = _str(
                    self.aifm_national_data[0].get("AIFM Reporting Member State", ""))

        # AIF_NATIONAL_DATA: same three scenarios
        if self.aif_national_data and not self.aif_national_codes:
            for nd in self.aif_national_data:
                self.aif_national_codes.append(nd)

        # ── INSTRUMENT enrichment: enrich POSITION records with static data ─
        # Build lookup by Instrument Name for fast enrichment
        self._instrument_lookup: dict[str, dict] = {}
        for instr in self.instruments:
            name = _str(instr.get("Instrument Name", ""))
            if name:
                self._instrument_lookup[name] = instr
        self._enrich_positions_from_instruments()

        # ── RISK_DEFAULT: build lookup for default risk descriptions ────────
        self._risk_default_lookup: dict[str, str] = {}
        for rd in self.risk_defaults:
            rtype = _str(rd.get("Risk measure type", ""))
            rdesc = _str(rd.get("Risk Measure description", ""))
            if rtype and rdesc:
                self._risk_default_lookup[rtype] = rdesc

    _KNOWN_SECTION_TAGS = frozenset({
        "AIFM_LIGHT", "AIF_LIGHT", "POSITION_COMPACT",
        "AIFM_ASSUMPTION", "AIF_ASSUMPTION", "SHARE_CLASS_COMPACT",
        "AIFM", "AIF", "AIFM_NATIONAL_CODE", "AIF_NATIONAL_CODE",
        "AIFM_NATIONAL_DATA", "AIF_NATIONAL_DATA",
        "HISTORICAL_RISK_PROFILE", "FINANCE", "SHARE_CLASS", "STRATEGY",
        "TURNOVER", "TRANSACTION", "RISK", "COUNTERPARTY_RISK_PROFILE",
        "INVESTOR_LIQUIDITY_PROFILE", "INVESTOR_REDEMPTION", "INVESTOR",
        "INVESTOR_AMOUNT",
        "STRESS_TEST", "PORTFOLIO_LIQUIDITY_PROFILE", "POSITION",
        "DOMINANT_INFLUENCE", "FINANCING_LIQUIDITY", "SPECIAL_ARRANGEMENT",
        "PRIME_BROKER", "CONTROLLED_STRUCTURE", "BORROW_SOURCE",
        "CUSTOM_FX_RATE",
        "MASTER",
        "INSTRUMENT", "RISK_DEFAULT",
    })

    # Maps record type string → (attribute name, append vs assign)
    _RECORD_DISPATCH = {
        "AIFM_LIGHT": ("aifm", "assign"),
        "AIFM": ("aifm", "assign"),
        "AIF_LIGHT": ("aifs", "append"),
        "AIF": ("aifs", "append"),
        "AIFM_NATIONAL_CODE": ("aifm_national_codes", "append"),
        "AIFM_NATIONAL_DATA": ("aifm_national_data", "append"),
        "AIF_NATIONAL_CODE": ("aif_national_codes", "append"),
        "AIF_NATIONAL_DATA": ("aif_national_data", "append"),
        "POSITION_COMPACT": ("positions", "append"),
        "AIFM_ASSUMPTION": ("aifm_assumptions", "append"),
        "AIF_ASSUMPTION": ("aif_assumptions", "append"),
        "SHARE_CLASS_COMPACT": ("share_classes", "append"),
        "SHARE_CLASS": ("share_classes", "append"),
        "HISTORICAL_RISK_PROFILE": ("historical_risk_profiles", "append"),
        "FINANCE": ("finance_records", "append"),
        "STRATEGY": ("strategies", "append"),
        "TURNOVER": ("turnovers", "append"),
        "TRANSACTION": ("transactions", "append"),
        "RISK": ("risks", "append"),
        "COUNTERPARTY_RISK_PROFILE": ("counterparty_risks", "append"),
        "INVESTOR_LIQUIDITY_PROFILE": ("investor_liquidity_profiles", "append"),
        "INVESTOR_REDEMPTION": ("investor_redemptions", "append"),
        "INVESTOR": ("investor_groups", "append"),
        "INVESTOR_AMOUNT": ("investor_amounts", "append"),
        "STRESS_TEST": ("stress_tests", "append"),
        "POSITION": ("portfolio_positions", "append"),
        "PORTFOLIO_LIQUIDITY_PROFILE": ("portfolio_liquidity_profiles", "append"),
        "CUSTOM_FX_RATE": ("custom_fx_rates", "append"),
        "MASTER": ("master_records", "append"),
        "SPECIAL_ARRANGEMENT": ("special_arrangements", "append"),
        "DOMINANT_INFLUENCE": ("dominant_influences", "append"),
        "PRIME_BROKER": ("prime_brokers", "append"),
        "CONTROLLED_STRUCTURE": ("controlled_structures", "append"),
        "BORROW_SOURCE": ("borrow_sources", "append"),
        "FINANCING_LIQUIDITY": ("financing_liquidity", "append"),
        "INSTRUMENT": ("instruments", "append"),
        "RISK_DEFAULT": ("risk_defaults", "append"),
    }

    @staticmethod
    def _normalize_section_tag(raw: str) -> str:
        """Extract and normalize a section tag from a header row.

        Returns the tag exactly as found (e.g. 'STRATEGYv2' stays 'STRATEGYv2',
        'STRATEGY' stays 'STRATEGY'), except that a leading '#' is stripped.
        """
        tag = raw.lstrip("#").strip().split()[0] if raw.lstrip("#").strip() else ""
        return tag

    @staticmethod
    def _dispatch_tag(section_tag: str) -> str:
        """Map a section tag to its dispatch key for _RECORD_DISPATCH.

        Variant tags map to their canonical dispatch key, e.g.:
        'STRATEGYv2' → 'STRATEGY', 'AIFM_LIGHT' → 'AIFM_LIGHT' (already in dispatch).
        """
        if section_tag.startswith("STRATEGY"):
            return "STRATEGY"
        return section_tag

    # ── INSTRUMENT enrichment ──────────────────────────────────────────────
    _INSTRUMENT_ENRICH_FIELDS = (
        "Sub-Asset Type of Position", "Market Type Code",
        "Market Identifier Code (MIC)", "Counterparty Type",
        "Counterparty Name", "Counterparty LEI Code",
        "Currency of the exposure", "Region or Country Code",
        "Instrument ISIN Code",
    )

    def _enrich_positions_from_instruments(self):
        """Enrich POSITION / POSITION_COMPACT records with INSTRUMENT static data.

        For each position, if the Instrument Name matches an INSTRUMENT record,
        any empty fields in the position are filled from the INSTRUMENT record.
        This allows users to define static instrument data once and leave those
        columns empty in the POSITION rows.
        """
        if not self._instrument_lookup:
            return
        for pos_list in (self.positions, self.portfolio_positions):
            for pos in pos_list:
                instr_name = _str(pos.get("Instrument Name", ""))
                if not instr_name or instr_name not in self._instrument_lookup:
                    continue
                instr = self._instrument_lookup[instr_name]
                for field in self._INSTRUMENT_ENRICH_FIELDS:
                    if not _str(pos.get(field, "")) and _str(instr.get(field, "")):
                        pos[field] = instr[field]

    # ── TRANSACTION → TURNOVER derivation ────────────────────────────────
    def _derive_turnovers_from_transactions(self, aif_id: str,
                                             period_start: str,
                                             period_end: str) -> list[dict]:
        """Delegate to derivation.aifmd_turnover (extracted Phase 3)."""
        return _derive_turnovers_fn(
            transactions=self.transactions,
            aif_id=aif_id,
            period_start=period_start,
            period_end=period_end,
            instrument_lookup=self._instrument_lookup,
        )

    # ── INVESTOR_AMOUNT → INVESTOR derivation ────────────────────────────
    def _derive_investor_pcts_from_amounts(self, aif_id: str) -> list[dict]:
        """Delegate to derivation.aifmd_investor (extracted Phase 3)."""
        return _derive_investor_pcts_fn(
            investor_amounts=self.investor_amounts,
            aif_id=aif_id,
        )

    # ── NCA-specific data accessors ──────────────────────────────────────
    def get_aifm_nca_data(self, reporting_member_state: str) -> dict:
        """Get AIFM registration data for a specific NCA.

        Returns the AIFM_NATIONAL_DATA record for the given NCA, or falls
        back to base AIFM fields if no per-NCA override exists. This
        enables multi-NCA reporting where Content Type, Period Type, and
        Reporting Code may differ per NCA.
        """
        for nd in self.aifm_national_data:
            rms = _str(nd.get("AIFM Reporting Member State", ""))
            if rms == reporting_member_state:
                return nd
        # Fallback: return base AIFM-level fields
        return {
            "AIFM Reporting Member State": self.reporting_member_state,
            "AIFM National Code": self.aifm_national_code,
            "AIFM Content Type": _str(self.aifm.get("AIFM Content Type", "")),
            "Reporting Period Type": self.reporting_period_type,
            "AIFM Reporting Code": _str(self.aifm.get("AIFM Reporting Code", "")),
        }

    def get_aif_nca_data(self, aif_id: str, reporting_member_state: str) -> dict:
        """Get AIF registration data for a specific NCA.

        Returns the AIF_NATIONAL_DATA record for the given AIF + NCA combo,
        or falls back to base AIF_NATIONAL_CODE if no per-NCA override exists.
        """
        for nd in self.aif_national_data:
            nd_aif = _str(nd.get("Custom AIF Identification", ""))
            rms = _str(nd.get("AIF Reporting Member State", ""))
            if nd_aif == aif_id and rms == reporting_member_state:
                return nd
        # Fallback: check AIF_NATIONAL_CODE records
        for nc in self.aif_national_codes:
            nc_aif = _str(nc.get("Custom AIF Identification", ""))
            rms = _str(nc.get("AIF Reporting Member State", ""))
            if nc_aif == aif_id and rms == reporting_member_state:
                return nc
        return {}

    def get_risk_default_description(self, risk_type: str) -> str | None:
        """Get the default risk measure description for a given risk type.

        Returns the RISK_DEFAULT description if one was defined, or None.
        Used to override the system-generated default description for
        field 147 (Risk Measure description).
        """
        return self._risk_default_lookup.get(risk_type)

    @staticmethod
    def _schema_tag(section_tag: str) -> str:
        """Map a parsed section tag to its key in the column schema.

        Each record type (including Light variants) has its own schema entry
        with the correct column layout. STRATEGY/STRATEGYv2 are kept distinct.
        No mapping is needed for most types — they match their schema key directly.
        """
        return section_tag

    def _make_record(self, section_tag: str, row, header_row=None) -> "MRecord":
        """Create a MRecord from a data row.

        Uses the column schema for canonical field names when available,
        falling back to header-based parsing otherwise.
        """
        schema_key = self._schema_tag(section_tag)
        schema_def = _COLUMN_SCHEMA.get(schema_key)
        schema_cols = schema_def["columns"] if schema_def else None

        return MRecord.from_row(
            record_type=section_tag,
            row=row,
            schema_cols=schema_cols,
            header_row=header_row,
        )

    def _parse_sheet(self, ws):
        """Parse a single worksheet, dispatching records to the correct lists.

        Column A determines row type:
          - '#...'      → section header (field names) or metadata (skip)
          - 'AIFM'      → data row for that record type
          - ''  / None   → skip (not a data row)
        """
        current_section = None
        header_row = None  # kept as fallback for schema-less parsing

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            a_val = row[0].value
            if a_val is None:
                continue

            a_str = _str(a_val)

            # Detect section headers (prefixed with #)
            if a_str.startswith("#"):
                tag = self._normalize_section_tag(a_str)
                dispatch_key = self._dispatch_tag(tag)
                if (dispatch_key in self._KNOWN_SECTION_TAGS
                    and not any(x in a_str.lower()
                                for x in ("fieldnum", "copyright", "start", "ignore"))):
                    current_section = tag  # preserve original tag for schema lookup
                    header_row = row
                continue

            # Data rows — only process if we have a current_section
            if not current_section:
                continue

            # Data row dispatch: column A value is the record type
            dispatch_key = self._dispatch_tag(a_str)
            dispatch = self._RECORD_DISPATCH.get(dispatch_key)
            if not dispatch:
                # Also try the raw value (for record types that match exactly)
                dispatch = self._RECORD_DISPATCH.get(a_str)
            if dispatch:
                attr_name, mode = dispatch
                record = self._make_record(current_section, row, header_row)
                if mode == "assign":
                    setattr(self, attr_name, record)
                else:
                    getattr(self, attr_name).append(record)

    def _parse_positions_sheet(self):
        """Parse the separate Positions sheet (full template)."""
        current_section = None
        header_row = None

        for row in self.positions_ws.iter_rows(min_row=1, max_row=self.positions_ws.max_row,
                                                values_only=False):
            a_val = row[0].value
            if a_val is None:
                continue

            a_str = _str(a_val)

            # Detect section headers
            if a_str.startswith("#"):
                tag = self._normalize_section_tag(a_str)
                if (tag in ("PORTFOLIO_LIQUIDITY_PROFILE", "POSITION")
                    and not any(x in a_str.lower() for x in ("fieldnum", "copyright", "start", "ignore"))):
                    current_section = tag
                    header_row = row
                continue

            # Data rows — use current_section for schema, a_str for dispatch
            if a_str in ("PORTFOLIO_LIQUIDITY_PROFILE", "POSITION"):
                record = self._make_record(current_section or a_str, row, header_row)
                dispatch = self._RECORD_DISPATCH.get(a_str)
                if dispatch:
                    attr_name, mode = dispatch
                    getattr(self, attr_name).append(record)

    def _parse_csv(self, csv_path: Path):
        """Parse a M adapter CSV file and dispatch records.

        CSV format:
          - Delimiter is ';' (M adapter standard) or ',' (auto-detected)
          - Lines starting with '#' are comments / headers
          - First field = record type (e.g. 'POSITION', 'AIFM', 'INPUT_FORMAT')
          - INPUT_FORMAT row is metadata → skip
          - Column A record type determines dispatch, just like Excel parsing
        """
        with open(csv_path, newline="", encoding="utf-8-sig") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            # Auto-detect delimiter: ';' vs ','
            semi_count = sample.count(";")
            comma_count = sample.count(",")
            delimiter = ";" if semi_count >= comma_count else ","
            reader = csv.reader(fh, delimiter=delimiter)

            current_section = None
            header_row_values = None

            for row in reader:
                if not row:
                    continue

                a_str = row[0].strip()

                # Skip empty rows
                if not a_str:
                    continue

                # Comment / header rows
                if a_str.startswith("#"):
                    tag = self._normalize_section_tag(a_str)
                    dispatch_key = self._dispatch_tag(tag)
                    if (dispatch_key in self._KNOWN_SECTION_TAGS
                        and not any(x in a_str.lower()
                                    for x in ("fieldnum", "copyright", "start", "ignore"))):
                        current_section = tag
                        header_row_values = row
                    continue

                # Skip INPUT_FORMAT metadata
                if a_str == "INPUT_FORMAT":
                    continue

                # Data rows — dispatch by record type in column A
                dispatch_key = self._dispatch_tag(a_str)
                dispatch = self._RECORD_DISPATCH.get(dispatch_key)
                if not dispatch:
                    dispatch = self._RECORD_DISPATCH.get(a_str)
                if dispatch:
                    section = current_section or a_str
                    # Build a MRecord from CSV row values
                    schema_key = self._schema_tag(section)
                    schema_def = _COLUMN_SCHEMA.get(schema_key)
                    schema_cols = schema_def["columns"] if schema_def else None
                    record = MRecord.from_csv_row(
                        record_type=section,
                        row_values=row,
                        schema_cols=schema_cols,
                    )
                    attr_name, mode = dispatch
                    if mode == "assign":
                        setattr(self, attr_name, record)
                    else:
                        getattr(self, attr_name).append(record)

    @staticmethod
    def _row_to_dict(header_row, data_row) -> dict:
        """Legacy helper — kept for backward compatibility."""
        result = {}
        if not header_row:
            return result
        for hcell, dcell in zip(header_row, data_row):
            h = hcell.value
            if h and not _str(h).startswith("#"):
                result[_str(h)] = dcell.value
        return result

    # ── Position helpers ─────────────────────────────────────────────────

    def _positions_for_aif(self, aif_id: str) -> list[dict]:
        """Get all positions for an AIF (light or full template)."""
        # Light template
        if self.positions:
            return [p for p in self.positions
                   if _str(p.get("Custom AIF Identification", "") or p.get("AIF ID", "")) == aif_id]
        # Full template — try both field name variants
        return [p for p in self.portfolio_positions
               if _str(p.get("Custom AIF Identification", "") or
                        p.get("AIF ID", "")) == aif_id]

    @staticmethod
    def _pos_val(p: dict) -> float:
        """Delegate to derivation.aifmd_portfolio.pos_val (extracted Phase 3)."""
        return _pos_val_fn(p)

    def _calc_aum(self, positions: list[dict]) -> int:
        """Delegate to derivation.aifmd_portfolio.calc_aum (extracted Phase 3)."""
        return _calc_aum_fn(positions)

    def _calc_nav(self, positions: list[dict]) -> int:
        """Delegate to derivation.aifmd_portfolio.calc_nav (extracted Phase 3)."""
        return _calc_nav_fn(positions)

    def _long_positions_sorted(self, positions: list[dict]) -> list[dict]:
        """Delegate to derivation.aifmd_portfolio.long_positions_sorted (extracted Phase 3)."""
        return _long_positions_sorted_fn(positions)

    def _aggregate_by_sub_asset(self, positions: list[dict],
                                reporting_member_state: str = "") -> list[dict]:
        """Delegate to derivation.aifmd_portfolio.aggregate_by_sub_asset (extracted Phase 3)."""
        return _aggregate_by_sub_asset_fn(positions, reporting_member_state)

    def _aggregate_by_asset_type(self, positions: list[dict],
                                reporting_member_state: str = "") -> list[dict]:
        """Delegate to derivation.aifmd_portfolio.aggregate_by_asset_type (extracted Phase 3)."""
        return _aggregate_by_asset_type_fn(positions, reporting_member_state)

    def _geo_focus(self, positions: list[dict], aum: int) -> dict:
        """Delegate to derivation.aifmd_portfolio.geo_focus (extracted Phase 3).

        Maps GeoFocusResult back to the state attributes expected by the
        XML builder mixins.
        """
        result = _geo_focus_fn(positions, aum)
        # Store raw values on self for use in XML generation (backward compat)
        self._last_geo_nav = result.nav_raw
        self._last_geo_aum_raw = result.aum_raw
        self._last_uk_nav = result.uk_nav
        self._last_uk_aum = result.uk_aum
        return result.aum_pct

    def _aif_ids_for_region(self, reporting_member_state: str = "") -> set | None:
        """Return AIF IDs whose positions should be included in an AIFM report.

        For multi-NCA (full) templates:
          - FCA (GB): only AIFs registered with the FCA.
          - ESMA (non-GB): AIFs registered with ANY ESMA NCA, de-duplicated by
            AIF ID so a fund registered in NL *and* BE is counted once.

        For single-NCA / light templates: returns None (include all AIFs).
        """
        if self.template_type != "FULL" or not self.aif_national_codes:
            return None  # single-NCA — include everything

        is_fca = (reporting_member_state == "GB")
        aif_ids: set[str] = set()
        for nc_rec in self.aif_national_codes:
            nc_rms = _str(nc_rec.get("AIF Reporting Member State", "") or
                          nc_rec.get("Reporting Member State", ""))
            nc_aif = _str(nc_rec.get("Custom AIF Identification", "") or
                          nc_rec.get("AIF ID", ""))
            if is_fca:
                # FCA: only AIFs that report to GB
                if nc_rms == "GB" and nc_aif:
                    aif_ids.add(nc_aif)
            else:
                # ESMA: any non-GB NCA (de-duplicated by set)
                if nc_rms != "GB" and nc_aif:
                    aif_ids.add(nc_aif)
        return aif_ids if aif_ids else None

    def _total_aum_all_aifs(self, reporting_member_state: str = "") -> int:
        """Sum AUM across AIFs, optionally filtered by NCA region."""
        allowed = self._aif_ids_for_region(reporting_member_state)
        total = 0
        for aif in self.aifs:
            aif_id = _str(aif.get("Custom AIF Identification", "") or aif.get("AIF ID", ""))
            if allowed is not None and aif_id not in allowed:
                continue
            if self.template_type == "FULL":
                aum_val = aif.get("Total AuM amount of the AIF in base currency")
                if aum_val:
                    total += _int_round(aum_val)
                    continue
            positions = self._positions_for_aif(aif_id)
            total += self._calc_aum(positions)
        return total

    def _get_fx_rate(self, aif_id: str = "") -> float:
        """Get FX EUR rate.

        Lookup order:
          1. AIFM record field (FXEURRate)
          2. AIF record field (FX EUR Rate)
          3. CUSTOM_FX_RATE record in template
          4. ECB Statistical Data Warehouse (online lookup)
          5. Fallback: 1.0 (assumes EUR base currency)
        """
        # 1. Try to get from AIFM record
        rate_str = _str(self.aifm.get("FXEURRate", ""))
        if rate_str:
            try:
                return float(rate_str)
            except:
                pass

        # 2. Try to get from AIF record if aif_id is specified
        if aif_id:
            for aif in self.aifs:
                if _str(aif.get("Custom AIF Identification", "") or aif.get("AIF ID", "")) == aif_id:
                    rate_str = _str(aif.get("FX EUR Rate", ""))
                    if rate_str:
                        try:
                            return float(rate_str)
                        except:
                            pass

        # 3. Try CUSTOM_FX_RATE records
        for fx_rec in self.custom_fx_rates:
            rate_str = _str(fx_rec.get("FX EUR Rate", "") or
                           fx_rec.get("FXEURRate", ""))
            if rate_str:
                try:
                    return float(rate_str)
                except:
                    pass

        # 4. Fetch from ECB if base currency is not EUR
        base_ccy = self._get_base_currency(aif_id)
        if base_ccy and base_ccy != "EUR":
            if base_ccy in self._ecb_fx_rate_cache:
                return self._ecb_fx_rate_cache[base_ccy]
            _, period_end = _reporting_period_dates(
                self.reporting_period_type, self.reporting_year)
            # 4a. Try live ECB API
            ecb_rate = _fetch_ecb_fx_rate(base_ccy, period_end)
            if ecb_rate is not None:
                self._ecb_fx_rate_cache[base_ccy] = ecb_rate
                return ecb_rate
            # 4b. Try hardcoded ECB fallback table
            fallback_rate = _get_ecb_fallback_rate(base_ccy, period_end)
            if fallback_rate is not None:
                self._ecb_fx_rate_cache[base_ccy] = fallback_rate
                return fallback_rate

        # 5. Fallback
        return 1.0

    def _get_base_currency(self, aif_id: str = "") -> str:
        """Get base currency from AIF record or default to EUR."""
        if aif_id:
            for aif in self.aifs:
                if _str(aif.get("Custom AIF Identification", "") or aif.get("AIF ID", "")) == aif_id:
                    ccy = _str(aif.get("Base currency of the AIF", ""))
                    if ccy:
                        return ccy
        # Default
        return "EUR"

    # ── Canonical model conversion methods ────────────────────────────────────

    def to_source_canonical(self) -> tuple["SourceCanonical", list["SourceCanonical"]]:
        """Export parsed M template data to Source Canonical entities.

        Returns:
            Tuple of (aifm_source, aif_sources) where:
            - aifm_source: SourceCanonical with ManagerStatic populated
            - aif_sources: List of SourceCanonical, one per AIF, with
              FundStatic, FundDynamic, and all collection entities populated
        """
        # Create AIFM-level SourceCanonical
        aifm_source = SourceCanonical()
        aifm_manager = ManagerStatic()

        # Populate manager from self.aifm dict
        if self.aifm:
            aifm_manager.set("name", _str(self.aifm_name), source="m_adapter",
                           priority=SourcePriority.IMPORTED)
            aifm_manager.set("jurisdiction", _str(self.aifm_jurisdiction), source="m_adapter",
                           priority=SourcePriority.IMPORTED)
            aifm_manager.set("national_code", _str(self.aifm_national_code), source="m_adapter",
                           priority=SourcePriority.IMPORTED)
            aifm_manager.set("lei", _str(self.aifm.get("AIFM LEI code")), source="m_adapter",
                           priority=SourcePriority.IMPORTED)

            # Derive EEA flag from jurisdiction
            eea_flag = _is_eea(self.aifm_jurisdiction)
            aifm_manager.set("eea_flag", eea_flag, source="m_adapter",
                           priority=SourcePriority.DERIVED)

            # BIC code
            bic = _str(self.aifm.get("AIFM BIC code"))
            if bic:
                aifm_manager.set("bic", bic, source="m_adapter",
                               priority=SourcePriority.IMPORTED)

            # NCA registration from first aifm_national_data entry
            if self.aifm_national_data:
                nca_entry = self.aifm_national_data[0]
                nca_content_type = _str(nca_entry.get("NCA content type", ""))
                if nca_content_type:
                    aifm_manager.set("nca_content_type", nca_content_type, source="m_adapter",
                                   priority=SourcePriority.IMPORTED)

                nca_period_type = _str(nca_entry.get("NCA period type", ""))
                if nca_period_type:
                    aifm_manager.set("nca_period_type", nca_period_type, source="m_adapter",
                                   priority=SourcePriority.IMPORTED)

                nca_reporting_code = _str(nca_entry.get("NCA reporting code", ""))
                if nca_reporting_code:
                    aifm_manager.set("nca_reporting_code", nca_reporting_code, source="m_adapter",
                                   priority=SourcePriority.IMPORTED)

                nca_sender_id = _str(nca_entry.get("NCA sender ID", ""))
                if nca_sender_id:
                    aifm_manager.set("nca_sender_id", nca_sender_id, source="m_adapter",
                                   priority=SourcePriority.IMPORTED)

        aifm_source.manager = aifm_manager

        # Create AIF-level SourceCanonical instances
        aif_sources = []
        for aif_idx, aif in enumerate(self.aifs):
            aif_source = SourceCanonical()

            # Get AIF ID for filtering related data
            aif_id = _str(aif.get("Custom AIF Identification", "") or aif.get("AIF ID", ""))

            # Populate FundStatic
            fund_static = FundStatic()

            fund_name = _str(aif.get("AIF Name", ""))
            if fund_name:
                fund_static.set("name", fund_name, source="m_adapter",
                              priority=SourcePriority.IMPORTED)

            # Domicile from multiple possible field names
            domicile = (_str(aif.get("Domicile of the AIF", "")) or
                       _str(aif.get("AIF Domicile", "")) or
                       self.aifm_jurisdiction)
            if domicile:
                fund_static.set("domicile", domicile, source="m_adapter",
                              priority=SourcePriority.IMPORTED)

            inception_date = (_str(aif.get("Inception date", "")) or
                             _str(aif.get("Inception Date", "")))
            if inception_date:
                fund_static.set("inception_date", inception_date, source="m_adapter",
                              priority=SourcePriority.IMPORTED)

            lei = (_str(aif.get("AIF LEI code", "")) or
                  _str(aif.get("AIF LEI Code", "")))
            if lei:
                fund_static.set("lei", lei, source="m_adapter",
                              priority=SourcePriority.IMPORTED)

            isin = _str(aif.get("AIF ISIN code", ""))
            if isin:
                fund_static.set("isin", isin, source="m_adapter",
                              priority=SourcePriority.IMPORTED)

            base_ccy = self._get_base_currency(aif_id)
            if base_ccy:
                fund_static.set("base_currency", base_ccy, source="m_adapter",
                              priority=SourcePriority.IMPORTED)

            predominant_type = _str(aif.get("Predominant AIF Type", ""))
            if predominant_type:
                fund_static.set("predominant_type", predominant_type, source="m_adapter",
                              priority=SourcePriority.IMPORTED)

            master_feeder_status = _str(aif.get("Master feeder status", ""))
            if master_feeder_status:
                fund_static.set("master_feeder_status", master_feeder_status, source="m_adapter",
                              priority=SourcePriority.IMPORTED)

            share_class_flag = _str(aif.get("AIF share class flag", ""))
            if share_class_flag:
                fund_static.set("share_class_flag", share_class_flag, source="m_adapter",
                              priority=SourcePriority.IMPORTED)

            # Derive EEA flag from domicile
            eea_flag = _is_eea(domicile)
            fund_static.set("eea_flag", eea_flag, source="m_adapter",
                          priority=SourcePriority.DERIVED)

            # NCA registration from aif_national_data matching this AIF
            for nca_entry in self.aif_national_data:
                entry_aif_id = _str(nca_entry.get("Custom AIF Identification", "") or
                                   nca_entry.get("AIF ID", ""))
                if entry_aif_id == aif_id:
                    nca_content_type = _str(nca_entry.get("NCA content type", ""))
                    if nca_content_type:
                        fund_static.set("nca_content_type", nca_content_type, source="m_adapter",
                                      priority=SourcePriority.IMPORTED)

                    nca_period_type = _str(nca_entry.get("NCA period type", ""))
                    if nca_period_type:
                        fund_static.set("nca_period_type", nca_period_type, source="m_adapter",
                                      priority=SourcePriority.IMPORTED)

                    nca_reporting_code = _str(nca_entry.get("NCA reporting code", ""))
                    if nca_reporting_code:
                        fund_static.set("nca_reporting_code", nca_reporting_code, source="m_adapter",
                                      priority=SourcePriority.IMPORTED)

                    nca_sender_id = _str(nca_entry.get("NCA sender ID", ""))
                    if nca_sender_id:
                        fund_static.set("nca_sender_id", nca_sender_id, source="m_adapter",
                                      priority=SourcePriority.IMPORTED)
                    break

            aif_source.fund_static = fund_static

            # Populate FundDynamic (for FULL templates)
            if self.template_type == "FULL":
                fund_dynamic = FundDynamic()
                # Map relevant fields from the AIF record to FundDynamic
                # These would include liquidity and trading-related fields
                aif_source.fund_dynamic = fund_dynamic

            # Populate collection entities
            aif_source.positions = self._populate_positions_for_aif(aif_id)
            aif_source.transactions = self._populate_transactions_for_aif(aif_id)
            aif_source.instruments = self._populate_instruments()
            aif_source.share_classes = self._populate_share_classes_for_aif(aif_id)
            aif_source.strategies = self._populate_strategies_for_aif(aif_id)
            aif_source.investors = self._populate_investors_for_aif(aif_id)
            aif_source.risk_measures = self._populate_risk_measures_for_aif(aif_id)
            aif_source.counterparties = self._populate_counterparties_for_aif(aif_id)
            aif_source.monthly_data = self._populate_monthly_data_for_aif(aif_id)
            aif_source.borrowing_sources = self._populate_borrowing_sources_for_aif(aif_id)
            aif_source.controlled_companies = self._populate_controlled_companies_for_aif(aif_id)
            aif_source.controlled_structures = self._populate_controlled_structures_for_aif(aif_id)

            aif_sources.append(aif_source)

        return aifm_source, aif_sources

    def _populate_positions_for_aif(self, aif_id: str) -> list["Position"]:
        """Map position dicts to Position entities for the given AIF."""
        positions = []
        for pos_dict in self._positions_for_aif(aif_id):
            pos = Position()
            # Map fields from pos_dict using set()
            for key, value in pos_dict.items():
                if value is not None:
                    pos.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
            positions.append(pos)
        return positions

    def _populate_transactions_for_aif(self, aif_id: str) -> list["Transaction"]:
        """Map transaction dicts to Transaction entities for the given AIF."""
        transactions = []
        for trans_dict in self.transactions:
            trans_aif_id = _str(trans_dict.get("Custom AIF Identification", "") or
                               trans_dict.get("AIF ID", ""))
            if trans_aif_id == aif_id:
                trans = Transaction()
                for key, value in trans_dict.items():
                    if value is not None:
                        trans.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
                transactions.append(trans)
        return transactions

    def _populate_instruments(self) -> list["Instrument"]:
        """Map instrument dicts to Instrument entities."""
        instruments = []
        for instr_dict in self.instruments:
            instr = Instrument()
            for key, value in instr_dict.items():
                if value is not None:
                    instr.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
            instruments.append(instr)
        return instruments

    def _populate_share_classes_for_aif(self, aif_id: str) -> list["ShareClass"]:
        """Map share class dicts to ShareClass entities for the given AIF."""
        share_classes = []
        for sc_dict in self.share_classes:
            sc_aif_id = _str(sc_dict.get("Custom AIF Identification", "") or
                            sc_dict.get("AIF ID", ""))
            if sc_aif_id == aif_id:
                sc = ShareClass()
                for key, value in sc_dict.items():
                    if value is not None:
                        sc.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
                share_classes.append(sc)
        return share_classes

    def _populate_strategies_for_aif(self, aif_id: str) -> list["Strategy"]:
        """Map strategy dicts to Strategy entities for the given AIF."""
        strategies = []
        for strat_dict in self.strategies:
            strat_aif_id = _str(strat_dict.get("Custom AIF Identification", "") or
                               strat_dict.get("AIF ID", ""))
            if strat_aif_id == aif_id:
                strat = Strategy()
                for key, value in strat_dict.items():
                    if value is not None:
                        strat.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
                strategies.append(strat)
        return strategies

    def _populate_investors_for_aif(self, aif_id: str) -> list["Investor"]:
        """Map investor dicts to Investor entities for the given AIF."""
        investors = []
        for inv_dict in self.investor_groups:
            inv_aif_id = _str(inv_dict.get("Custom AIF Identification", "") or
                             inv_dict.get("AIF ID", ""))
            if inv_aif_id == aif_id:
                inv = Investor()
                for key, value in inv_dict.items():
                    if value is not None:
                        inv.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)

                # If investor_amounts available, derive percentages
                if self.investor_amounts:
                    inv_pcts = self._derive_investor_pcts_from_amounts(aif_id)
                    for inv_id_key, pct_value in inv_pcts.items():
                        if pct_value is not None:
                            inv.set(inv_id_key, pct_value, source="m_adapter",
                                  priority=SourcePriority.DERIVED)

                investors.append(inv)
        return investors

    def _populate_risk_measures_for_aif(self, aif_id: str) -> list["RiskMeasure"]:
        """Map risk dicts to RiskMeasure entities for the given AIF."""
        risk_measures = []
        for risk_dict in self.risks:
            risk_aif_id = _str(risk_dict.get("Custom AIF Identification", "") or
                              risk_dict.get("AIF ID", ""))
            if risk_aif_id == aif_id:
                rm = RiskMeasure()
                for key, value in risk_dict.items():
                    if value is not None:
                        rm.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
                risk_measures.append(rm)
        return risk_measures

    def _populate_counterparties_for_aif(self, aif_id: str) -> list["Counterparty"]:
        """Map counterparty risk dicts to Counterparty entities for the given AIF."""
        counterparties = []
        for cp_dict in self.counterparty_risks:
            cp_aif_id = _str(cp_dict.get("Custom AIF Identification", "") or
                            cp_dict.get("AIF ID", ""))
            if cp_aif_id == aif_id:
                cp = Counterparty()
                for key, value in cp_dict.items():
                    if value is not None:
                        cp.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
                counterparties.append(cp)
        return counterparties

    def _populate_monthly_data_for_aif(self, aif_id: str) -> list["MonthlyData"]:
        """Map historical risk profile dicts to MonthlyData entities for the given AIF."""
        monthly_data = []
        for monthly_dict in self.historical_risk_profiles:
            monthly_aif_id = _str(monthly_dict.get("Custom AIF Identification", "") or
                                 monthly_dict.get("AIF ID", ""))
            if monthly_aif_id == aif_id:
                md = MonthlyData()
                for key, value in monthly_dict.items():
                    if value is not None:
                        md.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
                monthly_data.append(md)
        return monthly_data

    def _populate_borrowing_sources_for_aif(self, aif_id: str) -> list["BorrowingSource"]:
        """Map borrowing source dicts to BorrowingSource entities for the given AIF."""
        borrowing_sources = []
        for bs_dict in self.borrow_sources:
            bs_aif_id = _str(bs_dict.get("Custom AIF Identification", "") or
                            bs_dict.get("AIF ID", ""))
            if bs_aif_id == aif_id:
                bs = BorrowingSource()
                for key, value in bs_dict.items():
                    if value is not None:
                        bs.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
                borrowing_sources.append(bs)
        return borrowing_sources

    def _populate_controlled_companies_for_aif(self, aif_id: str) -> list["ControlledCompany"]:
        """Map dominant influence dicts to ControlledCompany entities for the given AIF."""
        controlled_companies = []
        for dc_dict in self.dominant_influences:
            dc_aif_id = _str(dc_dict.get("Custom AIF Identification", "") or
                            dc_dict.get("AIF ID", ""))
            if dc_aif_id == aif_id:
                cc = ControlledCompany()
                for key, value in dc_dict.items():
                    if value is not None:
                        cc.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
                controlled_companies.append(cc)
        return controlled_companies

    def _populate_controlled_structures_for_aif(self, aif_id: str) -> list["ControlledStructure"]:
        """Map controlled structure dicts to ControlledStructure entities for the given AIF."""
        controlled_structures = []
        for cs_dict in self.controlled_structures:
            cs_aif_id = _str(cs_dict.get("Custom AIF Identification", "") or
                            cs_dict.get("AIF ID", ""))
            if cs_aif_id == aif_id:
                cs = ControlledStructure()
                for key, value in cs_dict.items():
                    if value is not None:
                        cs.set(key, value, source="m_adapter", priority=SourcePriority.IMPORTED)
                controlled_structures.append(cs)
        return controlled_structures

    def _collect_aifm_report_fields(self) -> dict:
        """Collect AIFM report-specific fields (non-entity fields).

        Returns:
            Dictionary of field_id -> value for fields that don't come from
            source entities but are needed in the report.
        """
        report_fields = {}

        # Q1: Reporting Member State
        rms = self.reporting_member_state or ""
        if rms:
            report_fields["1"] = rms

        # Q2: Version (XSD version)
        report_fields["2"] = XSD_VERSION

        # Q3: Creation date
        report_fields["3"] = datetime.now().strftime("%Y-%m-%d")

        # Q4: Filing Type
        filing_type = self.filing_type or "INIT"
        report_fields["4"] = filing_type

        # Q5: Content Type (calculated based on template type and jurisdiction)
        if self.template_type != "FULL":
            content_type = "2"  # registered
        elif rms == self.aifm_jurisdiction:
            content_type = "1"  # authorised, home NCA
        else:
            content_type = "3"  # authorised, marketing NCA
        report_fields["5"] = content_type

        # Q6: Period Start Date
        period_start, period_end = _reporting_period_dates(self.reporting_period_type,
                                                           self.reporting_year)
        report_fields["6"] = period_start

        # Q7: Period End Date
        report_fields["7"] = period_end

        # Q8: Period Type
        period_type = self.reporting_period_type or "Y1"
        report_fields["8"] = period_type

        # Q9: Reporting Year (needed by from_canonical to reconstruct)
        report_fields["9"] = str(self.reporting_year) if self.reporting_year else ""

        return report_fields

    def _collect_aif_report_fields(self, aif: dict, aif_idx: int) -> dict:
        """Collect AIF report-specific fields for a given AIF.

        Args:
            aif: AIF record dict
            aif_idx: Index of AIF in self.aifs list

        Returns:
            Dictionary of field_id -> value for fields that don't come from
            source entities but are needed in the AIF report.
        """
        report_fields = {}

        # Filing and period information same as AIFM level
        rms = self.reporting_member_state or ""
        if rms:
            report_fields["1"] = rms

        report_fields["2"] = XSD_VERSION
        report_fields["3"] = datetime.now().strftime("%Y-%m-%d")

        filing_type = self.filing_type or "INIT"
        report_fields["4"] = filing_type

        if self.template_type != "FULL":
            content_type = "2"
        elif rms == self.aifm_jurisdiction:
            content_type = "1"
        else:
            content_type = "3"
        report_fields["5"] = content_type

        period_start, period_end = _reporting_period_dates(self.reporting_period_type,
                                                           self.reporting_year)
        report_fields["6"] = period_start
        report_fields["7"] = period_end

        period_type = self.reporting_period_type or "Y1"
        report_fields["8"] = period_type

        return report_fields

    def to_canonical_from_source(
        self,
        reference_store: "ReferenceStore | None" = None,
    ) -> tuple["CanonicalAIFMReport", list["CanonicalAIFReport"], "EnrichmentLog | None"]:
        """Two-step conversion: M template → Source Canonical → Report Canonical.

        This is the preferred path — preserves full lineage from source data
        through to report fields.

        When a ReferenceStore is provided, LEI enrichment runs at L1B
        (post-parsing, pre-projection). Empty LEI fields are matched against
        the GLEIF cache by normalized entity name. Single matches are
        auto-enriched with DERIVED priority; multiple matches are flagged
        as PENDING_USER_CHOICE for the review UI.

        Args:
            reference_store: Optional ReferenceStore with GLEIF cache.
                If None, LEI enrichment is skipped.

        Returns:
            Tuple of (aifm_report, aif_reports, enrichment_log) where:
            - aifm_report: CanonicalAIFMReport with all fields populated
            - aif_reports: List of CanonicalAIFReport with all fields populated
            - enrichment_log: EnrichmentLog if enrichment ran, else None
        """
        aifm_source, aif_sources = self.to_source_canonical()

        # ── L1B: LEI enrichment (optional) ────────────────────────
        enrichment_log = None
        if reference_store is not None:
            from shared.lei_enrichment import enrich_lei_fields, EnrichmentLog
            import logging
            _log = logging.getLogger("lei_enrichment")

            # Enrich AIFM-level canonical (manager LEI)
            aifm_enrich = enrich_lei_fields(aifm_source, reference_store)

            # Enrich each AIF-level canonical (fund LEI, counterparties, etc.)
            aif_enrichments = []
            for aif_source in aif_sources:
                aif_enrich = enrich_lei_fields(aif_source, reference_store)
                aif_enrichments.append(aif_enrich)

            # Merge into one enrichment log for the entire filing
            enrichment_log = EnrichmentLog()
            enrichment_log.actions.extend(aifm_enrich.actions)
            for ae in aif_enrichments:
                enrichment_log.actions.extend(ae.actions)
            enrichment_log.completed_at = aifm_enrich.completed_at

            _log.info(enrichment_log.summary())

        # ── Project to Report Canonical ───────────────────────────
        aifm_report_fields = self._collect_aifm_report_fields()
        aifm_report = project_aifm(aifm_source, report_fields=aifm_report_fields)

        # Project each AIF
        aif_reports = []
        for i, (aif, aif_source) in enumerate(zip(self.aifs, aif_sources)):
            aif_report_fields = self._collect_aif_report_fields(aif, i)
            aif_report = project_aif(aif_source, report_fields=aif_report_fields)
            aif_reports.append(aif_report)

        return aifm_report, aif_reports, enrichment_log

    def to_canonical_aifm(self) -> "CanonicalAIFMReport":
        """Export parsed AIFM data to a generic CanonicalAIFMReport.

        Maps all AIFM data from the parsed M template into the canonical model
        using ESMA question numbers as field IDs.
        """
        report = CanonicalAIFMReport()

        # AIFM scalar fields — map from self.aifm dict and derived fields
        # Q1: Reporting Member State
        rms = self.reporting_member_state or ""
        if rms:
            report.set_field("1", rms, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q2: Version (XSD version)
        report.set_field("2", XSD_VERSION, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q3: Creation date
        report.set_field("3", datetime.now().strftime("%Y-%m-%d"), source="m_adapter",
                        priority=SourcePriority.IMPORTED)

        # Q4: Filing Type
        filing_type = self.filing_type or "INIT"
        report.set_field("4", filing_type, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q5: Content Type (calculated based on template type and jurisdiction)
        if self.template_type != "FULL":
            content_type = "2"  # registered
        elif rms == self.aifm_jurisdiction:
            content_type = "1"  # authorised, home NCA
        else:
            content_type = "3"  # authorised, marketing NCA
        report.set_field("5", content_type, source="m_adapter", priority=SourcePriority.DERIVED)

        # Q6: Period Start Date
        period_start, period_end = _reporting_period_dates(self.reporting_period_type,
                                                           self.reporting_year)
        report.set_field("6", period_start, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q7: Period End Date
        report.set_field("7", period_end, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q8: Period Type
        period_type = self.reporting_period_type or "Y1"
        report.set_field("8", period_type, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q9: Year
        year = str(self.reporting_year) if self.reporting_year else ""
        if year:
            report.set_field("9", year, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q10-Q12: Change codes (from aifm)
        change_code_1 = _str(self.aifm.get("Change code 1", "") or self.aifm.get("Change Code 1", ""))
        if change_code_1:
            report.set_field("10", change_code_1, source="m_adapter", priority=SourcePriority.IMPORTED)

        change_code_2 = _str(self.aifm.get("Change code 2", "") or self.aifm.get("Change Code 2", ""))
        if change_code_2:
            report.set_field("11", change_code_2, source="m_adapter", priority=SourcePriority.IMPORTED)

        change_code_3 = _str(self.aifm.get("Change code 3", "") or self.aifm.get("Change Code 3", ""))
        if change_code_3:
            report.set_field("12", change_code_3, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q13: Last Reporting Flag
        last_reporting = _str(self.aifm.get("Last Reporting Flag", ""))
        if last_reporting:
            report.set_field("13", last_reporting, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q14-Q15: Assumptions (repeating group)
        if self.aifm_assumptions:
            for i, a_rec in enumerate(self.aifm_assumptions):
                q_num = _str(a_rec.get("Question Number", ""))
                a_desc = _str(a_rec.get("Assumption Description", ""))
                if q_num or a_desc:
                    report.add_group_item("assumptions", {
                        "question": q_num,
                        "description": a_desc,
                    }, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Store AIFM national codes in a custom group for roundtrip support
        if self.aifm_national_codes:
            for nc in self.aifm_national_codes:
                nc_dict = {}
                for key, val in nc.items():
                    if val is not None and _str(val):
                        nc_dict[key] = _str(val)
                if nc_dict:
                    report.add_group_item("aifm_national_codes", nc_dict, source="m_adapter",
                                        priority=SourcePriority.IMPORTED)

        # Store AIFM national data (per-NCA overrides) for multi-NCA support
        if self.aifm_national_data:
            for nd in self.aifm_national_data:
                nd_dict = {}
                for key, val in nd.items():
                    if val is not None and _str(val):
                        nd_dict[key] = _str(val)
                if nd_dict:
                    report.add_group_item("aifm_national_data", nd_dict, source="m_adapter",
                                        priority=SourcePriority.IMPORTED)

        # Q16: Reporting Code
        reporting_code = _str(self.aifm.get("AIFM Reporting Code", "")) or "1"
        report.set_field("16", reporting_code, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q17: Jurisdiction
        jurisdiction = self.aifm_jurisdiction or ""
        if jurisdiction:
            report.set_field("17", jurisdiction, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q18: National Code
        national_code = self.aifm_national_code or ""
        if national_code:
            report.set_field("18", national_code, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q19: AIFM Name
        aifm_name = self.aifm_name or ""
        if aifm_name:
            report.set_field("19", aifm_name, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q20: EEA Flag
        eea_flag = str(_is_eea(self.aifm_jurisdiction)).lower()
        report.set_field("20", eea_flag, source="m_adapter", priority=SourcePriority.DERIVED)

        # Q21: No Reporting Flag
        no_reporting = _str(self.aifm.get("AIFM no reporting flag (Nothing to report)", ""))
        if no_reporting:
            report.set_field("21", no_reporting, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q22: LEI code
        if self.aifm_lei:
            report.set_field("22", self.aifm_lei, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q23: BIC code
        bic = _str(self.aifm.get("AIFM BIC code", "") or self.aifm.get("AIFM BIC Code", ""))
        if bic:
            report.set_field("23", bic, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q24-Q25: Old identifiers
        old_id_1 = _str(self.aifm.get("AIFM old identifier 1", "") or self.aifm.get("AIFM Old Identifier 1", ""))
        if old_id_1:
            report.set_field("24", old_id_1, source="m_adapter", priority=SourcePriority.IMPORTED)

        old_id_2 = _str(self.aifm.get("AIFM old identifier 2", "") or self.aifm.get("AIFM Old Identifier 2", ""))
        if old_id_2:
            report.set_field("25", old_id_2, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q26-Q30: Principal Markets (up to 5)
        # Aggregate positions by sub-asset
        all_aif_ids = set()
        for aif in self.aifs:
            aif_id = _str(aif.get("Custom AIF Identification", "") or aif.get("AIF ID", ""))
            if aif_id:
                all_aif_ids.add(aif_id)

        all_positions = []
        for aif_id in all_aif_ids:
            all_positions.extend(self._positions_for_aif(aif_id))

        # Count positions by country/market
        market_counts = {}
        for pos in all_positions:
            market = _str(pos.get("Market of Listing", "") or pos.get("Listing Market", ""))
            if market:
                market_counts[market] = market_counts.get(market, 0) + 1

        # Sort by count descending and take top 5
        sorted_markets = sorted(market_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        for rank, (market, count) in enumerate(sorted_markets, 1):
            field_id = str(25 + rank)  # Q26 = field 26, etc.
            report.set_field(field_id, market, source="m_adapter",
                           priority=SourcePriority.DERIVED, note=f"Market rank {rank}")

        # Q31-Q35: Principal Instruments (up to 5)
        # Count positions by asset type
        asset_counts = {}
        for pos in all_positions:
            asset_type = _str(pos.get("Asset type", "") or pos.get("Asset Type", ""))
            if asset_type:
                asset_counts[asset_type] = asset_counts.get(asset_type, 0) + 1

        # Sort by count descending and take top 5
        sorted_assets = sorted(asset_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        for rank, (asset, count) in enumerate(sorted_assets, 1):
            field_id = str(30 + rank)  # Q31 = field 31, etc.
            report.set_field(field_id, asset, source="m_adapter",
                           priority=SourcePriority.DERIVED, note=f"Asset rank {rank}")

        # Q36: Total AuM in EUR
        total_aum = self._total_aum_all_aifs(rms)
        if total_aum:
            report.set_field("36", str(total_aum), source="m_adapter", priority=SourcePriority.DERIVED)

        # Q37: AuM base currency
        base_ccy = self._get_base_currency()
        if base_ccy:
            report.set_field("37", base_ccy, source="m_adapter", priority=SourcePriority.IMPORTED)

        # Q38: FX Rate Type
        fx_rate_type = _str(self.aifm.get("FX Rate Type", ""))
        if fx_rate_type:
            report.set_field("38", fx_rate_type, source="m_adapter", priority=SourcePriority.IMPORTED)

        return report

    def to_canonical_aifs(self) -> list["CanonicalAIFReport"]:
        """Export parsed AIF data to generic CanonicalAIFReport instances.

        Creates one canonical report per AIF in self.aifs, mapping all available
        fields from the M template using ESMA question numbers.
        """
        canonical_aifs = []

        for aif_index, aif in enumerate(self.aifs):
            report = CanonicalAIFReport()
            aif_id = _str(aif.get("Custom AIF Identification", "") or aif.get("AIF ID", ""))

            # Map basic AIF identification fields
            # Q1: AIF Identification (custom or AIF ID)
            if aif_id:
                report.set_field("1", aif_id, source="m_adapter", priority=SourcePriority.IMPORTED,
                               source_ref=f"AIF #{aif_index}")

            # Q2: AIF Name
            aif_name = _str(aif.get("AIF Name", ""))
            if aif_name:
                report.set_field("2", aif_name, source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q3: Domicile
            domicile = (_str(aif.get("Domicile of the AIF", "")) or
                       _str(aif.get("AIF Domicile", "")) or
                       _str(aif.get("Domicile", "")) or
                       self.aifm_jurisdiction or
                       self.reporting_member_state)
            if domicile:
                report.set_field("3", domicile, source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q4: Content Type
            content_type = _str(aif.get("AIF Content Type", "") or aif.get("AIF content type", ""))
            if content_type:
                report.set_field("4", content_type, source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q5: LEI
            aif_lei = _str(aif.get("AIF LEI code", "") or aif.get("AIF LEI Code", ""))
            if aif_lei:
                report.set_field("5", aif_lei, source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q6: Investment Strategy Code
            strategy_code = _str(aif.get("Investment strategy code", ""))
            if strategy_code:
                report.set_field("6", strategy_code, source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q7: AIF Type
            aif_type = _str(aif.get("AIF Type", ""))
            if aif_type:
                report.set_field("7", aif_type, source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q8: Base Currency
            base_ccy = self._get_base_currency(aif_id)
            if base_ccy:
                report.set_field("8", base_ccy, source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q9: Total AuM
            if self.template_type == "FULL":
                aum_val = aif.get("Total AuM amount of the AIF in base currency")
                aum = _int_round(aum_val) if aum_val else self._calc_aum(self._positions_for_aif(aif_id))
            else:
                aum = self._calc_aum(self._positions_for_aif(aif_id))
            if aum:
                report.set_field("9", str(aum), source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q10: NAV
            if self.template_type == "FULL":
                nav_val = aif.get("Total Net Asset Value of the AIF (NAV)")
                nav = _int_round(nav_val) if nav_val else self._calc_nav(self._positions_for_aif(aif_id))
            else:
                nav = self._calc_nav(self._positions_for_aif(aif_id))
            if nav:
                report.set_field("10", str(nav), source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q11: AIF Reporting Code
            aif_reporting_code = _str(aif.get("AIF Reporting Code", ""))
            if aif_reporting_code:
                report.set_field("11", aif_reporting_code, source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q12: Inception Date
            inception_date = _str(aif.get("Inception date", "") or aif.get("Inception Date", ""))
            if inception_date:
                report.set_field("12", inception_date, source="m_adapter", priority=SourcePriority.IMPORTED)

            # Q13: Last Reporting Flag
            aif_last_reporting = _str(aif.get("Last Reporting Flag", ""))
            if aif_last_reporting:
                report.set_field("13", aif_last_reporting, source="m_adapter", priority=SourcePriority.IMPORTED)

            # Store complete AIF dict as a special group for perfect roundtrip
            # We store the dict as a JSON string to preserve types
            import json
            aif_all_data = {}
            for key, val in aif.items():
                if val is not None:
                    aif_all_data[key] = val
            if aif_all_data:
                # Store as JSON to preserve types
                aif_json = json.dumps(aif_all_data, default=str)
                report.add_group_item("_aif_complete_data", {"_json": aif_json}, source="m_adapter",
                                    priority=SourcePriority.IMPORTED)

            # Store INSTRUMENT reference data (shared across AIFs, stored per report for completeness)
            if self.instruments:
                for instr in self.instruments:
                    instr_dict = {}
                    for key, val in instr.items():
                        if val is not None and _str(val):
                            instr_dict[key] = _str(val)
                    if instr_dict:
                        report.add_group_item("instruments", instr_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Map positions for this AIF
            positions = self._positions_for_aif(aif_id)
            if positions:
                for p_idx, pos in enumerate(positions):
                    pos_dict = {}

                    # Populate position fields from the position dict
                    # Use field IDs that would be appropriate for position items
                    # (this is a simplified mapping; full schema would map all position fields)
                    for key, val in pos.items():
                        if val is not None and _str(val):
                            pos_dict[key] = _str(val)

                    if pos_dict:
                        report.add_group_item("positions", pos_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Store AIF national codes for this AIF (used by from_canonical to reconstruct)
            aif_ncs = [nc for nc in self.aif_national_codes
                      if _str(nc.get("Custom AIF Identification", "") or
                             nc.get("AIF ID", "")) == aif_id]
            if aif_ncs:
                for nc in aif_ncs:
                    nc_dict = {}
                    for key, val in nc.items():
                        if val is not None and _str(val):
                            nc_dict[key] = _str(val)
                    if nc_dict:
                        report.add_group_item("aif_national_codes", nc_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Store AIF national data (per-NCA overrides) for multi-NCA support
            aif_nds = [nd for nd in self.aif_national_data
                      if _str(nd.get("Custom AIF Identification", "") or
                             nd.get("AIF ID", "")) == aif_id]
            if aif_nds:
                for nd in aif_nds:
                    nd_dict = {}
                    for key, val in nd.items():
                        if val is not None and _str(val):
                            nd_dict[key] = _str(val)
                    if nd_dict:
                        report.add_group_item("aif_national_data", nd_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Map share classes for this AIF
            aif_share_classes = [sc for sc in self.share_classes
                               if _str(sc.get("Custom AIF Identification", "") or
                                      sc.get("AIF ID", "")) == aif_id]
            if aif_share_classes:
                for sc in aif_share_classes:
                    sc_dict = {}
                    for key, val in sc.items():
                        if val is not None and _str(val):
                            sc_dict[key] = _str(val)
                    if sc_dict:
                        report.add_group_item("share_classes", sc_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Map strategies for this AIF
            aif_strategies = [st for st in self.strategies
                            if _str(st.get("Custom AIF Identification", "") or
                                   st.get("AIF ID", "")) == aif_id]
            if aif_strategies:
                for st in aif_strategies:
                    st_dict = {}
                    for key, val in st.items():
                        if val is not None and _str(val):
                            st_dict[key] = _str(val)
                    if st_dict:
                        report.add_group_item("strategies", st_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Map turnovers for this AIF
            # TRANSACTION records take precedence over TURNOVER records when present
            aif_turnovers = [t for t in self.turnovers
                           if _str(t.get("Custom AIF Identification", "") or
                                  t.get("AIF ID", "")) == aif_id]
            aif_has_transactions = any(
                _str(tx.get("Custom AIF Identification", "")) == aif_id
                for tx in self.transactions
            )
            if aif_has_transactions:
                # Derive turnovers from TRANSACTION records within reporting period
                period_start, period_end = _reporting_period_dates(
                    self.reporting_period_type, self.reporting_year)
                derived_turnovers = self._derive_turnovers_from_transactions(
                    aif_id, period_start, period_end)
                for t in derived_turnovers:
                    report.add_group_item("turnovers", t, source="m_adapter",
                                        priority=SourcePriority.DERIVED,
                                        note="Derived from TRANSACTION records")
                # Also store raw transactions for source traceability
                for tx in self.transactions:
                    if _str(tx.get("Custom AIF Identification", "")) == aif_id:
                        tx_dict = {}
                        for key, val in tx.items():
                            if val is not None and _str(val):
                                tx_dict[key] = _str(val) if not isinstance(val, (int, float)) else val
                        if tx_dict:
                            report.add_group_item("transactions", tx_dict, source="m_adapter",
                                                priority=SourcePriority.IMPORTED)
            elif aif_turnovers:
                for t in aif_turnovers:
                    t_dict = {}
                    for key, val in t.items():
                        if val is not None and _str(val):
                            t_dict[key] = _str(val)
                    if t_dict:
                        report.add_group_item("turnovers", t_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Map risks for this AIF
            # Apply RISK_DEFAULT descriptions where the RISK record has no description
            aif_risks = [r for r in self.risks
                       if _str(r.get("Custom AIF Identification", "") or
                              r.get("AIF ID", "")) == aif_id]
            if aif_risks:
                for r in aif_risks:
                    r_dict = {}
                    for key, val in r.items():
                        if val is not None and _str(val):
                            r_dict[key] = _str(val)
                    # Apply RISK_DEFAULT description if no description is set
                    risk_type = r_dict.get("Risk measure type", "")
                    if risk_type and not r_dict.get("Risk Measure description", ""):
                        default_desc = self.get_risk_default_description(risk_type)
                        if default_desc:
                            r_dict["Risk Measure description"] = default_desc
                    if r_dict:
                        report.add_group_item("risks", r_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Map counterparty risks for this AIF
            aif_cpr = [cp for cp in self.counterparty_risks
                      if _str(cp.get("Custom AIF Identification", "") or
                             cp.get("AIF ID", "")) == aif_id]
            if aif_cpr:
                for cp in aif_cpr:
                    cp_dict = {}
                    for key, val in cp.items():
                        if val is not None and _str(val):
                            cp_dict[key] = _str(val)
                    if cp_dict:
                        report.add_group_item("counterparty_risks", cp_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Map investor groups for this AIF
            # INVESTOR_AMOUNT records take precedence over INVESTOR records when present
            aif_investors = [inv for inv in self.investor_groups
                           if _str(inv.get("Custom AIF Identification", "") or
                                  inv.get("AIF ID", "")) == aif_id]
            aif_has_investor_amounts = any(
                _str(ia.get("Custom AIF Identification", "")) == aif_id
                for ia in self.investor_amounts
            )
            if aif_has_investor_amounts:
                # Derive investor percentages from INVESTOR_AMOUNT records
                derived_investors = self._derive_investor_pcts_from_amounts(aif_id)
                for inv in derived_investors:
                    report.add_group_item("investor_groups", inv, source="m_adapter",
                                        priority=SourcePriority.DERIVED,
                                        note="Derived from INVESTOR_AMOUNT records")
                # Also store raw investor amounts for source traceability
                for ia in self.investor_amounts:
                    if _str(ia.get("Custom AIF Identification", "")) == aif_id:
                        ia_dict = {}
                        for key, val in ia.items():
                            if val is not None and _str(val):
                                ia_dict[key] = _str(val) if not isinstance(val, (int, float)) else val
                        if ia_dict:
                            report.add_group_item("investor_amounts", ia_dict, source="m_adapter",
                                                priority=SourcePriority.IMPORTED)
            elif aif_investors:
                for inv in aif_investors:
                    inv_dict = {}
                    for key, val in inv.items():
                        if val is not None and _str(val):
                            inv_dict[key] = _str(val)
                    if inv_dict:
                        report.add_group_item("investor_groups", inv_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Map stress tests for this AIF
            aif_stress_tests = [st for st in self.stress_tests
                              if _str(st.get("Custom AIF Identification", "") or
                                     st.get("AIF ID", "")) == aif_id]
            if aif_stress_tests:
                for st in aif_stress_tests:
                    st_dict = {}
                    for key, val in st.items():
                        if val is not None and _str(val):
                            st_dict[key] = _str(val)
                    if st_dict:
                        report.add_group_item("stress_tests", st_dict, source="m_adapter",
                                            priority=SourcePriority.IMPORTED)

            # Map other full-template groups for this AIF
            if self.template_type == "FULL":
                # Historical risk profiles
                aif_hrp = [h for h in self.historical_risk_profiles
                         if _str(h.get("Custom AIF Identification", "") or
                                h.get("AIF ID", "")) == aif_id]
                if aif_hrp:
                    for h in aif_hrp:
                        h_dict = {}
                        for key, val in h.items():
                            if val is not None and _str(val):
                                h_dict[key] = _str(val)
                        if h_dict:
                            report.add_group_item("historical_risk_profiles", h_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

                # Portfolio liquidity profiles
                aif_plp = [pl for pl in self.portfolio_liquidity_profiles
                         if _str(pl.get("Custom AIF Identification", "") or
                                pl.get("AIF ID", "")) == aif_id]
                if aif_plp:
                    for pl in aif_plp:
                        pl_dict = {}
                        for key, val in pl.items():
                            if val is not None and _str(val):
                                pl_dict[key] = _str(val)
                        if pl_dict:
                            report.add_group_item("portfolio_liquidity_profiles", pl_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

                # Investor liquidity profiles
                aif_ilp = [il for il in self.investor_liquidity_profiles
                         if _str(il.get("Custom AIF Identification", "") or
                                il.get("AIF ID", "")) == aif_id]
                if aif_ilp:
                    for il in aif_ilp:
                        il_dict = {}
                        for key, val in il.items():
                            if val is not None and _str(val):
                                il_dict[key] = _str(val)
                        if il_dict:
                            report.add_group_item("investor_liquidity_profiles", il_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

                # Investor redemptions
                aif_ir = [ir for ir in self.investor_redemptions
                        if _str(ir.get("Custom AIF Identification", "") or
                               ir.get("AIF ID", "")) == aif_id]
                if aif_ir:
                    for ir in aif_ir:
                        ir_dict = {}
                        for key, val in ir.items():
                            if val is not None and _str(val):
                                ir_dict[key] = _str(val)
                        if ir_dict:
                            report.add_group_item("investor_redemptions", ir_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

                # Dominant influences
                aif_di = [di for di in self.dominant_influences
                        if _str(di.get("Custom AIF Identification", "") or
                               di.get("AIF ID", "")) == aif_id]
                if aif_di:
                    for di in aif_di:
                        di_dict = {}
                        for key, val in di.items():
                            if val is not None and _str(val):
                                di_dict[key] = _str(val)
                        if di_dict:
                            report.add_group_item("dominant_influences", di_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

                # Prime brokers
                aif_pb = [pb for pb in self.prime_brokers
                        if _str(pb.get("Custom AIF Identification", "") or
                               pb.get("AIF ID", "")) == aif_id]
                if aif_pb:
                    for pb in aif_pb:
                        pb_dict = {}
                        for key, val in pb.items():
                            if val is not None and _str(val):
                                pb_dict[key] = _str(val)
                        if pb_dict:
                            report.add_group_item("prime_brokers", pb_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

                # Controlled structures
                aif_cs = [cs for cs in self.controlled_structures
                        if _str(cs.get("Custom AIF Identification", "") or
                               cs.get("AIF ID", "")) == aif_id]
                if aif_cs:
                    for cs in aif_cs:
                        cs_dict = {}
                        for key, val in cs.items():
                            if val is not None and _str(val):
                                cs_dict[key] = _str(val)
                        if cs_dict:
                            report.add_group_item("controlled_structures", cs_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

                # Borrow sources
                aif_bs = [bs for bs in self.borrow_sources
                        if _str(bs.get("Custom AIF Identification", "") or
                               bs.get("AIF ID", "")) == aif_id]
                if aif_bs:
                    for bs in aif_bs:
                        bs_dict = {}
                        for key, val in bs.items():
                            if val is not None and _str(val):
                                bs_dict[key] = _str(val)
                        if bs_dict:
                            report.add_group_item("borrow_sources", bs_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

                # Financing liquidity
                aif_fl = [fl for fl in self.financing_liquidity
                        if _str(fl.get("Custom AIF Identification", "") or
                               fl.get("AIF ID", "")) == aif_id]
                if aif_fl:
                    for fl in aif_fl:
                        fl_dict = {}
                        for key, val in fl.items():
                            if val is not None and _str(val):
                                fl_dict[key] = _str(val)
                        if fl_dict:
                            report.add_group_item("financing_liquidity", fl_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

                # Special arrangements
                aif_sa = [sa for sa in self.special_arrangements
                        if _str(sa.get("Custom AIF Identification", "") or
                               sa.get("AIF ID", "")) == aif_id]
                if aif_sa:
                    for sa in aif_sa:
                        sa_dict = {}
                        for key, val in sa.items():
                            if val is not None and _str(val):
                                sa_dict[key] = _str(val)
                        if sa_dict:
                            report.add_group_item("special_arrangements", sa_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

                # Master records
                aif_mr = [mr for mr in self.master_records
                        if _str(mr.get("Custom AIF Identification", "") or
                               mr.get("AIF ID", "")) == aif_id]
                if aif_mr:
                    for mr in aif_mr:
                        mr_dict = {}
                        for key, val in mr.items():
                            if val is not None and _str(val):
                                mr_dict[key] = _str(val)
                        if mr_dict:
                            report.add_group_item("master_records", mr_dict,
                                                source="m_adapter", priority=SourcePriority.IMPORTED)

            canonical_aifs.append(report)

        return canonical_aifs

    @classmethod
    def from_canonical(cls, aifm_report: "CanonicalAIFMReport",
                      aif_reports: list["CanonicalAIFReport"],
                      path: str = "canonical") -> "MAdapter":
        """Reconstruct a MAdapter from canonical reports (for XML generation).

        This bypasses __init__ (no file parsing) and directly sets all the
        self.* attributes that the packaging mixins need. Used to generate XML
        from canonical reports without re-parsing the original file.

        Args:
            aifm_report: CanonicalAIFMReport with AIFM-level data
            aif_reports: List of CanonicalAIFReport with AIF-level data
            path: Fake path for logging (typically "canonical")

        Returns:
            MAdapter instance ready for XML generation
        """
        adapter = object.__new__(cls)

        # Initialize basic attributes
        adapter.path = Path(path)
        adapter.excel_paths = []
        adapter.csv_paths = []
        adapter._workbooks = []
        adapter.wb = None
        adapter.ws = None
        adapter.positions_ws = None
        adapter.extra_data_sheets = []

        # Determine template type (FULL if any aif_report has full-only groups)
        adapter.template_type = "FULL" if aif_reports else "LIGHT"

        # Initialize all data structures
        adapter.aifm = {}
        adapter.aifs = []
        adapter.positions = []
        adapter.aifm_assumptions = []
        adapter.aif_assumptions = []
        adapter.share_classes = []
        adapter.aifm_national_codes = []
        adapter.aif_national_codes = []
        adapter.historical_risk_profiles = []
        adapter.finance_records = []
        adapter.strategies = []
        adapter.turnovers = []
        adapter.risks = []
        adapter.counterparty_risks = []
        adapter.investor_liquidity_profiles = []
        adapter.investor_redemptions = []
        adapter.investor_groups = []
        adapter.stress_tests = []
        adapter.portfolio_liquidity_profiles = []
        adapter.portfolio_positions = []
        adapter.custom_fx_rates = []
        adapter.master_records = []
        adapter.special_arrangements = []
        adapter.dominant_influences = []
        adapter.prime_brokers = []
        adapter.controlled_structures = []
        adapter.borrow_sources = []
        adapter.financing_liquidity = []
        adapter._ecb_fx_rate_cache = {}

        # Populate AIFM from canonical report
        aifm_fields = aifm_report.fields

        # Map AIFM scalar fields back to the aifm dict
        if aifm_report.get_value("1"):
            adapter.aifm["AIF(M) Reporting Member State"] = aifm_report.get_value("1")
        if aifm_report.get_value("4"):
            adapter.aifm["Filing Type"] = aifm_report.get_value("4")
        if aifm_report.get_value("8"):
            adapter.aifm["Reporting Period Type"] = aifm_report.get_value("8")
        if aifm_report.get_value("9"):
            try:
                adapter.aifm["Reporting Period Year"] = int(aifm_report.get_value("9"))
            except (ValueError, TypeError):
                pass
        if aifm_report.get_value("13"):
            adapter.aifm["Last Reporting Flag"] = aifm_report.get_value("13")
        if aifm_report.get_value("16"):
            adapter.aifm["AIFM Reporting Code"] = aifm_report.get_value("16")
        if aifm_report.get_value("17"):
            adapter.aifm["AIFM Jurisdiction"] = aifm_report.get_value("17")
        if aifm_report.get_value("18"):
            adapter.aifm["AIFM National Code"] = aifm_report.get_value("18")
        if aifm_report.get_value("19"):
            adapter.aifm["AIFM Name"] = aifm_report.get_value("19")
        if aifm_report.get_value("21"):
            adapter.aifm["AIFM no reporting flag (Nothing to report)"] = aifm_report.get_value("21")
        if aifm_report.get_value("22"):
            adapter.aifm["AIFM LEI code"] = aifm_report.get_value("22")
        if aifm_report.get_value("23"):
            adapter.aifm["AIFM BIC code"] = aifm_report.get_value("23")
        if aifm_report.get_value("24"):
            adapter.aifm["AIFM old identifier 1"] = aifm_report.get_value("24")
        if aifm_report.get_value("25"):
            adapter.aifm["AIFM old identifier 2"] = aifm_report.get_value("25")

        # Restore assumptions
        assumptions_group = aifm_report.get_group_values("assumptions")
        for a in assumptions_group:
            assumption_rec = {}
            if "question" in a:
                assumption_rec["Question Number"] = a["question"]
            if "description" in a:
                assumption_rec["Assumption Description"] = a["description"]
            if assumption_rec:
                adapter.aifm_assumptions.append(assumption_rec)

        # Restore AIFM national codes
        aifm_ncs_group = aifm_report.get_group_values("aifm_national_codes")
        for nc in aifm_ncs_group:
            adapter.aifm_national_codes.append(nc)

        # Populate AIFs from canonical reports
        for aif_report in aif_reports:
            # Create MRecord instead of plain dict to preserve case-insensitive lookups
            aif_dict = MRecord()

            # Map AIF scalar fields
            if aif_report.get_value("1"):
                aif_dict["Custom AIF Identification"] = aif_report.get_value("1")
            if aif_report.get_value("2"):
                aif_dict["AIF Name"] = aif_report.get_value("2")
            if aif_report.get_value("3"):
                aif_dict["Domicile of the AIF"] = aif_report.get_value("3")
            if aif_report.get_value("4"):
                aif_dict["AIF Content Type"] = aif_report.get_value("4")
            if aif_report.get_value("5"):
                aif_dict["AIF LEI code"] = aif_report.get_value("5")
            if aif_report.get_value("6"):
                aif_dict["Investment strategy code"] = aif_report.get_value("6")
            if aif_report.get_value("7"):
                aif_dict["AIF Type"] = aif_report.get_value("7")
            if aif_report.get_value("8"):
                aif_dict["Base currency of the AIF"] = aif_report.get_value("8")
            if aif_report.get_value("9"):
                aif_dict["Total AuM amount of the AIF in base currency"] = aif_report.get_value("9")
            if aif_report.get_value("10"):
                aif_dict["Total Net Asset Value of the AIF (NAV)"] = aif_report.get_value("10")
            if aif_report.get_value("11"):
                aif_dict["AIF Reporting Code"] = aif_report.get_value("11")
            if aif_report.get_value("12"):
                aif_dict["Inception date"] = aif_report.get_value("12")
            if aif_report.get_value("13"):
                aif_dict["Last Reporting Flag"] = aif_report.get_value("13")

            # Restore complete AIF dict from special storage
            aif_complete_group = aif_report.get_group_values("_aif_complete_data")
            if aif_complete_group:
                # Parse JSON to restore original types
                import json
                json_str = aif_complete_group[0].get("_json", "{}")
                try:
                    complete_data = json.loads(json_str)
                    for key, val in complete_data.items():
                        if key not in aif_dict or aif_dict[key] is None:
                            aif_dict[key] = val
                except (json.JSONDecodeError, TypeError):
                    pass

            # Rebuild the case-insensitive map after populating all keys
            if isinstance(aif_dict, MRecord):
                aif_dict._rebuild_lower_map()

            adapter.aifs.append(aif_dict)

            aif_id = aif_dict.get("Custom AIF Identification", "")

            # Restore AIF national codes for this AIF
            aif_ncs_group = aif_report.get_group_values("aif_national_codes")
            for nc in aif_ncs_group:
                adapter.aif_national_codes.append(nc)

            # Restore repeating groups for this AIF
            # Positions
            positions_group = aif_report.get_group_values("positions")
            for pos in positions_group:
                adapter.positions.append(pos)

            # Share classes
            share_classes_group = aif_report.get_group_values("share_classes")
            for sc in share_classes_group:
                adapter.share_classes.append(sc)

            # Strategies
            strategies_group = aif_report.get_group_values("strategies")
            for st in strategies_group:
                adapter.strategies.append(st)

            # Turnovers
            turnovers_group = aif_report.get_group_values("turnovers")
            for t in turnovers_group:
                adapter.turnovers.append(t)

            # Risks
            risks_group = aif_report.get_group_values("risks")
            for r in risks_group:
                adapter.risks.append(r)

            # Counterparty risks
            cpr_group = aif_report.get_group_values("counterparty_risks")
            for cp in cpr_group:
                adapter.counterparty_risks.append(cp)

            # Investor groups
            investor_groups = aif_report.get_group_values("investor_groups")
            for ig in investor_groups:
                adapter.investor_groups.append(ig)

            # Stress tests
            stress_tests = aif_report.get_group_values("stress_tests")
            for st in stress_tests:
                adapter.stress_tests.append(st)

            # Full-template only groups
            if adapter.template_type == "FULL":
                # Historical risk profiles
                hrp_group = aif_report.get_group_values("historical_risk_profiles")
                for h in hrp_group:
                    adapter.historical_risk_profiles.append(h)

                # Portfolio liquidity profiles
                plp_group = aif_report.get_group_values("portfolio_liquidity_profiles")
                for pl in plp_group:
                    adapter.portfolio_liquidity_profiles.append(pl)

                # Investor liquidity profiles
                ilp_group = aif_report.get_group_values("investor_liquidity_profiles")
                for il in ilp_group:
                    adapter.investor_liquidity_profiles.append(il)

                # Investor redemptions
                ir_group = aif_report.get_group_values("investor_redemptions")
                for ir in ir_group:
                    adapter.investor_redemptions.append(ir)

                # Dominant influences
                di_group = aif_report.get_group_values("dominant_influences")
                for di in di_group:
                    adapter.dominant_influences.append(di)

                # Prime brokers
                pb_group = aif_report.get_group_values("prime_brokers")
                for pb in pb_group:
                    adapter.prime_brokers.append(pb)

                # Controlled structures
                cs_group = aif_report.get_group_values("controlled_structures")
                for cs in cs_group:
                    adapter.controlled_structures.append(cs)

                # Borrow sources
                bs_group = aif_report.get_group_values("borrow_sources")
                for bs in bs_group:
                    adapter.borrow_sources.append(bs)

                # Financing liquidity
                fl_group = aif_report.get_group_values("financing_liquidity")
                for fl in fl_group:
                    adapter.financing_liquidity.append(fl)

                # Special arrangements
                sa_group = aif_report.get_group_values("special_arrangements")
                for sa in sa_group:
                    adapter.special_arrangements.append(sa)

                # Master records
                mr_group = aif_report.get_group_values("master_records")
                for mr in mr_group:
                    adapter.master_records.append(mr)

        # Derive fields from canonical data
        adapter.reporting_member_state = aifm_report.get_value("1") or ""
        adapter.reporting_year = int(aifm_report.get_value("9")) if aifm_report.get_value("9") else 0
        adapter.reporting_period_type = aifm_report.get_value("8") or "Y1"
        adapter.aifm_national_code = aifm_report.get_value("18") or ""
        adapter.aifm_name = aifm_report.get_value("19") or ""
        adapter.aifm_jurisdiction = aifm_report.get_value("17") or adapter.reporting_member_state
        adapter.filing_type = aifm_report.get_value("4") or "INIT"
        adapter.aifm_lei = aifm_report.get_value("22") or ""
        adapter.aifm_bic = aifm_report.get_value("23") or ""

        return adapter

# ── CLI entry point ──────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json
    import sys

    if len(sys.argv) < 2:
        print("Usage: python m_adapter.py <template.xlsx> [output_dir]")
        sys.exit(1)

    xlsx = sys.argv[1]
    out_dir = sys.argv[2] if len(sys.argv) > 2 else None

    adapter = MAdapter(xlsx)
    print("Template summary:")
    print(json.dumps(adapter.summary(), indent=2))

    result = adapter.generate_all(out_dir)
    print(f"\nGenerated: {result}")

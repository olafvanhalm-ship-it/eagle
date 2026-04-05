#!/usr/bin/env python3
"""
AIFMD Golden Set Regression Suite — True End-to-End
====================================================
For each suite, reads the M adapter Excel template, generates AIFM/AIF XML
(and NCA packaging), then:

  1. Compares generated XML against reference XML byte-by-byte (ignoring
     cosmetic differences like timestamps & comments).
  2. Checks file naming conventions (AIFM_CC_NNN_YYYYMMDD.xml).
  3. Verifies NCA packaging (DE → .gz, BE → .zip, NL multi-AIF → .zip).
  4. Runs XSD & DQF validation on the generated XML.
  5. Compares DQF/naming findings against expected_findings.yaml baseline.

Usage:
    python run_regression_suite.py                        # Run all suites
    python run_regression_suite.py --scope realdata        # Quick: M examples + authorised_anon only
    python run_regression_suite.py --scope synthetic       # Synthetic suites only
    python run_regression_suite.py --suite m_light_nl      # Run one
    python run_regression_suite.py --with-enrichment       # Run with LEI enrichment from GLEIF seed
    python run_regression_suite.py --update-baseline       # Accept current as baseline
    python run_regression_suite.py --capture-expected       # Save current findings as expected baseline
    python run_regression_suite.py --compliance-report     # Generate Excel compliance report
    python run_regression_suite.py --freshness             # Reference data age

Evidence output (per run):
    golden_set/evidence/regression_{timestamp}.yaml
    golden_set/evidence/compliance_{timestamp}.xlsx   (with --compliance-report)
"""

import argparse
import gzip
import hashlib
import warnings

# Suppress openpyxl's Data Validation warning (the templates use dropdowns
# that openpyxl doesn't support — harmless, cell values read fine)
warnings.filterwarnings("ignore",
                        message="Data Validation extension is not supported",
                        category=UserWarning,
                        module="openpyxl")
import re
import shutil
import subprocess
import sys
import tempfile
import traceback
import zipfile
import yaml
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET

# ── Paths ────────────────────────────────────────────────────────────────────

_SCRIPT_DIR = Path(__file__).resolve().parent
_GOLDEN_SET_DIR = _SCRIPT_DIR / "golden_set"

# Auto-detect project root (directory containing Application/ or Blueprint/)
_PROJECT_ROOT = _SCRIPT_DIR
for _p in [_SCRIPT_DIR] + [_SCRIPT_DIR.parents[i] for i in range(5)]:
    if (_p / "Application").is_dir():
        _PROJECT_ROOT = _p
        break
    if (_p / "Blueprint").is_dir():
        _PROJECT_ROOT = _p
        break

_APP_ROOT = _PROJECT_ROOT / "Application"
_VALIDATOR = _APP_ROOT / "validation" / "validate_aifmd_xml.py"

# Add Application root to sys.path so we can import shared modules
if str(_APP_ROOT) not in sys.path:
    sys.path.insert(0, str(_APP_ROOT))

from aifmd_packaging.aifmd_nca_packaging import NCA_PACKAGING_CONFIG, get_expected_extensions

_AIFMD_ANNEX_IV_DIR = _APP_ROOT / "regulation" / "aifmd" / "annex_iv"
if not _AIFMD_ANNEX_IV_DIR.is_dir():
    # Fallback: try old flat location, then legacy Blueprint
    _AIFMD_ANNEX_IV_DIR = _APP_ROOT / "regulation" / "aifmd annex iv"
    if not _AIFMD_ANNEX_IV_DIR.is_dir():
        _AIFMD_ANNEX_IV_DIR = _PROJECT_ROOT / "Blueprint"

DEFAULT_RULES_PATH = _AIFMD_ANNEX_IV_DIR / "aifmd_validation_rules.yaml"
if not DEFAULT_RULES_PATH.exists():
    DEFAULT_RULES_PATH = _AIFMD_ANNEX_IV_DIR / "aifmd_annex_iv_validation_rules.yaml"
NCA_OVERRIDES_DIR = _AIFMD_ANNEX_IV_DIR / "nca_overrides"
if not NCA_OVERRIDES_DIR.is_dir():
    NCA_OVERRIDES_DIR = _AIFMD_ANNEX_IV_DIR
EXPECTED_FINDINGS_PATH = _GOLDEN_SET_DIR / "expected_findings.yaml"
_GLEIF_SEED_PATH = _GOLDEN_SET_DIR / "gleif_seed.yaml"


# ── LEI Enrichment Seed Loader ──────────────────────────────────────────────

def _load_gleif_seed_store():
    """Load gleif_seed.yaml into an in-memory SQLite ReferenceStore.

    Returns a ReferenceStore populated with the synthetic GLEIF cache
    entries defined in the seed file, or None if the seed file is missing
    or the ReferenceStore cannot be imported.
    """
    if not _GLEIF_SEED_PATH.exists():
        print(f"  [WARN] GLEIF seed file not found: {_GLEIF_SEED_PATH.name}")
        return None

    # Ensure Application/ is on sys.path for shared module imports
    if str(_APP_ROOT) not in sys.path:
        sys.path.insert(0, str(_APP_ROOT))

    try:
        from shared.reference_store import ReferenceStore
    except ImportError as e:
        print(f"  [WARN] Cannot import ReferenceStore: {e}")
        return None

    with open(str(_GLEIF_SEED_PATH), "r", encoding="utf-8") as f:
        seed_records = yaml.safe_load(f) or []

    # Create in-memory SQLite store and populate with seed data
    store = ReferenceStore.sqlite(":memory:")

    # Convert seed records to the format expected by upsert_lei
    lei_records = []
    for rec in seed_records:
        lei_records.append({
            "lei": rec["lei"],
            "legal_name": rec["legal_name"],
            "entity_status": rec.get("entity_status", "ACTIVE"),
            "country": rec.get("country", ""),
            "registration_authority": None,
            "last_update": None,
        })

    store.upsert_lei(lei_records)
    return store


# ── Expected Findings Baseline ───────────────────────────────────────────────

def _load_expected_findings() -> dict:
    """Load the expected findings baseline from YAML."""
    if not EXPECTED_FINDINGS_PATH.exists():
        return {}
    with open(str(EXPECTED_FINDINGS_PATH), "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _get_expected_count(expected: dict, metric: str) -> int:
    """Extract the expected count for a metric from the baseline entry."""
    val = expected.get(metric, 0)
    if isinstance(val, dict):
        return val.get("count", 0)
    return int(val) if val else 0


def _get_expected_rationale(baseline: dict, expected: dict, metric: str) -> str:
    """Resolve rationale text, following $ref to the rationale catalogue."""
    val = expected.get(metric, {})
    if not isinstance(val, dict):
        return ""
    raw = val.get("rationale", "")
    if isinstance(raw, str) and raw.startswith("$"):
        catalogue = baseline.get("rationale", {})
        return catalogue.get(raw[1:], raw).strip()
    return str(raw).strip() if raw else ""


def _compare_findings(suite_name: str, validation: dict, baseline: dict,
                      is_multi_nca: bool = False) -> dict:
    """Compare actual validation metrics against expected findings baseline.

    Returns a dict with:
      - matched: list of {metric, expected, actual, rationale}
      - regressions: list of {metric, expected, actual, direction}
    """
    suite_baseline = baseline.get("suites", {}).get(suite_name, {})
    if not suite_baseline:
        return {"matched": [], "regressions": [],
                "note": "No baseline defined — run with --capture-expected"}

    if is_multi_nca and suite_baseline.get("multi_nca"):
        return _compare_multi_nca(suite_name, validation, suite_baseline, baseline)

    expected = suite_baseline.get("expected", {})
    return _compare_single(validation, expected, baseline)


def _compare_single(validation: dict, expected: dict, baseline: dict) -> dict:
    """Compare a single-NCA validation result against expected findings."""
    matched = []
    regressions = []
    metrics = ["xsd_valid", "xsd_invalid", "dqf_fail", "dqf_missing",
               "dqf_warning", "file_naming_fail"]

    for metric in metrics:
        exp = _get_expected_count(expected, metric)
        act = validation.get(metric, 0)
        rationale = _get_expected_rationale(baseline, expected, metric)
        entry = {"metric": metric, "expected": exp, "actual": act,
                 "rationale": rationale}
        if act == exp:
            matched.append(entry)
        else:
            entry["direction"] = "NEW" if act > exp else "RESOLVED"
            regressions.append(entry)

    return {"matched": matched, "regressions": regressions}


def _compare_multi_nca(suite_name: str, validation: dict,
                       suite_baseline: dict, baseline: dict) -> dict:
    """Compare multi-NCA validation results against per-NCA expected findings."""
    all_matched = []
    all_regressions = []

    for nca_key, nca_val in validation.items():
        if not isinstance(nca_val, dict) or "xsd_valid" not in nca_val:
            continue
        nca_expected = suite_baseline.get("per_nca", {}).get(nca_key, {}).get("expected", {})
        result = _compare_single(nca_val, nca_expected, baseline)
        for m in result["matched"]:
            m["nca"] = nca_key
            all_matched.append(m)
        for r in result["regressions"]:
            r["nca"] = nca_key
            all_regressions.append(r)

    return {"matched": all_matched, "regressions": all_regressions}


# ── Suite Configuration ──────────────────────────────────────────────────────
# Each suite maps to a golden_set/ subdirectory that MUST contain:
#   - At least one .xlsx (or .xlsx + .csv for split templates)
#   - Reference XML files (AIFM_*.xml, AIF_*.xml) and/or packaged outputs
#
# multi_nca is special: one template produces outputs for 5 NCAs.
# Its reference XMLs live in sub-directories (m_multi_nca_{nca}/).

SUITES = {
    # ── M example templates (fictitious data) ──────────────────
    "m_light_nl": {
        "description": "M Light Template, 2 AIFs with positions, RE strategy (NL/AFM)",
        "category": "realdata",
        "nca": "NL",
        "directory": "m_light_nl",
        "expected_packaging": ["zip"],   # multi-AIF → ZIP
    },
    "m_full_nl": {
        "description": "M Full Template, 2 AIFs with positions, RE strategy (NL/AFM)",
        "category": "realdata",
        "nca": "NL",
        "directory": "m_full_nl",
        "expected_packaging": ["zip"],
    },
    "m_masterfeeder_nl": {
        "description": "Explicit Master-Feeder template, feeder with master AIF identification (NL/AFM)",
        "category": "realdata",
        "nca": "NL",
        "directory": "m_masterfeeder_nl",
        "expected_packaging": ["zip"],
    },
    "m_positions_nl": {
        "description": "Light template with full position detail, single AIF (NL/AFM)",
        "category": "realdata",
        "nca": "NL",
        "directory": "m_positions_nl",
        "expected_packaging": [],  # single AIF → plain XML
    },
    "m_split_nl": {
        "description": "Split template (Excel+CSV positions), empty national codes (NL/AFM)",
        "category": "realdata",
        "nca": "NL",
        "directory": "m_split_nl",
        "expected_packaging": ["zip"],
    },
    "m_cancellation_nl": {
        "description": "Cancellation filing, ESMA CancellationRecordInfo structure (NL/AFM)",
        "category": "realdata",
        "nca": "NL",
        "directory": "m_cancellation_nl",
        "expected_packaging": [],
    },
    "m_amnd_full_nl": {
        "description": "Full Template AMND filing, 2 AIFs with positions, RE strategy (NL/AFM)",
        "category": "realdata",
        "nca": "NL",
        "directory": "m_amnd_full_nl",
        "expected_packaging": ["zip"],
    },
    "m_amnd_light_nl": {
        "description": "Light Template AMND filing, 1 AIF (#-excluded fund), RE strategy (NL/AFM)",
        "category": "realdata",
        "nca": "NL",
        "directory": "m_amnd_light_nl",
        "expected_packaging": [],
    },

    # ── Multi-NCA template (single template → 5 NCA outputs) ────────────
    "m_multi_nca": {
        "description": "Multi-jurisdiction template: NL+BE+DE+GB+LU outputs from single template",
        "category": "realdata",
        "nca": "ALL",      # special: iterates all NCAs
        "directory": "m_multi_nca",
        "expected_packaging": ["zip", "gz"],  # BE→zip, DE→gz, NL→zip, LU→zip; GB→consolidated XML (no zip)
        "multi_nca": True,
        "reference_subdirs": {
            "NL": "m_multi_nca_nl",
            "BE": "m_multi_nca_be",
            "DE": "m_multi_nca_de",
            "GB": "m_multi_nca_gb",
            "LU": "m_multi_nca_lu",
        },
    },

    # ── Synthetic system tests (exhaustive domain value coverage) ────────
    "synthetic_light_nl_init": {
        "description": "Synthetic Light Template, 8 AIFs, NL/AFM, INIT — exhaustive domain values",
        "category": "synthetic",
        "nca": "NL",
        "directory": "synthetic_light_nl_init",
        "expected_packaging": ["zip"],
    },
    "synthetic_light_de_init": {
        "description": "Synthetic Light Template, 8 AIFs, DE/BaFin, INIT",
        "category": "synthetic",
        "nca": "DE",
        "directory": "synthetic_light_de_init",
        "expected_packaging": ["gz"],
    },
    "synthetic_light_be_init": {
        "description": "Synthetic Light Template, 8 AIFs, BE/FSMA, INIT",
        "category": "synthetic",
        "nca": "BE",
        "directory": "synthetic_light_be_init",
        "expected_packaging": ["zip"],
    },
    "synthetic_full_nl_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), NL/AFM, INIT — all Art 24 sections",
        "category": "synthetic",
        "nca": "NL",
        "directory": "synthetic_full_nl_init",
        "expected_packaging": ["zip"],
    },
    "synthetic_full_de_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), DE/BaFin, INIT",
        "category": "synthetic",
        "nca": "DE",
        "directory": "synthetic_full_de_init",
        "expected_packaging": ["gz"],
    },
    "synthetic_full_be_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), BE/FSMA, INIT",
        "category": "synthetic",
        "nca": "BE",
        "directory": "synthetic_full_be_init",
        "expected_packaging": ["zip"],
    },
    "synthetic_amnd_nl": {
        "description": "Synthetic Full Template, AMND filing, NL/AFM — obligation change codes",
        "category": "synthetic",
        "nca": "NL",
        "directory": "synthetic_amnd_nl",
        "expected_packaging": ["zip"],
    },
    "synthetic_cancel_nl": {
        "description": "Synthetic Light Template, CANCEL filing, NL/AFM",
        "category": "synthetic",
        "nca": "NL",
        "directory": "synthetic_cancel_nl",
        "expected_packaging": ["zip"],
    },
    "synthetic_light_gb_init": {
        "description": "Synthetic FCA Light Template, 8 AIFs, GB/FCA, INIT — v2 XSD, consolidated AIF",
        "category": "synthetic",
        "nca": "GB",
        "directory": "synthetic_light_gb_init",
        "expected_packaging": [],  # FCA consolidated: all AIFs in single XML, no ZIP needed
    },
    "synthetic_full_gb_init": {
        "description": "Synthetic FCA Full Template, 6 AIFs (CT=4/5), GB/FCA, INIT — v2 XSD, consolidated AIF",
        "category": "synthetic",
        "nca": "GB",
        "directory": "synthetic_full_gb_init",
        "expected_packaging": [],  # FCA consolidated: all AIFs in single XML, no ZIP needed
    },
    "synthetic_negative_crossfield": {
        "description": "Synthetic negative tests — deliberately invalid cross-field combinations",
        "category": "synthetic",
        "nca": "NL",
        "directory": "synthetic_negative_crossfield",
        "expected_packaging": ["zip"],
    },
    # ── Remaining 27 ESMA NCAs (light + full) ──────────────────────────
    "synthetic_light_at_init": {
        "description": "Synthetic Light Template, 8 AIFs, AT/FMA, INIT",
        "category": "synthetic",
        "nca": "AT", "directory": "synthetic_light_at_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_at_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), AT/FMA, INIT",
        "category": "synthetic",
        "nca": "AT", "directory": "synthetic_full_at_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_bg_init": {
        "description": "Synthetic Light Template, 8 AIFs, BG/FSC, INIT",
        "category": "synthetic",
        "nca": "BG", "directory": "synthetic_light_bg_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_bg_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), BG/FSC, INIT",
        "category": "synthetic",
        "nca": "BG", "directory": "synthetic_full_bg_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_cy_init": {
        "description": "Synthetic Light Template, 8 AIFs, CY/CySEC, INIT",
        "category": "synthetic",
        "nca": "CY", "directory": "synthetic_light_cy_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_cy_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), CY/CySEC, INIT",
        "category": "synthetic",
        "nca": "CY", "directory": "synthetic_full_cy_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_cz_init": {
        "description": "Synthetic Light Template, 8 AIFs, CZ/CNB, INIT",
        "category": "synthetic",
        "nca": "CZ", "directory": "synthetic_light_cz_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_cz_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), CZ/CNB, INIT",
        "category": "synthetic",
        "nca": "CZ", "directory": "synthetic_full_cz_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_dk_init": {
        "description": "Synthetic Light Template, 8 AIFs, DK/Finanstilsynet, INIT",
        "category": "synthetic",
        "nca": "DK", "directory": "synthetic_light_dk_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_dk_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), DK/Finanstilsynet, INIT",
        "category": "synthetic",
        "nca": "DK", "directory": "synthetic_full_dk_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_ee_init": {
        "description": "Synthetic Light Template, 8 AIFs, EE/FSA, INIT",
        "category": "synthetic",
        "nca": "EE", "directory": "synthetic_light_ee_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_ee_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), EE/FSA, INIT",
        "category": "synthetic",
        "nca": "EE", "directory": "synthetic_full_ee_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_es_init": {
        "description": "Synthetic Light Template, 8 AIFs, ES/CNMV, INIT",
        "category": "synthetic",
        "nca": "ES", "directory": "synthetic_light_es_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_es_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), ES/CNMV, INIT",
        "category": "synthetic",
        "nca": "ES", "directory": "synthetic_full_es_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_fi_init": {
        "description": "Synthetic Light Template, 8 AIFs, FI/FIN-FSA, INIT",
        "category": "synthetic",
        "nca": "FI", "directory": "synthetic_light_fi_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_fi_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), FI/FIN-FSA, INIT",
        "category": "synthetic",
        "nca": "FI", "directory": "synthetic_full_fi_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_fr_init": {
        "description": "Synthetic Light Template, 8 AIFs, FR/AMF, INIT",
        "category": "synthetic",
        "nca": "FR", "directory": "synthetic_light_fr_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_fr_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), FR/AMF, INIT",
        "category": "synthetic",
        "nca": "FR", "directory": "synthetic_full_fr_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_gr_init": {
        "description": "Synthetic Light Template, 8 AIFs, GR/HCMC, INIT",
        "category": "synthetic",
        "nca": "GR", "directory": "synthetic_light_gr_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_gr_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), GR/HCMC, INIT",
        "category": "synthetic",
        "nca": "GR", "directory": "synthetic_full_gr_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_hr_init": {
        "description": "Synthetic Light Template, 8 AIFs, HR/HANFA, INIT",
        "category": "synthetic",
        "nca": "HR", "directory": "synthetic_light_hr_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_hr_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), HR/HANFA, INIT",
        "category": "synthetic",
        "nca": "HR", "directory": "synthetic_full_hr_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_hu_init": {
        "description": "Synthetic Light Template, 8 AIFs, HU/MNB, INIT",
        "category": "synthetic",
        "nca": "HU", "directory": "synthetic_light_hu_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_hu_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), HU/MNB, INIT",
        "category": "synthetic",
        "nca": "HU", "directory": "synthetic_full_hu_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_ie_init": {
        "description": "Synthetic Light Template, 8 AIFs, IE/CBI, INIT",
        "category": "synthetic",
        "nca": "IE", "directory": "synthetic_light_ie_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_ie_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), IE/CBI, INIT",
        "category": "synthetic",
        "nca": "IE", "directory": "synthetic_full_ie_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_is_init": {
        "description": "Synthetic Light Template, 8 AIFs, IS/CBI, INIT",
        "category": "synthetic",
        "nca": "IS", "directory": "synthetic_light_is_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_is_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), IS/CBI, INIT",
        "category": "synthetic",
        "nca": "IS", "directory": "synthetic_full_is_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_it_init": {
        "description": "Synthetic Light Template, 8 AIFs, IT/CONSOB, INIT",
        "category": "synthetic",
        "nca": "IT", "directory": "synthetic_light_it_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_it_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), IT/CONSOB, INIT",
        "category": "synthetic",
        "nca": "IT", "directory": "synthetic_full_it_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_li_init": {
        "description": "Synthetic Light Template, 8 AIFs, LI/FMA, INIT",
        "category": "synthetic",
        "nca": "LI", "directory": "synthetic_light_li_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_li_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), LI/FMA, INIT",
        "category": "synthetic",
        "nca": "LI", "directory": "synthetic_full_li_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_lt_init": {
        "description": "Synthetic Light Template, 8 AIFs, LT/BoL, INIT",
        "category": "synthetic",
        "nca": "LT", "directory": "synthetic_light_lt_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_lt_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), LT/BoL, INIT",
        "category": "synthetic",
        "nca": "LT", "directory": "synthetic_full_lt_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_lu_init": {
        "description": "Synthetic Light Template, 8 AIFs, LU/CSSF, INIT",
        "category": "synthetic",
        "nca": "LU", "directory": "synthetic_light_lu_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_lu_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), LU/CSSF, INIT",
        "category": "synthetic",
        "nca": "LU", "directory": "synthetic_full_lu_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_lv_init": {
        "description": "Synthetic Light Template, 8 AIFs, LV/BoL, INIT",
        "category": "synthetic",
        "nca": "LV", "directory": "synthetic_light_lv_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_lv_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), LV/BoL, INIT",
        "category": "synthetic",
        "nca": "LV", "directory": "synthetic_full_lv_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_mt_init": {
        "description": "Synthetic Light Template, 8 AIFs, MT/MFSA, INIT",
        "category": "synthetic",
        "nca": "MT", "directory": "synthetic_light_mt_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_mt_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), MT/MFSA, INIT",
        "category": "synthetic",
        "nca": "MT", "directory": "synthetic_full_mt_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_no_init": {
        "description": "Synthetic Light Template, 8 AIFs, NO/Finanstilsynet, INIT",
        "category": "synthetic",
        "nca": "NO", "directory": "synthetic_light_no_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_no_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), NO/Finanstilsynet, INIT",
        "category": "synthetic",
        "nca": "NO", "directory": "synthetic_full_no_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_pl_init": {
        "description": "Synthetic Light Template, 8 AIFs, PL/KNF, INIT",
        "category": "synthetic",
        "nca": "PL", "directory": "synthetic_light_pl_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_pl_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), PL/KNF, INIT",
        "category": "synthetic",
        "nca": "PL", "directory": "synthetic_full_pl_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_pt_init": {
        "description": "Synthetic Light Template, 8 AIFs, PT/CMVM, INIT",
        "category": "synthetic",
        "nca": "PT", "directory": "synthetic_light_pt_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_pt_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), PT/CMVM, INIT",
        "category": "synthetic",
        "nca": "PT", "directory": "synthetic_full_pt_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_ro_init": {
        "description": "Synthetic Light Template, 8 AIFs, RO/ASF, INIT",
        "category": "synthetic",
        "nca": "RO", "directory": "synthetic_light_ro_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_ro_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), RO/ASF, INIT",
        "category": "synthetic",
        "nca": "RO", "directory": "synthetic_full_ro_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_se_init": {
        "description": "Synthetic Light Template, 8 AIFs, SE/FI, INIT",
        "category": "synthetic",
        "nca": "SE", "directory": "synthetic_light_se_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_se_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), SE/FI, INIT",
        "category": "synthetic",
        "nca": "SE", "directory": "synthetic_full_se_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_si_init": {
        "description": "Synthetic Light Template, 8 AIFs, SI/ATVP, INIT",
        "category": "synthetic",
        "nca": "SI", "directory": "synthetic_light_si_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_si_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), SI/ATVP, INIT",
        "category": "synthetic",
        "nca": "SI", "directory": "synthetic_full_si_init", "expected_packaging": ["zip"],
    },
    "synthetic_light_sk_init": {
        "description": "Synthetic Light Template, 8 AIFs, SK/NBS, INIT",
        "category": "synthetic",
        "nca": "SK", "directory": "synthetic_light_sk_init", "expected_packaging": ["zip"],
    },
    "synthetic_full_sk_init": {
        "description": "Synthetic Full Template, 6 AIFs (CT=4/5), SK/NBS, INIT",
        "category": "synthetic",
        "nca": "SK", "directory": "synthetic_full_sk_init", "expected_packaging": ["zip"],
    },

    # ── Real client templates (authorised_anon) ─────────────────────────
    # authorised_anon contains multiple independent templates (each a different
    # AIFM). Each must be tested as its own adapter instance with its own
    # reference XMLs. TODO: implement per-template sub-suite discovery so each
    # template runs in isolation. For now, these are tested manually or via
    # --suite with individual template names.
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def _find_templates(suite_dir: Path) -> list[Path]:
    """Find all input files (.xlsx + .csv) in a suite directory."""
    files = sorted(suite_dir.glob("*.xlsx")) + sorted(suite_dir.glob("*.csv"))
    return files


def _find_reference_xmls(suite_dir: Path) -> list[Path]:
    """Find reference XML files in a suite directory (AIFM_*.xml, AIF_*.xml, AIF_REPORTS_*.xml)."""
    return sorted(
        f for f in suite_dir.glob("*.xml")
        if f.name.startswith("AIFM_") or f.name.startswith("AIF_")
    )


def _find_reference_packages(suite_dir: Path) -> list[Path]:
    """Find reference packages (.zip, .gz) in a suite directory."""
    return sorted(
        f for f in suite_dir.iterdir()
        if f.suffix in (".zip", ".gz")
        and (f.name.startswith("AIFM_") or f.name.startswith("AIF_")
             or f.name.startswith("AIFMD_"))
    )


# ── XML Comparison (semantic, ignoring cosmetic differences) ─────────────────

_FILENAME_PATTERN = re.compile(
    r"^(AIFMD|AIF_REPORTS|AIFM|AIF)_([A-Z]{2})_([A-Za-z0-9_#.-]+)_(\d{8})(_master)?\.(xml|gz|zip)$"
)


def _normalise_xml(xml_path: Path) -> str:
    """Parse and re-serialise XML for structural comparison.

    Strips:
      - XML comments (timestamps, version info)
      - Leading/trailing whitespace in element text
      - Attribute ordering differences
      - &apos; vs ' differences (handled by XML parser)
    """
    tree = ET.parse(str(xml_path))
    root = tree.getroot()

    # Remove comments (ET doesn't preserve them, so this is automatic)
    # Normalise whitespace in text/tail
    for el in root.iter():
        if el.text:
            el.text = el.text.strip()
        if el.tail:
            el.tail = el.tail.strip()

    return ET.tostring(root, encoding="unicode", xml_declaration=False)


def _compare_xml(generated: Path, reference: Path) -> list[str]:
    """Compare two XML files structurally. Returns list of difference descriptions."""
    diffs = []
    try:
        gen_norm = _normalise_xml(generated)
        ref_norm = _normalise_xml(reference)
    except ET.ParseError as e:
        return [f"XML parse error: {e}"]

    if gen_norm == ref_norm:
        return []

    # Detailed element-by-element comparison
    gen_tree = ET.ElementTree(ET.fromstring(gen_norm))
    ref_tree = ET.ElementTree(ET.fromstring(ref_norm))

    gen_elems = list(gen_tree.iter())
    ref_elems = list(ref_tree.iter())

    if len(gen_elems) != len(ref_elems):
        diffs.append(f"Element count: generated={len(gen_elems)}, reference={len(ref_elems)}")

    # Walk both trees in parallel
    for i, (ge, re_) in enumerate(zip(gen_elems, ref_elems)):
        tag_g = ge.tag.split("}")[-1] if "}" in ge.tag else ge.tag
        tag_r = re_.tag.split("}")[-1] if "}" in re_.tag else re_.tag
        if tag_g != tag_r:
            diffs.append(f"Element #{i}: tag mismatch: {tag_g} vs {tag_r}")
            continue
        text_g = (ge.text or "").strip()
        text_r = (re_.text or "").strip()
        if text_g != text_r:
            diffs.append(f"<{tag_g}>: '{text_g}' vs '{text_r}'")

    # Limit output
    if len(diffs) > 20:
        total = len(diffs)
        diffs = diffs[:20]
        diffs.append(f"... and {total - 20} more differences")

    return diffs


def _check_file_naming(filename: str) -> tuple[bool, str]:
    """Verify AIFMD file naming convention: TYPE_CC_CODE_YYYYMMDD.ext"""
    m = _FILENAME_PATTERN.match(filename)
    if not m:
        return False, f"Naming convention violation: {filename}"
    return True, ""


def _check_nca_packaging(generated_dir: Path, nca: str, num_aifs: int) -> list[str]:
    """Check that NCA packaging rules are followed.

    Uses NCA_PACKAGING_CONFIG as the single source of truth for packaging
    requirements across all 31 AIFMD jurisdictions.
    """
    issues = []
    files_in_dir = list(generated_dir.iterdir())
    gz_files = [f for f in files_in_dir if f.suffix == ".gz"]
    zip_files = [f for f in files_in_dir if f.suffix == ".zip"]
    xml_files = [f for f in files_in_dir if f.suffix == ".xml"
                 and (f.name.startswith("AIFM_") or f.name.startswith("AIF_")
                      or f.name.startswith("AIFMD_"))]

    spec = NCA_PACKAGING_CONFIG.get(nca, {})
    pkg_type = spec.get("packaging", "xml")
    nca_name = spec.get("nca_name", nca)

    if pkg_type == "gzip":
        if not gz_files:
            issues.append(f"{nca}/{nca_name}: expected .gz files, found none")
        # Each XML should have a corresponding .gz
        expected_gz = len(xml_files)
        if gz_files and len(gz_files) < expected_gz:
            issues.append(
                f"{nca}/{nca_name}: expected {expected_gz} .gz files, found {len(gz_files)}")

    elif pkg_type in ("zip", "zip-in-zip"):
        if not zip_files:
            issues.append(f"{nca}/{nca_name}: expected .zip files, found none")

    elif pkg_type == "xml":
        # Plain XML — ZIP only expected when multi-AIF
        zip_when_multi = spec.get("zip_when_multi", True)
        if num_aifs > 1 and zip_when_multi and not zip_files:
            issues.append(
                f"Multi-AIF ({nca}): expected .zip bundle for {num_aifs} AIFs, found none")

    return issues


# ── Validation Runner ────────────────────────────────────────────────────────

def run_validation(xml_files: list[Path], nca: str) -> dict:
    """
    Run the 4-layer validator on XML files.
    Parses the console output to extract metrics.
    """
    if not xml_files:
        return {"error": "No XML files to validate"}

    # Ensure Logging directory exists
    logging_dir = _SCRIPT_DIR / "Logging"
    logging_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = logging_dir / f"aifmd_validation_{ts}.xlsx"

    cmd = [
        sys.executable, str(_VALIDATOR),
        "--nca", nca,
        "--output", str(output_path),
    ] + [str(f) for f in xml_files]

    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=120,
            cwd=str(_APP_ROOT)
        )
        output = result.stdout + result.stderr
        if not output.strip():
            # No output at all — log the command and return code for debugging
            output = (f"[validator returned no output; exit code {result.returncode}]\n"
                      f"[cmd: {' '.join(cmd[:4])} ...{len(xml_files)} files]\n"
                      f"[stderr: {result.stderr[:300] if result.stderr else '(empty)'}]")
    except subprocess.TimeoutExpired:
        return {"error": "Validation timed out (120s)"}

    metrics = {
        "total_files": len(xml_files),
        "xsd_valid": 0,
        "xsd_invalid": 0,
        "dqf_fail": 0,
        "dqf_missing": 0,
        "dqf_warning": 0,
        "dqf_pass": 0,
        "file_naming_fail": 0,
        "file_naming_warn": 0,
        "rule_integrity": "UNKNOWN",
        "raw_output": output,
    }

    for line in output.split("\n"):
        line = line.strip()
        if "XSD: VALID" in line:
            metrics["xsd_valid"] += 1
        elif "XSD: INVALID" in line:
            metrics["xsd_invalid"] += 1
        elif "DQF:" in line and "rules |" in line:
            m = re.search(r"FAIL=(\d+)", line)
            if m:
                metrics["dqf_fail"] += int(m.group(1))
            m = re.search(r"MISSING=(\d+)", line)
            if m:
                metrics["dqf_missing"] += int(m.group(1))
            m = re.search(r"WARNING=(\d+)", line)
            if m:
                metrics["dqf_warning"] += int(m.group(1))
            m = re.search(r"PASS=(\d+)", line)
            if m:
                metrics["dqf_pass"] += int(m.group(1))
        elif "File:" in line and "FAIL" in line:
            m = re.search(r"(\d+) FAIL", line)
            if m:
                metrics["file_naming_fail"] += int(m.group(1))
        elif "File:" in line and "WARNING" in line:
            m = re.search(r"(\d+) WARNING", line)
            if m:
                metrics["file_naming_warn"] += int(m.group(1))
        elif "Rule Integrity:" in line:
            if "VERIFIED" in line:
                metrics["rule_integrity"] = "VERIFIED"
            elif "UNVERIFIED" in line:
                metrics["rule_integrity"] = "UNVERIFIED"
            elif "FIRST RUN" in line:
                metrics["rule_integrity"] = "FIRST_RUN"

    return metrics


# ── E2E Test Runner ──────────────────────────────────────────────────────────

def run_e2e_suite(suite_name: str, suite_config: dict,
                  reference_store=None) -> dict:
    """Run a full E2E test for a single suite.

    Steps:
      1. Find template (.xlsx + optional .csv) in suite directory.
      2. Import template via MAdapter and generate XML to temp dir.
         When reference_store is provided, uses the canonical path
         (to_canonical_from_source → build_from_canonical) with LEI
         enrichment at L1B.
      3. Check file naming conventions on generated files.
      4. Check NCA packaging rules.
      5. Compare generated XML against reference XML.
      6. Run XSD/DQF validation on generated XML.

    Args:
        suite_name: Name of the suite (key in SUITES).
        suite_config: Suite configuration dict.
        reference_store: Optional ReferenceStore with GLEIF cache for LEI
            enrichment. If None, uses the standard generate_all() path.

    Returns a result dict with status, metrics, diffs, and issues.
    """
    suite_dir = _GOLDEN_SET_DIR / suite_config["directory"]
    result = {
        "status": "UNKNOWN",
        "generation": {},
        "naming_issues": [],
        "packaging_issues": [],
        "xml_diffs": {},
        "validation": {},
        "error": None,
    }

    # ── Step 1: Find template ────────────────────────────────────────────
    templates = _find_templates(suite_dir)
    if not templates:
        result["status"] = "SKIPPED"
        result["error"] = f"No .xlsx template found in {suite_dir.name}/"
        return result

    xlsx_files = [t for t in templates if t.suffix.lower() in (".xlsx", ".xls")]
    csv_files = [t for t in templates if t.suffix.lower() == ".csv"]

    # ── Step 2: Generate XML via MAdapter ────────────────────────────────
    tmp_dir = Path(tempfile.mkdtemp(prefix=f"eagle_e2e_{suite_name}_"))
    enrichment_log = None
    try:
        # Import MAdapter (add to path if needed)
        if str(_SCRIPT_DIR) not in sys.path:
            sys.path.insert(0, str(_SCRIPT_DIR))
        if str(_APP_ROOT) not in sys.path:
            sys.path.insert(0, str(_APP_ROOT))

        from m_adapter import MAdapter

        input_paths = [str(f) for f in xlsx_files + csv_files]
        adapter = MAdapter(*input_paths)

        # Run LEI enrichment analysis (if store provided)
        if reference_store is not None:
            try:
                _aifm_report, _aif_reports, enrichment_log = \
                    adapter.to_canonical_from_source(
                        reference_store=reference_store)
            except Exception as enr_err:
                # Enrichment failure should not block generation
                import logging
                logging.getLogger(__name__).warning(
                    f"Enrichment analysis failed: {enr_err}")

        # Standard generation path (always used for XML output)
        gen_result = adapter.generate_all(str(tmp_dir))

        generated_xmls = sorted(
            f for f in tmp_dir.glob("*.xml")
            if f.name.startswith("AIFM_") or f.name.startswith("AIF_")
        )
        generated_packages = sorted(
            f for f in tmp_dir.iterdir()
            if f.suffix in (".zip", ".gz")
        )

        result["generation"] = {
            "aifm_xmls": len(gen_result.get("aifm_xmls", [])),
            "aif_xmls": len(gen_result.get("aif_xmls", [])),
            "packages": len(gen_result.get("aif_zips", []))
                        + len(gen_result.get("gz_files", [])),
            "output_dir": str(tmp_dir),
        }

        # Capture LEI enrichment results if enrichment ran
        if enrichment_log is not None:
            enrichment_summary = {
                "enriched": 0, "pending_user_choice": 0,
                "no_match": 0, "skipped": 0, "error": 0,
            }
            for action in enrichment_log.actions:
                key = action.status.value.lower()
                enrichment_summary[key] = enrichment_summary.get(key, 0) + 1
            enrichment_summary["total"] = len(enrichment_log.actions)
            result["enrichment"] = enrichment_summary

    except Exception as e:
        result["status"] = "ERROR"
        result["error"] = f"Generation failed: {e}\n{traceback.format_exc()}"
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return result

    # ── Step 3: Check file naming ────────────────────────────────────────
    for f in generated_xmls + generated_packages:
        ok, msg = _check_file_naming(f.name)
        if not ok:
            result["naming_issues"].append(msg)

    # ── Step 4: Check NCA packaging ──────────────────────────────────────
    nca = suite_config["nca"]
    is_multi_nca = suite_config.get("multi_nca", False)

    if not is_multi_nca and nca != "ALL":
        num_aifs = len(gen_result.get("aif_xmls", []))
        pkg_issues = _check_nca_packaging(tmp_dir, nca, num_aifs)
        result["packaging_issues"].extend(pkg_issues)

    # ── Step 5: Compare generated XML against reference ──────────────────
    if is_multi_nca:
        # Multi-NCA: compare against per-NCA sub-directories
        for sub_nca, sub_dir_name in suite_config.get("reference_subdirs", {}).items():
            ref_dir = _GOLDEN_SET_DIR / sub_dir_name
            if not ref_dir.is_dir():
                result["xml_diffs"][f"{sub_nca}/"] = [f"Reference dir {sub_dir_name}/ not found"]
                continue
            ref_xmls = _find_reference_xmls(ref_dir)
            for ref_xml in ref_xmls:
                gen_xml = tmp_dir / ref_xml.name
                if not gen_xml.exists():
                    result["xml_diffs"][ref_xml.name] = [f"Generated file missing"]
                else:
                    diffs = _compare_xml(gen_xml, ref_xml)
                    if diffs:
                        result["xml_diffs"][ref_xml.name] = diffs

            # Check packaging per NCA sub-dir
            ref_pkgs = _find_reference_packages(ref_dir)
            for ref_pkg in ref_pkgs:
                gen_pkg = tmp_dir / ref_pkg.name
                if not gen_pkg.exists():
                    result["packaging_issues"].append(
                        f"{sub_nca}: expected package {ref_pkg.name} not generated")
    else:
        # Standard suite: compare against reference XMLs in same directory
        ref_xmls = _find_reference_xmls(suite_dir)
        for ref_xml in ref_xmls:
            gen_xml = tmp_dir / ref_xml.name
            if not gen_xml.exists():
                result["xml_diffs"][ref_xml.name] = [f"Generated file missing"]
            else:
                diffs = _compare_xml(gen_xml, ref_xml)
                if diffs:
                    result["xml_diffs"][ref_xml.name] = diffs

        # Check that no extra XMLs were generated
        ref_names = {f.name for f in ref_xmls}
        for gen_xml in generated_xmls:
            if gen_xml.name not in ref_names:
                result["xml_diffs"][gen_xml.name] = ["Extra file: not in reference set"]

    # ── Step 6: Run XSD/DQF validation ───────────────────────────────────
    if generated_xmls:
        # For multi-NCA we validate per-NCA
        if is_multi_nca:
            all_validation = {}
            for sub_nca in suite_config.get("reference_subdirs", {}).keys():
                nca_xmls = [
                    f for f in generated_xmls
                    if f"_{sub_nca}_" in f.name
                ]
                if nca_xmls:
                    val = run_validation(nca_xmls, sub_nca)
                    all_validation[sub_nca] = val
            result["validation"] = all_validation
        else:
            result["validation"] = run_validation(generated_xmls, nca)

    # ── Determine status (baseline-aware) ──────────────────────────────
    has_diffs = bool(result["xml_diffs"])
    has_naming = bool(result["naming_issues"])
    has_packaging = bool(result["packaging_issues"])
    has_gen_error = bool(result["error"])

    # Check validation (handle both single and multi-NCA)
    has_xsd_error = False
    if is_multi_nca and isinstance(result["validation"], dict):
        for nca_key, val in result["validation"].items():
            if isinstance(val, dict):
                has_xsd_error = has_xsd_error or val.get("xsd_invalid", 0) > 0
    elif isinstance(result["validation"], dict):
        has_xsd_error = result["validation"].get("xsd_invalid", 0) > 0

    # Compare findings against expected baseline
    baseline = _load_expected_findings()
    comparison = _compare_findings(
        suite_name, result["validation"], baseline, is_multi_nca)
    result["baseline_comparison"] = comparison
    has_regression = bool(comparison.get("regressions"))

    # XSD invalids that match the expected baseline are intentional (e.g. negative tests)
    xsd_invalid_expected = not has_regression and any(
        m.get("metric") == "xsd_invalid" and m.get("expected", 0) == m.get("actual", 0)
        for m in comparison.get("matched", [])
    )
    has_unexpected_xsd_error = has_xsd_error and not xsd_invalid_expected

    if has_gen_error:
        result["status"] = "ERROR"
    elif has_diffs or has_unexpected_xsd_error:
        result["status"] = "FAIL"
    elif has_regression:
        # Findings changed versus baseline — needs investigation
        result["status"] = "REGRESSION"
    elif has_naming or has_packaging:
        result["status"] = "WARNING"
    else:
        # All findings match baseline (including expected FAILs/WARNINGs)
        result["status"] = "PASS"

    # Cleanup temp dir
    shutil.rmtree(tmp_dir, ignore_errors=True)

    return result


# ── Evidence Record ──────────────────────────────────────────────────────────

def save_evidence(all_results: dict, output_dir: Path) -> Path:
    """Save regression evidence record."""
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    evidence_path = output_dir / f"regression_{ts}.yaml"

    # Check if any suite has enrichment data
    has_enrichment = any(
        data.get("enrichment") for data in all_results.values()
    )
    evidence = {
        "timestamp": datetime.now().isoformat(),
        "python_version": sys.version.split()[0],
        "mode": "E2E+enrichment" if has_enrichment else "E2E",
        "suites": {},
    }

    for name, data in all_results.items():
        suite_ev = {
            "status": data["status"],
        }
        if data.get("generation"):
            suite_ev["generation"] = data["generation"]
        if data.get("naming_issues"):
            suite_ev["naming_issues"] = data["naming_issues"]
        if data.get("packaging_issues"):
            suite_ev["packaging_issues"] = data["packaging_issues"]
        if data.get("xml_diffs"):
            # Only store first few diffs per file for brevity
            suite_ev["xml_diffs"] = {
                k: v[:5] for k, v in data["xml_diffs"].items()
            }
        if data.get("error"):
            suite_ev["error"] = data["error"]
        # Strip raw_output from validation for evidence
        if isinstance(data.get("validation"), dict):
            val = {}
            for k, v in data["validation"].items():
                if k == "raw_output":
                    continue
                if isinstance(v, dict):
                    val[k] = {kk: vv for kk, vv in v.items() if kk != "raw_output"}
                else:
                    val[k] = v
            if val:
                suite_ev["validation"] = val
        # Include baseline comparison result
        if data.get("baseline_comparison"):
            bc = data["baseline_comparison"]
            suite_ev["baseline_check"] = {
                "matched": len(bc.get("matched", [])),
                "regressions": len(bc.get("regressions", [])),
            }
            if bc.get("regressions"):
                suite_ev["baseline_regressions"] = bc["regressions"]
            if bc.get("note"):
                suite_ev["baseline_note"] = bc["note"]
        # Include enrichment summary
        if data.get("enrichment"):
            suite_ev["enrichment"] = data["enrichment"]
        evidence["suites"][name] = suite_ev

    with open(str(evidence_path), "w", encoding="utf-8") as f:
        yaml.dump(evidence, f, default_flow_style=False, sort_keys=False,
                  allow_unicode=True, width=120)

    return evidence_path


# ── Reference Data Freshness ─────────────────────────────────────────────────

def check_reference_data_freshness() -> dict:
    """Check reference data freshness. Returns staleness status."""
    freshness = {}
    _aifmd_xsd = _AIFMD_ANNEX_IV_DIR / "xsd"
    for label, xsd_dir in [("ESMA XSD", _aifmd_xsd / "esma_1.2" if (_aifmd_xsd / "esma_1.2").is_dir() else _APP_ROOT / "XSD" / "ESMA 1.2"),
                            ("FCA XSD", _aifmd_xsd / "fca_2.0" if (_aifmd_xsd / "fca_2.0").is_dir() else _APP_ROOT / "XSD" / "FCA 2.0")]:
        if xsd_dir.is_dir():
            xsd_files = list(xsd_dir.glob("*.xsd"))
            if xsd_files:
                newest = max(f.stat().st_mtime for f in xsd_files)
                age_days = (datetime.now().timestamp() - newest) / 86400
                status = "GREEN" if age_days < 90 else "AMBER" if age_days < 365 else "RED"
                freshness[label] = {
                    "age_days": round(age_days),
                    "status": status,
                    "last_modified": datetime.fromtimestamp(newest).isoformat()[:10],
                }

    if DEFAULT_RULES_PATH.exists():
        age_days = (datetime.now().timestamp() -
                    DEFAULT_RULES_PATH.stat().st_mtime) / 86400
        status = "GREEN" if age_days < 30 else "AMBER" if age_days < 90 else "RED"
        freshness["Validation Rules"] = {
            "age_days": round(age_days),
            "status": status,
            "last_modified": datetime.fromtimestamp(
                DEFAULT_RULES_PATH.stat().st_mtime).isoformat()[:10],
        }

    nca_files = sorted(Path(NCA_OVERRIDES_DIR).glob(
        "aifmd_nca_overrides_*.yaml"))
    if nca_files:
        oldest = max(f.stat().st_mtime for f in nca_files)
        age_days = (datetime.now().timestamp() - oldest) / 86400
        status = "GREEN" if age_days < 30 else "AMBER" if age_days < 90 else "RED"
        freshness["NCA Overrides"] = {
            "age_days": round(age_days),
            "status": status,
            "file_count": len(nca_files),
            "last_modified": datetime.fromtimestamp(oldest).isoformat()[:10],
        }

    return freshness


# ── Main ─────────────────────────────────────────────────────────────────────

def _capture_expected_findings(all_results: dict) -> Path:
    """Save current actual findings as the new expected baseline."""
    baseline = _load_expected_findings() if EXPECTED_FINDINGS_PATH.exists() else {}
    rationale_catalogue = baseline.get("rationale", {})

    new_baseline = {
        "schema_version": 2,
        "reviewed_by": baseline.get("reviewed_by"),
        "reviewed_at": datetime.now().isoformat()[:10],
        "next_review": None,
        "rationale": rationale_catalogue,
        "suites": {},
    }

    for name, data in all_results.items():
        val = data.get("validation", {})
        suite_config = SUITES.get(name, {})
        is_multi = suite_config.get("multi_nca", False)

        if is_multi and isinstance(val, dict):
            per_nca = {}
            for nca_key, nca_val in val.items():
                if isinstance(nca_val, dict) and "xsd_valid" in nca_val:
                    per_nca[nca_key] = {"expected": {
                        "xsd_valid": nca_val.get("xsd_valid", 0),
                        "xsd_invalid": nca_val.get("xsd_invalid", 0),
                        "dqf_fail": {"count": nca_val.get("dqf_fail", 0)},
                        "dqf_missing": {"count": nca_val.get("dqf_missing", 0)},
                        "dqf_warning": {"count": nca_val.get("dqf_warning", 0)},
                        "file_naming_fail": {"count": nca_val.get("file_naming_fail", 0)},
                    }}
            new_baseline["suites"][name] = {
                "description": suite_config.get("description", ""),
                "multi_nca": True,
                "per_nca": per_nca,
            }
        elif isinstance(val, dict):
            new_baseline["suites"][name] = {
                "description": suite_config.get("description", ""),
                "expected": {
                    "xsd_valid": val.get("xsd_valid", 0),
                    "xsd_invalid": val.get("xsd_invalid", 0),
                    "dqf_fail": {"count": val.get("dqf_fail", 0)},
                    "dqf_missing": {"count": val.get("dqf_missing", 0)},
                    "dqf_warning": {"count": val.get("dqf_warning", 0)},
                    "file_naming_fail": {"count": val.get("file_naming_fail", 0)},
                },
            }

    with open(str(EXPECTED_FINDINGS_PATH), "w", encoding="utf-8") as f:
        yaml.dump(new_baseline, f, default_flow_style=False, sort_keys=False,
                  allow_unicode=True, width=120)

    return EXPECTED_FINDINGS_PATH


def generate_compliance_report(all_results: dict, output_dir: Path) -> Path:
    """Generate an Excel compliance evidence report.

    The report contains:
      - Summary tab: suite-level pass/fail with baseline comparison
      - Findings tab: every expected finding with rationale and actual match status
      - Regressions tab: any unexpected deviations (should be empty for a clean run)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [!] openpyxl not installed — skipping compliance report")
        return None

    wb = Workbook()
    baseline = _load_expected_findings()

    # ── Styles ────────────────────────────────────────────────────────────
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="2B5797")
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    reg_fill = PatternFill("solid", fgColor="FF6B6B")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))
    wrap = Alignment(wrap_text=True, vertical="top")

    def _style_header(ws, cols):
        for c, (label, width) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=label)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = width

    # ── Summary tab ───────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Suite Summary"
    cols = [("Suite", 22), ("Status", 12), ("XSD Valid", 10), ("XSD Invalid", 10),
            ("DQF Pass", 10), ("DQF Fail", 10), ("DQF Missing", 10),
            ("DQF Warning", 10), ("Naming Fail", 10),
            ("Baseline Match", 14), ("Regressions", 12)]
    _style_header(ws_sum, cols)

    row = 2
    for name, data in all_results.items():
        bc = data.get("baseline_comparison", {})
        val = data.get("validation", {})
        status = data.get("status", "UNKNOWN")

        # Aggregate validation metrics for multi-NCA
        suite_config = SUITES.get(name, {})
        is_multi = suite_config.get("multi_nca", False)
        if is_multi:
            agg = {"xsd_valid": 0, "xsd_invalid": 0, "dqf_pass": 0,
                   "dqf_fail": 0, "dqf_missing": 0, "dqf_warning": 0,
                   "file_naming_fail": 0}
            for nca_key, nca_val in val.items():
                if isinstance(nca_val, dict):
                    for k in agg:
                        agg[k] += nca_val.get(k, 0)
            val = agg

        ws_sum.cell(row=row, column=1, value=name)
        status_cell = ws_sum.cell(row=row, column=2, value=status)
        fill_map = {"PASS": pass_fill, "FAIL": fail_fill, "WARNING": warn_fill,
                    "REGRESSION": reg_fill, "ERROR": fail_fill}
        status_cell.fill = fill_map.get(status, warn_fill)
        ws_sum.cell(row=row, column=3, value=val.get("xsd_valid", 0))
        ws_sum.cell(row=row, column=4, value=val.get("xsd_invalid", 0))
        ws_sum.cell(row=row, column=5, value=val.get("dqf_pass", 0))
        ws_sum.cell(row=row, column=6, value=val.get("dqf_fail", 0))
        ws_sum.cell(row=row, column=7, value=val.get("dqf_missing", 0))
        ws_sum.cell(row=row, column=8, value=val.get("dqf_warning", 0))
        ws_sum.cell(row=row, column=9, value=val.get("file_naming_fail", 0))
        ws_sum.cell(row=row, column=10, value=len(bc.get("matched", [])))
        reg_count = len(bc.get("regressions", []))
        reg_cell = ws_sum.cell(row=row, column=11, value=reg_count)
        if reg_count > 0:
            reg_cell.fill = reg_fill
            reg_cell.font = Font(name="Arial", bold=True, color="FFFFFF")

        for c in range(1, 12):
            ws_sum.cell(row=row, column=c).border = thin_border
        row += 1

    # ── Expected Findings tab ─────────────────────────────────────────────
    ws_exp = wb.create_sheet("Expected Findings")
    cols = [("Suite", 22), ("NCA", 6), ("Metric", 16), ("Expected", 10),
            ("Actual", 10), ("Match", 8), ("Rationale", 60)]
    _style_header(ws_exp, cols)

    row = 2
    for name, data in all_results.items():
        bc = data.get("baseline_comparison", {})
        for entry in bc.get("matched", []) + bc.get("regressions", []):
            ws_exp.cell(row=row, column=1, value=name)
            ws_exp.cell(row=row, column=2, value=entry.get("nca", ""))
            ws_exp.cell(row=row, column=3, value=entry["metric"])
            ws_exp.cell(row=row, column=4, value=entry["expected"])
            ws_exp.cell(row=row, column=5, value=entry["actual"])
            is_match = entry["expected"] == entry["actual"]
            match_cell = ws_exp.cell(row=row, column=6,
                                     value="YES" if is_match else "NO")
            match_cell.fill = pass_fill if is_match else fail_fill
            rat_cell = ws_exp.cell(row=row, column=7,
                                   value=entry.get("rationale", ""))
            rat_cell.alignment = wrap
            for c in range(1, 8):
                ws_exp.cell(row=row, column=c).border = thin_border
            row += 1

    # ── Regressions tab ──────────────────────────────────────────────────
    ws_reg = wb.create_sheet("Regressions")
    cols = [("Suite", 22), ("NCA", 6), ("Metric", 16), ("Expected", 10),
            ("Actual", 10), ("Direction", 10), ("Action Required", 40)]
    _style_header(ws_reg, cols)

    row = 2
    any_regression = False
    for name, data in all_results.items():
        bc = data.get("baseline_comparison", {})
        for entry in bc.get("regressions", []):
            any_regression = True
            ws_reg.cell(row=row, column=1, value=name)
            ws_reg.cell(row=row, column=2, value=entry.get("nca", ""))
            ws_reg.cell(row=row, column=3, value=entry["metric"])
            ws_reg.cell(row=row, column=4, value=entry["expected"])
            ws_reg.cell(row=row, column=5, value=entry["actual"])
            direction = entry.get("direction", "")
            dir_cell = ws_reg.cell(row=row, column=6, value=direction)
            dir_cell.fill = reg_fill if direction == "NEW" else pass_fill
            action = ("Investigate: new finding not in baseline"
                      if direction == "NEW" else
                      "Verify: expected finding resolved — update baseline if intentional")
            ws_reg.cell(row=row, column=7, value=action)
            for c in range(1, 8):
                ws_reg.cell(row=row, column=c).border = thin_border
            row += 1

    if not any_regression:
        ws_reg.cell(row=2, column=1, value="No regressions detected")
        ws_reg.cell(row=2, column=1).fill = pass_fill
        ws_reg.cell(row=2, column=1).font = Font(name="Arial", bold=True)

    # ── Baseline Info tab ─────────────────────────────────────────────────
    ws_info = wb.create_sheet("Baseline Info")
    cols = [("Field", 22), ("Value", 50)]
    _style_header(ws_info, cols)
    info_rows = [
        ("Schema Version", baseline.get("schema_version", "")),
        ("Reviewed By", baseline.get("reviewed_by", "NOT YET REVIEWED")),
        ("Reviewed At", baseline.get("reviewed_at", "")),
        ("Next Review", baseline.get("next_review", "")),
        ("Report Generated", datetime.now().isoformat()),
        ("Baseline File", str(EXPECTED_FINDINGS_PATH)),
    ]
    # Add rationale catalogue
    for key, text in baseline.get("rationale", {}).items():
        info_rows.append((f"Rationale: {key}", str(text).strip()))

    for i, (field, value) in enumerate(info_rows, 2):
        ws_info.cell(row=i, column=1, value=field).font = Font(name="Arial", bold=True)
        val_cell = ws_info.cell(row=i, column=2, value=str(value))
        val_cell.alignment = wrap
        for c in range(1, 3):
            ws_info.cell(row=i, column=c).border = thin_border

    # Save
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = output_dir / f"compliance_{ts}.xlsx"
    wb.save(str(report_path))
    return report_path


def main():
    parser = argparse.ArgumentParser(
        description="AIFMD Golden Set Regression Suite — True End-to-End")
    parser.add_argument("--suite", default="",
                        help="Run only this suite (default: all)")
    parser.add_argument("--scope", default="all",
                        choices=["all", "realdata", "synthetic"],
                        help="Run only realdata (M examples + authorised_anon) "
                             "or synthetic suites (default: all)")
    parser.add_argument("--update-baseline", action="store_true",
                        help="Accept current results as new expected baseline")
    parser.add_argument("--capture-expected", action="store_true",
                        help="Save current findings as new expected_findings.yaml")
    parser.add_argument("--compliance-report", action="store_true",
                        help="Generate Excel compliance evidence report")
    parser.add_argument("--freshness", action="store_true",
                        help="Show reference data freshness report")
    parser.add_argument("--with-enrichment", action="store_true",
                        help="Run with LEI enrichment from gleif_seed.yaml")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Show detailed diff output")
    args = parser.parse_args()

    # Reference data freshness check
    if args.freshness:
        freshness = check_reference_data_freshness()
        print(f"\n{'='*70}")
        print(f"  Reference Data Freshness Report")
        print(f"{'='*70}\n")
        for source, info in freshness.items():
            icon = {"GREEN": "G", "AMBER": "A", "RED": "R"}.get(info["status"], "?")
            print(f"  [{icon}] {source}: {info['age_days']} days old "
                  f"({info['status']}) — last modified: {info['last_modified']}")
        print()
        return

    print(f"\n{'='*70}")
    print(f"  AIFMD Golden Set Regression Suite — E2E Mode")
    print(f"{'='*70}")
    print(f"  Timestamp: {datetime.now().isoformat()}")
    print(f"  Golden set: {_GOLDEN_SET_DIR}")

    # Determine which suites to run
    suites_to_run = {}
    if args.suite:
        if args.suite in SUITES:
            suites_to_run[args.suite] = SUITES[args.suite]
        else:
            print(f"\n  ERROR: Suite '{args.suite}' not found. Available: "
                  f"{', '.join(SUITES.keys())}")
            sys.exit(1)
    elif args.scope != "all":
        suites_to_run = {
            k: v for k, v in SUITES.items()
            if v.get("category") == args.scope
        }
        print(f"  Scope: {args.scope} ({len(suites_to_run)} suites)")
    else:
        suites_to_run = SUITES

    # Reference data freshness summary
    freshness = check_reference_data_freshness()
    print(f"\n  Reference Data Freshness:")
    for source, info in freshness.items():
        icon = {"GREEN": "G", "AMBER": "A", "RED": "R"}.get(info["status"], "?")
        print(f"    [{icon}] {source}: {info['age_days']}d ({info['status']})")

    # Load GLEIF seed store for enrichment (if requested)
    gleif_store = None
    if args.with_enrichment:
        gleif_store = _load_gleif_seed_store()
        if gleif_store:
            count = gleif_store.get_lei_count()
            print(f"\n  LEI Enrichment: ENABLED ({count} seed records from gleif_seed.yaml)")
        else:
            print(f"\n  LEI Enrichment: FAILED (could not load seed store)")

    # Run suites
    all_results = {}
    total_pass = 0
    total_fail = 0
    total_warn = 0
    total_error = 0
    total_skip = 0

    for suite_name, suite_config in suites_to_run.items():
        suite_dir = _GOLDEN_SET_DIR / suite_config["directory"]

        print(f"\n{'_'*60}")
        print(f"  Suite: {suite_name}")
        print(f"  {suite_config['description']}")

        if not suite_dir.is_dir():
            print(f"  [SKIP] Directory not found")
            all_results[suite_name] = {"status": "SKIPPED", "error": "Directory not found"}
            total_skip += 1
            continue

        # Count templates
        templates = _find_templates(suite_dir)
        print(f"  Templates: {len(templates)} file(s)")

        # Run E2E (with optional enrichment)
        result = run_e2e_suite(suite_name, suite_config,
                               reference_store=gleif_store)
        all_results[suite_name] = result

        # Display result
        status = result["status"]
        gen = result.get("generation", {})
        status_icon = {
            "PASS": "[PASS]", "FAIL": "[FAIL]", "WARNING": "[WARN]",
            "ERROR": "[ERR ]", "SKIPPED": "[SKIP]", "REGRESSION": "[REGR]",
        }.get(status, "[????]")

        print(f"  {status_icon} {status}")

        if gen:
            print(f"     Generated: {gen.get('aifm_xmls', 0)} AIFM + "
                  f"{gen.get('aif_xmls', 0)} AIF XMLs, "
                  f"{gen.get('packages', 0)} packages")

        if result.get("xml_diffs"):
            print(f"     XML diffs: {len(result['xml_diffs'])} file(s)")
            if args.verbose:
                for fname, diffs in result["xml_diffs"].items():
                    print(f"       {fname}:")
                    for d in diffs[:5]:
                        print(f"         {d}")

        if result.get("naming_issues"):
            print(f"     Naming issues: {len(result['naming_issues'])}")
            for issue in result["naming_issues"][:3]:
                print(f"       {issue}")

        if result.get("packaging_issues"):
            print(f"     Packaging issues: {len(result['packaging_issues'])}")
            for issue in result["packaging_issues"][:3]:
                print(f"       {issue}")

        if result.get("error"):
            print(f"     Error: {result['error'][:200]}")

        # Validation summary
        val = result.get("validation", {})
        if isinstance(val, dict) and not any(isinstance(v, dict) for v in val.values()):
            # Single NCA validation
            if "xsd_valid" in val:
                print(f"     Validation: XSD {val['xsd_valid']} valid, "
                      f"{val.get('xsd_invalid', 0)} invalid | "
                      f"DQF {val.get('dqf_pass', 0)} pass, "
                      f"{val.get('dqf_fail', 0)} fail")
                # Show diagnostic if validation produced no results
                if val.get("xsd_valid", 0) == 0 and val.get("xsd_invalid", 0) == 0:
                    raw = val.get("raw_output", "")
                    if raw:
                        print(f"     [DIAG] Validator raw output (first 2000 chars):")
                        for diag_line in raw[:2000].split("\n"):
                            if diag_line.strip():
                                print(f"       | {diag_line.rstrip()}")
        elif isinstance(val, dict):
            # Multi-NCA validation
            for nca_key, nca_val in val.items():
                if isinstance(nca_val, dict) and "xsd_valid" in nca_val:
                    print(f"     Validation [{nca_key}]: XSD {nca_val['xsd_valid']} valid, "
                          f"{nca_val.get('xsd_invalid', 0)} invalid | "
                          f"DQF {nca_val.get('dqf_pass', 0)} pass, "
                          f"{nca_val.get('dqf_fail', 0)} fail")

        # Enrichment summary
        enr = result.get("enrichment")
        if enr:
            parts = []
            if enr.get("enriched"):
                parts.append(f"{enr['enriched']} enriched")
            if enr.get("pending_user_choice"):
                parts.append(f"{enr['pending_user_choice']} pending")
            if enr.get("no_match"):
                parts.append(f"{enr['no_match']} no-match")
            if enr.get("skipped"):
                parts.append(f"{enr['skipped']} skipped")
            if enr.get("error"):
                parts.append(f"{enr['error']} error")
            print(f"     Enrichment: {enr.get('total', 0)} fields — "
                  f"{', '.join(parts)}")

        # Baseline comparison summary
        bc = result.get("baseline_comparison", {})
        matched = len(bc.get("matched", []))
        regressions = bc.get("regressions", [])
        if matched or regressions:
            print(f"     Baseline: {matched} matched, {len(regressions)} regressions")
            for reg in regressions[:5]:
                nca_prefix = f"[{reg['nca']}] " if reg.get("nca") else ""
                print(f"       {nca_prefix}{reg['metric']}: "
                      f"expected={reg['expected']} actual={reg['actual']} "
                      f"({reg.get('direction', '')})")

        # Tally
        if status == "PASS":
            total_pass += 1
        elif status == "FAIL":
            total_fail += 1
        elif status == "REGRESSION":
            total_fail += 1   # regressions count as failures
        elif status == "WARNING":
            total_warn += 1
        elif status == "ERROR":
            total_error += 1
        else:
            total_skip += 1

    # Summary
    total = len(suites_to_run)
    print(f"\n{'='*70}")
    print(f"  E2E REGRESSION SUITE SUMMARY")
    print(f"{'='*70}")
    print(f"  Total suites: {total}")
    print(f"  [PASS] {total_pass}")
    if total_warn:
        print(f"  [WARN] {total_warn}")
    if total_fail:
        print(f"  [FAIL] {total_fail}")
    if total_error:
        print(f"  [ERR ] {total_error}")
    if total_skip:
        print(f"  [SKIP] {total_skip}")

    overall = "PASS" if (total_fail == 0 and total_error == 0) else "FAIL"
    print(f"\n  Overall: {'ALL PASS' if overall == 'PASS' else 'FAILURES DETECTED'}")

    # Save evidence
    evidence_dir = _GOLDEN_SET_DIR / "evidence"
    evidence_path = save_evidence(all_results, evidence_dir)
    print(f"  Evidence: {evidence_path.name}")

    # Capture expected findings if requested
    if args.capture_expected:
        ef_path = _capture_expected_findings(all_results)
        print(f"  Expected findings baseline saved: {ef_path.name}")
        print(f"  NOTE: Set reviewed_by and next_review in the file before approval.")

    # Generate compliance report if requested
    if args.compliance_report:
        report_path = generate_compliance_report(all_results, evidence_dir)
        if report_path:
            print(f"  Compliance report: {report_path.name}")

    print(f"{'='*70}\n")

    sys.exit(0 if overall == "PASS" else 1)


if __name__ == "__main__":
    main()
